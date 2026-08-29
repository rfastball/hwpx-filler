"""Excel/CSV 행 성형 계약(#183)."""
from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path

import pytest
from openpyxl import Workbook

from hwpxfiller.data.excel import ExcelDataSource, _cell_text


def _xlsx(path: Path, rows: list[list[object]]) -> Path:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()
    return path


@pytest.mark.parametrize("suffix", [".csv", ".xlsx"])
def test_custom_header_row_blank_and_ragged_rows_share_policy(
    tmp_path: Path, suffix: str
) -> None:
    path = tmp_path / f"rows{suffix}"
    if suffix == ".csv":
        path.write_text(
            "설명 줄\n공고명,금액,담당자\n전산장비,1000\n,,\n비품,2000,김담당\n",
            encoding="utf-8-sig",
        )
    else:
        _xlsx(
            path,
            [
                ["설명 줄"],
                ["공고명", "금액", "담당자"],
                ["전산장비", 1000],
                [None, None, None],
                ["비품", 2000, "김담당"],
            ],
        )

    source = ExcelDataSource(str(path), header_row=2)

    assert source.fields() == ["공고명", "금액", "담당자"]
    assert source.records() == [
        {"공고명": "전산장비", "금액": "1000", "담당자": ""},
        {"공고명": "비품", "금액": "2000", "담당자": "김담당"},
    ]


def test_csv_bom_and_unicode_headers_are_preserved(tmp_path: Path) -> None:
    path = tmp_path / "unicode.csv"
    path.write_text("공고명,예산￦,담당자𠮷\n장비,10,김\n", encoding="utf-8-sig")

    source = ExcelDataSource(str(path))

    assert source.fields() == ["공고명", "예산￦", "담당자𠮷"]
    assert source.records()[0]["담당자𠮷"] == "김"


@pytest.mark.parametrize(
    ("headers", "message"),
    [(["공고명", "", "담당자"], "빈 헤더"), (["공고명", "공고명"], "중복 헤더")],
)
@pytest.mark.parametrize("suffix", [".csv", ".xlsx"])
def test_blank_and_duplicate_headers_fail_loudly(
    tmp_path: Path, suffix: str, headers: list[str], message: str
) -> None:
    path = tmp_path / f"bad{suffix}"
    if suffix == ".csv":
        path.write_text(",".join(headers) + "\n값,값,값\n", encoding="utf-8")
    else:
        _xlsx(path, [headers, ["값"] * len(headers)])

    with pytest.raises(ValueError, match=message):
        ExcelDataSource(str(path)).fields()


@pytest.mark.parametrize("suffix", [".csv", ".xlsx"])
def test_nonblank_cells_beyond_header_fail_loudly(tmp_path: Path, suffix: str) -> None:
    path = tmp_path / f"wide{suffix}"
    if suffix == ".csv":
        path.write_text("공고명\n장비,숨은 값\n", encoding="utf-8")
    else:
        _xlsx(path, [["공고명"], ["장비", "숨은 값"]])

    # CSV는 ragged overflow로, XLSX는 worksheet max_column이 헤더까지 넓어져 빈 헤더로
    # 관측된다. 어느 쪽도 이름 없는 값을 조용히 버리지 않는다.
    with pytest.raises(ValueError, match="빈 헤더|헤더보다 값이 많은 행"):
        ExcelDataSource(str(path)).records()


def test_xlsx_scalar_conversion_is_deterministic(tmp_path: Path) -> None:
    # openpyxl은 날짜 셀을 datetime으로 읽는다. 공용 scalar 정책의 순수 date/time
    # 입력은 어댑터 단위에서 직접 고정한다.
    assert _cell_text(date(2026, 7, 22)) == "2026-07-22"
    assert _cell_text(time(9, 5, 6)) == "09:05:06"

    path = _xlsx(
        tmp_path / "scalars.xlsx",
        [
            ["날짜", "시각", "정수", "실수", "참거짓"],
            [date(2026, 7, 22), datetime(2026, 7, 22, 9, 5, 6), 1000, 1.25, True],
        ],
    )

    # 날짜 셀은 openpyxl 이 자정 datetime 으로 돌려주지만 표시 서식에 시각이 없다 —
    # 사용자가 쓴 적 없는 00:00:00 을 지어내지 않는다(U4 계열1-25).
    assert ExcelDataSource(str(path)).records() == [
        {
            "날짜": "2026-07-22",
            "시각": "2026-07-22 09:05:06",
            "정수": "1000",
            "실수": "1.25",
            "참거짓": "TRUE",
        }
    ]


def test_midnight_datetime_keeps_its_time_when_the_format_shows_one(tmp_path: Path) -> None:
    """자정 억제는 **서식에 시각이 없을 때만**이다 — 실제 자정 값을 잃지 않는다."""
    path = _xlsx(tmp_path / "midnight.xlsx", [["마감"], [datetime(2026, 7, 22, 0, 0)]])
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    ws.cell(row=2, column=1).number_format = "yyyy-mm-dd hh:mm"
    wb.save(path)
    wb.close()

    assert ExcelDataSource(str(path)).records() == [{"마감": "2026-07-22 00:00:00"}]


def test_date_only_cell_does_not_leak_a_fabricated_time_into_the_display_form(
    tmp_path: Path,
) -> None:
    """읽기 경계의 자정 억제가 표시형까지 관통하는가 — 결함 25 의 재현 회귀."""
    from hwpxfiller.domain.mapping import apply_transform

    path = _xlsx(tmp_path / "dates.xlsx", [["계약일"], [date(2026, 7, 22)]])
    value = ExcelDataSource(str(path)).records()[0]["계약일"]

    assert value == "2026-07-22"
    assert apply_transform("date", value) == "2026. 7. 22."


def test_xlsx_formula_without_cached_value_fails_loudly(tmp_path: Path) -> None:
    path = _xlsx(tmp_path / "formula.xlsx", [["합계"], ["=1+2"]])

    with pytest.raises(ValueError, match="계산값이 저장돼 있지 않습니다") as caught:
        ExcelDataSource(str(path)).records()

    message = str(caught.value)
    assert "행 2 열 1" in message
    # 조치 없는 통보는 사용자를 막다른 길에 세운다(COPY_STYLE_GUIDE §2 오류=①문제 ②조치).
    assert "엑셀에서 이 파일을 열어 다시 저장하면" in message
    assert "다시 불러오세요" in message
    assert "cache" not in message


def test_xlsx_missing_formula_values_are_reported_in_one_pass(tmp_path: Path) -> None:
    # 첫 위반에서 끊으면 사용자는 한 셀씩 고쳐 다시 여는 왕복을 반복한다.
    path = _xlsx(
        tmp_path / "many.xlsx",
        [["합계", "잔액"], ["=1+2", 10], [20, "=A3*2"], ["=SUM(A2:A3)", "=B2+B3"]],
    )

    with pytest.raises(ValueError) as caught:
        ExcelDataSource(str(path)).records()

    message = str(caught.value)
    for coordinate in ("행 2 열 1", "행 3 열 2", "행 4 열 1", "행 4 열 2"):
        assert coordinate in message
    assert "외 " not in message  # 상한(5) 이내면 요약하지 않는다


def test_xlsx_missing_formula_values_beyond_limit_are_summarized(tmp_path: Path) -> None:
    path = _xlsx(
        tmp_path / "flood.xlsx",
        [["합계"], *[[f"=1+{n}"] for n in range(8)]],
    )

    with pytest.raises(ValueError) as caught:
        ExcelDataSource(str(path)).records()

    message = str(caught.value)
    assert "행 2 열 1" in message
    assert "행 6 열 1" in message  # 상한 5번째
    assert "행 7 열 1" not in message  # 초과분은 나열하지 않는다
    assert "외 3개" in message


def test_csv_and_xlsx_same_logical_data_have_parity(tmp_path: Path) -> None:
    csv_path = tmp_path / "same.csv"
    csv_path.write_text(
        "공고명,금액,완료\n전산장비,1000,TRUE\n비품,1.25,FALSE\n",
        encoding="utf-8-sig",
    )
    xlsx_path = _xlsx(
        tmp_path / "same.xlsx",
        [["공고명", "금액", "완료"], ["전산장비", 1000, True], ["비품", 1.25, False]],
    )

    csv_source = ExcelDataSource(str(csv_path))
    xlsx_source = ExcelDataSource(str(xlsx_path))

    assert csv_source.fields() == xlsx_source.fields()
    assert csv_source.records() == xlsx_source.records()


def test_minimal_xlsm_is_read_through_supported_adapter(tmp_path: Path) -> None:
    path = _xlsx(tmp_path / "minimal.xlsm", [["공고명", "금액"], ["전산장비", 1000]])

    source = ExcelDataSource(str(path))

    assert source.fields() == ["공고명", "금액"]
    assert source.records() == [{"공고명": "전산장비", "금액": "1000"}]


@pytest.mark.parametrize("header_row", [0, -1, True, 1.5])
def test_header_row_must_be_positive_integer(header_row: object) -> None:
    with pytest.raises(ValueError, match="1 이상의 정수"):
        ExcelDataSource("unused.csv", header_row=header_row)  # type: ignore[arg-type]
