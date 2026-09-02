"""작업 에디터 화면 컨트롤러 계약 가드 — pywebview/Qt 불필요(헤드리스).

에픽 #20 화면 #15·#16 이관의 회귀 심. 3단계 마법사 게이트(스키마·PARTIAL·매핑 확정·저장)를
링1 VM 그대로 구동해 창 없이 확인한다(R-flow 슬라이스 5 블록 2 — 데이터 선택이 매핑 단계
관문으로 접힘: 템플릿 0 → 매핑 1 → 저장 2). 실 HWPX 픽스처(COMPILED·PARTIAL)로 게이트 분기를 탄다.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest

from _web_source import REPO_ROOT, SOURCE_JS_DIR
from hwpxfiller.external.job_store import JobRegistry, encode_job
from hwpxfiller.external.text_registry import TextTemplateRegistry
from hwpxfiller.external.template_files import TemplateFileStore
from hwpxfiller.external.template_inspection import (
    HWPX_TEMPLATE_OPS,
    inspect_hwpx_template,
)
from hwpxfiller.gui.template_manager_state import TemplateManagerViewModel
from hwpxfiller.webapp.screen_editor import EditorController
from hwpxfiller.webapp.template_groups import TemplateGroupModel

REPO = REPO_ROOT
TPL_COMPILED = REPO / "tests" / "corpus" / "scenario" / "templates" / "구매요청서.hwpx"
TPL_PARTIAL = REPO / "tests" / "fixtures" / "template_v1.hwpx"
MULTI_SHEET = REPO / "tests" / "fixtures" / "multi_sheet.xlsx"
_NOW = datetime(2026, 8, 11, 12, 34, 56)


def _clock() -> datetime:
    return _NOW


def _controller(
    tmp_path: Path, *, after_mapping_saved=None, binding_confirm_pending=None
) -> "tuple[EditorController, list]":
    pushes: list = []
    reg = JobRegistry(tmp_path / "jobs")
    # 빈 라이브러리 VM·격리 TXT 레지스트리 주입 — 기본(표준 라이브러리 지연 생성)이 실
    # 사용자 폴더를 스캔하면 테스트가 개발 머신 상태에 좌우된다(PR-4 리뷰 F5: 격리·결정성).
    ctrl = EditorController(
        reg, lambda s, snap: pushes.append((s, snap)),
        clock=_clock,
        template_library=TemplateManagerViewModel(
            paths=[],
            inspect_template=inspect_hwpx_template,
            file_ops=HWPX_TEMPLATE_OPS,
        ),
        text_registry=TextTemplateRegistry(tmp_path / "text_templates"),
        after_mapping_saved=after_mapping_saved,
        binding_confirm_pending=binding_confirm_pending,
    )
    return ctrl, pushes


def _mount_data(ctrl: EditorController) -> None:
    """이 세션에 데이터를 연결한다 — 저장 게이트가 결속을 요구한다(#932 U4-C S2-3).

    낙찰현황 시트의 열 셋(업체명·낙찰금액·계약일)은 COMPILED 템플릿의 필드와 하나도 겹치지
    않아 자동 제안이 서지 않는다. 그래서 이 마운트는 매핑 판정을 바꾸지 않고 **결속만**
    세운다 — 구 「데이터 없이 진행」 세션의 상수·비움 흐름이 그대로 산다.
    """
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")


def _mount_other_data(ctrl: EditorController) -> None:
    """1단계 게이트를 여는 **다른** 결속(공고목록 시트) — 뒤이어 낙찰현황으로 갈아탄다.

    U6-B(#976) 이후 초안은 데이터 없이 「연결 확인」으로 갈 수 없다(:meth:`can_advance`).
    데이터 **교체**가 매핑 모델을 어떻게 다시 세우는지 재는 시험들은 그래서 첫 결속을
    따로 세운 뒤 갈아탄다 — 같은 시트를 두 번 마운트하면 정체 키가 안 움직여 재생성
    자체가 일어나지 않는다(`_model_key_now`).
    """
    ctrl.load_data_path(str(MULTI_SHEET))


def _txt_template(tmp_path: Path, name: str = "기안", body: "str | None" = None) -> Path:
    """격리 TXT 템플릿 픽스처 — `_controller` 가 주입하는 레지스트리 루트에 쓴다."""
    root = tmp_path / "text_templates"
    root.mkdir(parents=True, exist_ok=True)
    p = root / f"{name}.txt"
    p.write_text(
        body if body is not None else "건명: {{건명}}\n금액: {{금액}}\n담당: {{건명}}",
        encoding="utf-8",
    )
    return p


def test_compiled_template_opens_advance_gate(tmp_path):
    ctrl, pushes = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    snap = pushes[-1][1]
    assert snap["field_count"] == 10
    assert snap["gate"] is None and not snap["raw_block"]
    # U6-B(#976): 1단계가 묻는 질문은 「어느 템플릿을 어느 데이터에?」 하나라 템플릿만으로는
    # 열리지 않는다. 사유도 Python 이 낸다 — 고칠 자리가 좌·우로 갈리므로 한 문장에 합치지
    # 않는다(저장 게이트가 이미 요구하던 것을 같은 순서로 앞당겨 세운 것이다).
    assert ctrl.can_advance("template") is False
    assert snap["pairing"]["advance_block_reason"] == "오른쪽에서 데이터를 고르세요."
    _mount_data(ctrl)
    assert ctrl.can_advance("template") is True
    assert ctrl.snapshot()["pairing"]["advance_block_reason"] == ""


def test_snapshot_exposes_structured_fields(tmp_path):
    """1단계 구조화 표(#16 98DDFE96) — 스냅샷이 필드별 명세를 실어야 한다.

    나열식 요약(schema_summary)은 헤더로 존치하되, 표 렌더가 소비할 fields 배열이
    필드 수만큼·정해진 키로 있어야 한다. 템플릿 로드 전엔 빈 배열.
    """
    ctrl, pushes = _controller(tmp_path)
    assert ctrl.snapshot()["fields"] == []  # 스키마 없으면 빈 배열
    ctrl.load_template_path(str(TPL_COMPILED))
    snap = pushes[-1][1]
    fields = snap["fields"]
    assert isinstance(fields, list) and len(fields) == snap["field_count"]
    assert snap["schema_summary"]  # 헤더 요약은 존치
    for f in fields:
        assert set(f) >= {"name", "inferred_type", "in_table", "occurrences", "context"}
        assert isinstance(f["name"], str) and f["name"]
        assert isinstance(f["in_table"], bool)


def test_snapshot_exposes_sample_rows_projected_and_capped(tmp_path):
    """2단계 데이터 미리보기(#16) — 스냅샷이 source_fields 순서로 투영한 샘플 행을 싣는다.

    데이터 로드 전엔 빈 배열, 로드 후엔 record_count 를 넘지 않는 소량(≤_SAMPLE_ROWS)의
    문자열 셀 행. 각 행 폭은 컬럼 수와 일치(투영 정합).
    """
    from hwpxfiller.webapp.screen_editor import _SAMPLE_ROWS

    ctrl, pushes = _controller(tmp_path)
    assert ctrl.snapshot()["sample_rows"] == []  # 데이터 없으면 빈 배열
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    snap = pushes[-1][1]
    cols = snap["source_fields"]
    sample = snap["sample_rows"]
    assert 0 < len(sample) <= min(snap["record_count"], _SAMPLE_ROWS)
    for row in sample:
        assert len(row) == len(cols)  # source_fields 순서로 정확히 투영
        assert all(isinstance(c, str) for c in row)  # 렌더 esc 안전 위해 문자열


def test_partial_template_blocks_until_acked(tmp_path):
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_PARTIAL))
    assert ctrl.can_advance("template") is False  # PARTIAL → 게이트 닫힘
    gate = ctrl.snapshot()["gate"]
    assert gate and gate["unmet"] and not gate["acked"]
    # 게이트 미통과 상태에서 전진 요청은 시끄럽게 거부(confirm-or-alarm).
    with pytest.raises(ValueError, match="조건을 아직 채우지 못해"):
        ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("ack_gate", {})
    assert ctrl.can_advance("template") is False       # 게이트는 열렸고 데이터가 남았다
    _mount_data(ctrl)
    assert ctrl.can_advance("template") is True
    assert ctrl.snapshot()["gate"]["acked"] is True


def test_load_data_honors_confirmed_sheet(tmp_path):
    """다중 시트 확정 게이트(#33) — load_data_path(sheet=) 가 확정 시트를 관통 로드.

    첫 시트(공고목록)가 아닌 낙찰현황을 확정하면 그 시트의 필드·레코드가 온다 —
    조용한 첫 시트 강등이 아니라 확정값이 반영됨을 못박는다.
    """
    ctrl, pushes = _controller(tmp_path)
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    snap = pushes[-1][1]
    assert snap["source_fields"] == ["업체명", "낙찰금액", "계약일"]
    assert snap["record_count"] == 3
    # 대조군: 시트 미지정(None)은 첫 시트(공고목록) — 브리지가 모호할 때만 확정을 요구하므로
    # 컨트롤러 계약 자체는 None=첫/유일 시트로 유지된다.
    ctrl2, pushes2 = _controller(tmp_path)
    ctrl2.load_data_path(str(MULTI_SHEET))
    assert pushes2[-1][1]["source_fields"] == ["공고명", "추정가격"]


def test_full_new_job_flow_schema_only_const(tmp_path):
    """템플릿→매핑(관문 데이터 없이 진행, 상수 1행+비움 확정)→저장 end-to-end."""
    ctrl, pushes = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    # 데이터 연결은 저장 게이트(#932 U4-C S2-3)이자 **1단계 전진 게이트**다(U6-B #976) —
    # 상수만 쓰는 작업도 예외가 아니다. 그래서 고르기 단계에서 짝이 먼저 선다.
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})   # 매핑 진입(모델 초안 생성)
    snap = ctrl.snapshot()
    assert snap["section"] == "binding" and len(snap["rows"]) == 10
    assert snap["schema_only"] is False

    # 0행에 고정값 부여(내용 생성).
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "테스트값"})
    assert ctrl.snapshot()["rows"][0]["has_content"] is True

    # 모두 확정 → 내용 행 즉시 확정, 나머지는 비움 승격 후보로 반환(이름게이트).
    result = ctrl.dispatch("confirm_all", {})
    assert len(result["blanks"]) == 9
    assert ctrl.snapshot()["is_complete"] is False  # 비움 미확정
    ctrl.dispatch("confirm_blanks", {"fields": result["blanks"]})
    assert ctrl.snapshot()["is_complete"] is True

    # 저장.
    ctrl.dispatch("goto_section", {"section": "filename"})
    ctrl.dispatch("set_name", {"name": "테스트작업"})
    ctrl.dispatch("set_pattern", {"pattern": "문서-{{수요기관}}"})
    res = ctrl.dispatch("save", {})
    assert res["ok"] is True and res["saved_name"] == "테스트작업"
    assert JobRegistry(tmp_path / "jobs").exists("테스트작업")


def test_full_new_job_flow_today_system_token(tmp_path):
    """U4-E1 #939 — 「오늘 날짜」 유형 end-to-end: 선택 → 서식 → 미리보기 → 저장 왕복.

    const 흐름과 동형이되 **고정값 입력이 없다**는 것이 이 유형의 요점이다: 데이터 열도
    사람이 친 리터럴도 없이 행이 값을 낸다. 미리보기는 컨트롤러 clock(``_NOW``) 기준이라
    실 시각에 흔들리지 않는다.
    """
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    # 유형 목록은 링0 TYPES 를 그대로 싣는다 — 프런트가 자기 목록을 발명하지 않는다.
    assert "today" in ctrl.snapshot()["type_options"]
    # 서식 목록은 date 어휘를 공유한다(판정 1) — 없으면 프런트 서식 셀이 통째 비활성.
    fmt = ctrl.snapshot()["fmt_options"]
    assert fmt["today"] == fmt["date"] and fmt["today"]

    ctrl.dispatch("set_type", {"index": 0, "type": "today"})
    row = ctrl.snapshot()["rows"][0]
    # 소스도 상수도 없는데 내용이 있다 — 이 한 줄이 blank 강등(값 소실)의 회귀 심이다.
    assert row["source"] == "" and row["const"] == ""
    assert row["has_content"] is True
    assert row["preview"] == "2026. 8. 11. 12:34"      # clock 기준 기본 서식

    ctrl.dispatch("set_fmt", {"index": 0, "fmt": "%Y-%m-%d"})
    assert ctrl.snapshot()["rows"][0]["preview"] == "2026-08-11"

    result = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": result["blanks"]})
    ctrl.dispatch("goto_section", {"section": "filename"})
    ctrl.dispatch("set_name", {"name": "오늘작업"})
    ctrl.dispatch("set_pattern", {"pattern": "문서-{{수요기관}}"})
    assert ctrl.dispatch("save", {})["ok"] is True

    # 저장 왕복 — 유형·표시형이 durable 로 살아 돌아온다(blank 강등 없음).
    saved = JobRegistry(tmp_path / "jobs").load("오늘작업")
    today_rows = [m for m in saved.mapping.mappings if m.type == "today"]
    assert len(today_rows) == 1
    assert (today_rows[0].fmt, today_rows[0].source, today_rows[0].const) == (
        "%Y-%m-%d", "", "",
    )
    assert today_rows[0].template_field in saved.mapping.template_fields()
    assert saved.mapping.apply({}, now=_NOW)[today_rows[0].template_field] == "2026-08-11"


def test_unconfirm_all_restores_exact_previous_confirmed_set(tmp_path):
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl)                                   # 1단계 게이트(U6-B)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_confirmed", {"index": 1, "confirmed": True})
    ctrl.dispatch("set_confirmed", {"index": 4, "confirmed": True})
    result = ctrl.dispatch("unconfirm_all", {})
    assert result == {"undo_count": 2}
    assert ctrl.snapshot()["unconfirm_undo_count"] == 2
    restored = ctrl.dispatch("restore_confirmed", {})
    assert restored == {"restored": 2}
    rows = ctrl.snapshot()["rows"]
    assert [i for i, row in enumerate(rows) if row["confirmed"]] == [1, 4]
    assert ctrl.snapshot()["unconfirm_undo_count"] == 0


def test_unconfirm_undo_slot_dies_with_model_rebuild(tmp_path):
    """#273 리뷰 — 「모두 해제」 undo 슬롯은 **이전 모델의** 숫자 인덱스라, 템플릿/데이터
    교체로 모델이 재생성되면 소멸해야 한다. 살아남으면 아직 보이는 「되돌리기」가 새 입력의
    행들을 검토 없이 확정해 '키 변경 시 전원 미확정' 불변식을 우회한다(조용한 게이트 우회)."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_other_data(ctrl)                              # 1단계 게이트(U6-B) — 갈아탈 첫 결속
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_confirmed", {"index": 1, "confirmed": True})
    assert ctrl.dispatch("unconfirm_all", {}) == {"undo_count": 1}
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")  # 키 변경 → 모델 재생성
    assert ctrl.snapshot()["unconfirm_undo_count"] == 0      # 슬롯 소멸(버튼 근거 사라짐)
    assert ctrl.dispatch("restore_confirmed", {}) == {"restored": 0}
    assert all(row["confirmed"] is False for row in ctrl.snapshot()["rows"])


def test_new_draft_with_data_anchors_the_mounted_data_in_the_same_wizard(tmp_path):
    """U2 §2.4(#349) — 「이 데이터로 새 작업」은 **기존 마법사**에 데이터만 미리 세운다.

    확인할 것 셋: ①새 마법사를 짓지 않았다(단계·초안 성질 그대로, 1단계=템플릿에서 시작)
    ②2단계 관문이 그리는 앵커(`data_name`·`data_sheet`·헤더)가 이미 서 있다 ③진입 문맥이
    배너의 원천으로 살아 있다(사유·증거·복귀처).
    """
    ctrl, pushes = _controller(tmp_path)
    ctrl.new_draft_with_data(
        {"path": str(MULTI_SHEET), "sheet": "낙찰현황", "header_row": 0},
        entry_reason="document_browser_new_work",
        evidence={"데이터": "multi_sheet.xlsx"},
        return_context={"surface": "data"},
    )
    snap = pushes[-1][1]
    assert snap["is_draft"] is True and snap["editing_origin"] == ""
    assert snap["section"] == "template"          # 마법사는 1단계부터 — 순서 의존은 그대로
    assert snap["template_path"] == "" and snap["name"] == ""
    # 2단계 관문 앵커(dataGateway 가 그리는 값) — 「이 데이터로」가 실제로 그 데이터다.
    assert snap["data_name"] == "multi_sheet.xlsx" and snap["data_sheet"] == "낙찰현황"
    assert snap["source_fields"] == ["업체명", "낙찰금액", "계약일"]
    assert snap["record_count"] == 3
    ctx = snap["context"]
    assert ctx["entry_reason"] == "document_browser_new_work"
    assert ctx["evidence"] == {"데이터": "multi_sheet.xlsx"} and ctx["work"] == ""
    assert ctx["return_context"] == {"surface": "data"}
    # 템플릿을 고르고 2단계로 가면 매핑 모델이 **그 데이터의 헤더**로 선다(관문 재선택 불요).
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.dispatch("goto_section", {"section": "binding"})
    snap = ctrl.snapshot()
    assert snap["schema_only"] is False
    assert snap["active_source_fields"] == ["업체명", "낙찰금액", "계약일"]


def test_new_draft_with_data_validates_before_it_destroys(tmp_path):
    """배선 실수가 남의 세션을 조용히 지우지 않는다 — 문맥 검증이 `_reset` 보다 먼저다.

    미배선·미지 사유는 fail-closed 인데(링1), 그 거절이 초기화 **뒤에** 나면 사용자는
    아무 것도 못 얻고 편집 중이던 것만 잃는다. 거절 시 세션은 손대지 않은 채 남아야 한다.
    """
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.dispatch("set_name", {"name": "쓰던 작업"})
    with pytest.raises(ValueError, match="배선되지 않았습니다"):
        ctrl.new_draft_with_data({"path": str(MULTI_SHEET)}, entry_reason="workbench_result")
    assert ctrl.job_name == "쓰던 작업"                    # 세션 생존
    assert ctrl.template_path == str(TPL_COMPILED)
    assert ctrl.data_path == ""                            # 새 데이터도 서지 않았다


def test_anchored_draft_survives_the_real_template_pick(tmp_path):
    """#349 리뷰 3R P1 — 1단계에서 템플릿을 고르는 **실 dispatch** 를 타도 앵커가 산다.

    앞 라운드의 테스트는 `load_template_path` 를 직접 불러 이 결함을 통과시켰다: 실 UX 의
    다음 행동은 피커의 `use_library_template` 이고 그 경로는 `new_job_session` → `_reset()`
    이라, 「이 템플릿으로」를 누른 **모든 사용자**가 데이터 앵커와 진입 문맥을 잃었다.
    그래서 여기서는 링2 액션을 그대로 태운다 — 계약을 지키는 코드가 아니라 사용자가 밟는
    경로를 센다.

    끊기는 것은 종전 그대로임도 함께 못박는다(혼합 세션 금지는 살아 있다): 이름은 남지 않는다.
    """
    ctrl, _ = _controller_lib(tmp_path, paths=[TPL_COMPILED, TPL_PARTIAL])
    ctrl.new_draft_with_data(
        {"path": str(MULTI_SHEET), "sheet": "낙찰현황"},
        entry_reason="document_browser_new_work",
        evidence={"데이터": "multi_sheet.xlsx"},
        return_context={"surface": "data"},
    )
    ctrl.dispatch("set_name", {"name": "쓰던 이름"})

    ctrl.dispatch("use_library_template", {"path": str(TPL_COMPILED)})   # 실 UX 경로
    snap = ctrl.snapshot()
    assert snap["template_name"] == TPL_COMPILED.name
    assert snap["data_name"] == "multi_sheet.xlsx" and snap["data_sheet"] == "낙찰현황"
    assert snap["source_fields"] == ["업체명", "낙찰금액", "계약일"]
    assert snap["record_count"] == 3
    assert snap["context"]["entry_reason"] == "document_browser_new_work"
    assert snap["context"]["evidence"] == {"데이터": "multi_sheet.xlsx"}
    assert snap["name"] == ""            # 이름·매핑은 종전대로 끊긴다(혼합 세션 금지)

    # 마음을 바꿔 다른 템플릿을 골라도 앵커는 산다 — 문맥까지 되살아나야 성립하는 성질이다.
    ctrl.dispatch("use_library_template", {"path": str(TPL_PARTIAL)})
    snap = ctrl.snapshot()
    assert snap["data_name"] == "multi_sheet.xlsx"
    assert snap["context"]["entry_reason"] == "document_browser_new_work"

    # 대조군: 앵커 없이 시작한 보통 초안은 종전대로 데이터가 끊긴다(계약 무변경).
    plain, _ = _controller_lib(tmp_path, paths=[TPL_COMPILED, TPL_PARTIAL])
    plain.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    plain.dispatch("use_library_template", {"path": str(TPL_COMPILED)})
    assert plain.snapshot()["data_name"] == ""


def test_repair_entry_data_also_survives_the_template_pick(tmp_path):
    """#878 — 앵커 승계는 **진입 사유**가 열고 저장본 유무는 보지 않는다.

    종전 가드는 초안(`session.base is None`)만 앵커로 봤는데, 데이터를 들고 오는 진입이
    초안 하나뿐이라 그 조건이 사유 조건과 구별되지 않았다. 수리 진입은 저장된 작업을 여는데
    그 데이터도 사람이 아니라 진입이 들고 온 것이고, 템플릿을 갈아 끼우는 것은 그 세션의
    정상 진행이다 — 거기서 데이터가 끊기면 인계가 한 걸음 만에 무효가 된다.
    """
    from hwpxfiller.domain.job import Job

    ctrl, _ = _controller_lib(tmp_path, paths=[TPL_COMPILED, TPL_PARTIAL])
    ctrl.registry.save(Job(name="수리대상", template_path=str(TPL_COMPILED)))
    ctrl.load_job(
        "수리대상",
        entry_reason="document_browser_repair",
        return_context={"surface": "data"},
        source_ref={"path": str(MULTI_SHEET), "sheet": "낙찰현황"},
    )
    assert ctrl.snapshot()["data_name"] == "multi_sheet.xlsx"

    ctrl.dispatch("use_library_template", {"path": str(TPL_PARTIAL)})   # 실 UX 경로
    snap = ctrl.snapshot()
    assert snap["template_name"] == TPL_PARTIAL.name
    assert snap["data_name"] == "multi_sheet.xlsx" and snap["data_sheet"] == "낙찰현황"
    assert snap["record_count"] == 3
    assert snap["context"]["entry_reason"] == "document_browser_repair"

    # 대조군: 사람이 관문에서 고른 데이터는 종전대로 끊긴다(계약 무변경).
    plain, _ = _controller_lib(tmp_path, paths=[TPL_COMPILED, TPL_PARTIAL])
    plain.registry.save(Job(name="자발", template_path=str(TPL_COMPILED)))
    plain.load_job("자발")
    plain.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    plain.dispatch("use_library_template", {"path": str(TPL_PARTIAL)})
    assert plain.snapshot()["data_name"] == ""


def test_new_draft_carries_the_whole_reference_not_just_the_path(tmp_path):
    """#349 리뷰 P1 — 참조를 경로로 줄이면 **다른 헤더**의 데이터로 마법사가 선다.

    등록 데이터의 엑셀 참조는 `header_row` 를 들 수 있다. 그것을 떨어뜨리고 다시 열면
    사용자가 「문서 만들기」에서 본 열과 마법사가 앵커한 열이 갈리는데, 그 어긋남은 화면
    어디에도 표시가 없다(조용히 다른 데이터). 모델 정체 키도 같은 성분을 든다 — 두 판의
    헤더 이름이 우연히 겹치면 키가 불변이라 이전 기준의 확정 행이 그대로 산다.
    """
    xlsx = tmp_path / "머리2행.xlsx"
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["2026년 발주 목록", "작성 2026-07", "비고"])   # 1행 = 표제 줄(헤더 아님)
    ws.append(["부서", "사업명", "금액"])        # 2행 = 진짜 헤더
    ws.append(["총무과", "책상", "100"])
    wb.save(xlsx)

    ctrl, _ = _controller(tmp_path)
    ctrl.new_draft_with_data({"path": str(xlsx), "sheet": "", "header_row": 2})
    assert ctrl.source_fields == ["부서", "사업명", "금액"]
    assert ctrl.data_header_row == 2
    assert ctrl._model_key_now()[3] == 2      # 정체 키 성분 — 누락은 조용한 게이트 우회다
    # 대조군: 같은 파일을 경로만으로 열면 제목 줄이 헤더가 된다 = **다른 데이터**.
    other, _ = _controller(tmp_path)
    other.new_draft_with_data({"path": str(xlsx)})
    assert other.source_fields != ctrl.source_fields
    assert other.data_header_row == 0


def test_gateway_data_pick_rebuilds_mapping_in_place(tmp_path):
    """3단계 접기(블록 2 결정 11·12): 매핑 진입 후 관문에서 데이터를 고르면 매핑표가 그
    자리에서 다시 선다 — 컬럼·자동 제안 반영, 스키마온리 탈출, 전환 없음(라이브 순서 가드).

    헬퍼(_complete_with_data)는 데이터를 먼저 로드하고 진입하지만, 실 UX 는 진입 후 관문에서
    겨눔한다 — 그때 load_data_path 가 모델 존재를 보고 _ensure_model 로 재구성해야 한다.
    """
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_other_data(ctrl)                                            # 1단계 게이트(U6-B)
    ctrl.dispatch("goto_section", {"section": "binding"})              # 매핑 진입(첫 결속)
    snap = ctrl.snapshot()
    assert snap["section"] == "binding" and snap["schema_only"] is False
    assert snap["source_fields"] == ["공고명", "추정가격"] and snap["rows"]

    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")  # 고르기에서 데이터 갈아탐 → in-place 재구성
    snap = ctrl.snapshot()
    assert snap["section"] == "binding"                              # 여전히 매핑(단계 전환 없음)
    assert snap["schema_only"] is False                  # 새 데이터 반영
    assert snap["source_fields"] == ["업체명", "낙찰금액", "계약일"]
    assert snap["active_source_fields"] == ["업체명", "낙찰금액", "계약일"]  # 소스 후보 채워짐


def test_same_file_different_sheet_repick_demotes_confirmed(tmp_path):
    """3단계 접기 리뷰 F1 — 정체 키에 시트 포함: 같은 workbook 의 다른 시트로 관문 재겨눔할 때
    헤더명이 같아도 확정 매핑이 조용히 살아남지 않는다(조용한 게이트 우회 차단).

    두 시트의 헤더명이 같고(업체명·금액) 데이터만 다르면, data_sheet 를 키에서 빼면
    source_fields 불변→키 불변→_ensure_model 조기 반환→확정 유지→이전 시트 기준 저장·실행되는
    조용한 우회가 된다(슬라이스 4 '정체 키 성분 누락' 교훈). 시트를 키 성분으로 넣어 재구성·강등.
    """
    from openpyxl import Workbook

    xlsx = tmp_path / "twin_headers.xlsx"
    wb = Workbook()
    a = wb.active
    a.title = "1월"
    a.append(["업체명", "금액"])
    a.append(["갑상사", "100"])
    b = wb.create_sheet("2월")
    b.append(["업체명", "금액"])        # 동일 헤더명, 다른 데이터
    b.append(["을상사", "999"])
    wb.save(xlsx)

    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(xlsx), sheet="1월")
    ctrl.dispatch("goto_section", {"section": "binding"})            # 매핑 진입(1월 데이터)
    ctrl.dispatch("set_source", {"index": 0, "source": "금액"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    assert ctrl.snapshot()["is_complete"] is True

    ctrl.load_data_path(str(xlsx), sheet="2월")         # 같은 파일 다른 시트로 관문 재겨눔
    snap = ctrl.snapshot()
    assert snap["is_complete"] is False                # 확정이 조용히 살아남지 않음(재구성)
    assert all(row["confirmed"] is False for row in snap["rows"])
    assert snap["notice"] and "다시 확정" in snap["notice"]["text"]


def test_save_gate_blocks_incomplete_and_unnamed(tmp_path):
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    # 미확정 매핑 → 저장 차단(구체 사유 재진술).
    res = ctrl.dispatch("save", {})
    assert res["ok"] is False and "확정" in res["block_reason"]
    # 전부 비움 확정(내용 0) → 이름 있어도 '채울 값 없음' 차단.
    ctrl.dispatch("confirm_all", {})
    blanks = ctrl.snapshot()  # confirm_all 이 content 0 → 모두 blanks
    ctrl.dispatch("confirm_blanks", {"fields": [r["template_field"] for r in blanks["rows"]]})
    ctrl.dispatch("set_name", {"name": "빈작업"})
    ctrl.dispatch("set_pattern", {"pattern": "x-{{수요기관}}"})
    res = ctrl.dispatch("save", {})
    assert res["ok"] is False and "비움" in res["block_reason"]


def test_overwrite_confirm_flow(tmp_path):
    ctrl, _ = _controller(tmp_path)
    # 첫 저장.
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("set_name", {"name": "중복작업"})
    ctrl.dispatch("set_pattern", {"pattern": "p-{{수요기관}}"})
    assert ctrl.dispatch("save", {})["ok"] is True

    # **새 세션**에서 같은 이름 저장 → 덮어쓰기 확인 요구(조용한 덮어쓰기 금지).
    # 저장 착지가 편집 세션이 된 뒤(PR-2 리뷰 F2)로는 같은 세션의 같은 이름 재저장은
    # 자기-갱신(확인 불요)이 맞다 — 충돌 시나리오는 새 세션으로 재현한다.
    ctrl.dispatch("new_session", {})
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v2"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("set_name", {"name": "중복작업"})
    ctrl.dispatch("set_pattern", {"pattern": "p-{{수요기관}}"})
    res = ctrl.dispatch("save", {})
    assert res["ok"] is False and res.get("needs_overwrite") is True
    assert "덮어" in res["overwrite_text"]
    # 확인 후 재호출 → 저장.
    assert ctrl.dispatch(
        "save", {"confirm_overwrite": True, "confirmed_overwrite_text": res["overwrite_text"]}
    )["ok"] is True


def _save_named(ctrl: EditorController, name: str) -> dict:
    """이름 하나로 새 작업을 저장하는 최소 흐름(테스트 헬퍼)."""
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("set_name", {"name": name})
    ctrl.dispatch("set_pattern", {"pattern": "p-{{수요기관}}"})
    return ctrl.dispatch("save", {})


def test_slug_collision_different_name_restates_victim_then_saves(tmp_path):
    """다른 이름이 같은 slug 로 충돌하면 victim 을 재진술 확인하고, 확정 시 저장된다(#1).

    core 가드가 확정 저장 경로에서 allow_overwrite=True 로 통과하는지까지 검증 —
    확인했는데 JobSlugCollisionError 로 터지면 흐름이 깨진다.
    """
    ctrl, _ = _controller(tmp_path)
    assert _save_named(ctrl, "예산/2026")["ok"] is True

    res = _save_named(ctrl, "예산_2026")  # slug 동일 → 충돌
    assert res["ok"] is False and res.get("needs_overwrite") is True
    # 입력 이름·victim 이름이 모두 재진술된다(거짓 확인 방지).
    assert "예산_2026" in res["overwrite_text"] and "예산/2026" in res["overwrite_text"]
    # 확정 → allow_overwrite 로 core 가드 통과, 저장 성공(크래시 없음).
    assert ctrl.dispatch(
        "save", {"confirm_overwrite": True, "confirmed_overwrite_text": res["overwrite_text"]}
    )["ok"] is True
    assert JobRegistry(tmp_path / "jobs").exists("예산_2026")


# ------------------------------------------------------------ #25 세션 혼합 방지
def _build_complete_session(ctrl, name: str) -> None:
    """COMPILED 템플릿으로 저장 가능한 완결 세션 구성(저장 직전까지) — 혼합 테스트 준비."""
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("goto_section", {"section": "filename"})
    ctrl.dispatch("set_name", {"name": name})
    ctrl.dispatch("set_pattern", {"pattern": "p-{{수요기관}}"})


def test_has_unsaved_work_tracks_session_lifecycle(tmp_path):
    ctrl, _ = _controller(tmp_path)
    assert ctrl.has_unsaved_work() is False              # 갓 초기화 — 버릴 것 없음
    ctrl.load_template_path(str(TPL_COMPILED))
    assert ctrl.has_unsaved_work() is False              # 템플릿만 로드 — 아직 세션 아님
    _mount_data(ctrl)                                    # 1단계 게이트(U6-B)
    ctrl.dispatch("goto_section", {"section": "binding"})  # 매핑 모델 생성 → 진행 중 세션
    assert ctrl.has_unsaved_work() is True
    assert ctrl.snapshot()["dirty"] is True              # 스냅샷의 얼굴은 dirty 하나다


def test_handed_over_data_is_the_draft_baseline_not_an_unsaved_change(tmp_path):
    """#945 F7 — 「이 데이터로 새 작업」 무조작 진입은 미저장이 아니다(초안 갈래 정렬).

    저장본 갈래는 #878 에서 이미 그렇게 센다(`_extras_of` — 진입이 들고 온 데이터는 사람이
    고른 적이 없으므로 기준 그 자체). 초안 갈래만 `data_path` 를 날것으로 세어, 데이터를
    골라 새 작업으로 들어온 사람은 아무것도 손대지 않고도 첫 이탈부터 「버리고 계속」을
    물었다. 이제 두 갈래가 같은 기준선을 쓴다 — 1단계 템플릿 선택(마법사의 정상 진행)에서도
    앵커와 함께 기준선이 건너간다.
    """
    ctrl, _ = _controller_lib(tmp_path, paths=[TPL_COMPILED, TPL_PARTIAL])
    ctrl.new_draft_with_data(
        {"path": str(MULTI_SHEET), "sheet": "낙찰현황"},
        entry_reason="document_browser_new_work",
        evidence={"데이터": "multi_sheet.xlsx"},
        return_context={"surface": "data"},
    )
    assert ctrl.data_path == str(MULTI_SHEET)             # 데이터는 서 있고
    assert ctrl.has_unsaved_work() is False               # 사람이 손댄 것은 없다
    assert ctrl.snapshot()["dirty"] is False              # 스냅샷의 얼굴도 같은 값

    ctrl.dispatch("use_library_template", {"path": str(TPL_COMPILED)})   # 실 UX 경로
    assert ctrl.snapshot()["data_name"] == "multi_sheet.xlsx"  # 앵커 생존(계약 무변경)
    assert ctrl.has_unsaved_work() is False, "템플릿만 고른 진행이 미저장으로 섭니다."

    ctrl.dispatch("set_name", {"name": "손댄 이름"})       # 양성 대조 — 손대면 즉시 미저장
    assert ctrl.has_unsaved_work() is True


def test_swapping_the_handed_over_data_returns_the_draft_to_unsaved(tmp_path):
    """면제는 **진입이 세운 그 참조**에만 선다 — 관문에서 갈아타면 다시 미저장이다.

    면제를 「데이터가 있으면 안 센다」로 넓히면 사람이 고른 데이터를 조용히 버리는 이탈이
    생긴다(같은 결함의 반대 얼굴). 경로가 같아도 시트가 갈리면 다른 데이터다(#33).
    """
    ctrl, _ = _controller(tmp_path)
    ctrl.new_draft_with_data(
        {"path": str(MULTI_SHEET), "sheet": "낙찰현황"},
        entry_reason="document_browser_new_work",
    )
    assert ctrl.has_unsaved_work() is False
    ctrl.load_data_path(str(MULTI_SHEET), sheet="공고목록")   # 사람이 관문에서 갈아탄다
    assert ctrl.data_path == str(MULTI_SHEET)                 # 경로는 그대로인데
    assert ctrl.has_unsaved_work() is True                    # 시트가 갈렸다 = 사람의 선택


def test_new_job_session_atomically_clears_prior_session_and_blocks_mixed_save(tmp_path):
    """템플릿 A 진행 세션 → new_job_session(B) 는 이름·데이터·매핑·단계를 원자 초기화(#25)."""
    ctrl, _ = _controller(tmp_path)
    _build_complete_session(ctrl, "작업A")
    assert ctrl.snapshot()["is_complete"] is True and ctrl.has_unsaved_work() is True

    ctrl.new_job_session(str(TPL_PARTIAL))               # 다른 템플릿으로 새 세션
    snap = ctrl.snapshot()
    assert snap["section"] == "template"                             # 단계 초기화
    assert snap["name"] == ""                            # 이름 소거(A 잔존 없음)
    assert snap["rows"] == [] and snap["is_complete"] is False  # 구 매핑 모델 폐기
    assert snap["data_path"] == ""                       # 데이터 소거
    res = ctrl.dispatch("save", {})
    assert res["ok"] is False and "확정" in res["block_reason"]  # 모델 리셋 → 미확정 차단


def test_new_session_action_resets_prior_session(tmp_path):
    """홈 「＋ 새 작업」의 new_session 액션 — 진행 세션 전량 초기화(F10).

    종전 홈 버튼은 bare nav 라 직전 세션(이름·데이터·매핑·단계)이 그대로 복원돼
    라벨 '새'가 사실상 '이전 작성 계속'이었다. 초기화 뒤엔 미저장 판정도 소거된다
    (방금 저장 직후처럼 — 다음 「새 작업」이 불필요한 확인을 띄우지 않게).
    """
    from hwpxfiller.domain.job import DEFAULT_FILENAME_PATTERN

    ctrl, pushes = _controller(tmp_path)
    _build_complete_session(ctrl, "작업A")
    assert ctrl.has_unsaved_work() is True
    ctrl.dispatch("new_session", {})
    snap = pushes[-1][1]                                 # dispatch 말미 자동 푸시
    assert ctrl.has_unsaved_work() is False
    assert snap["section"] == "template" and snap["name"] == ""
    assert snap["rows"] == [] and snap["data_path"] == ""
    assert snap["pattern"] == DEFAULT_FILENAME_PATTERN   # 패턴도 기본으로 복원


def test_discard_session_cancels_new_wizard_but_rejects_saved_edit(tmp_path):
    """신규 마법사 취소는 휘발 상태를 실제 폐기하고, 저장 작업 편집에는 오용되지 않는다."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl)                                    # 1단계 게이트(U6-B)
    ctrl.dispatch("goto_section", {"section": "binding"})
    assert ctrl.has_unsaved_work() is True
    ctrl.dispatch("discard_session", {})
    snap = ctrl.snapshot()
    assert snap["section"] == "template" and ctrl.template_path == "" and ctrl.model is None
    assert ctrl.has_unsaved_work() is False

    # 편집 모드는 별도 비파괴 복귀 계약(T2)을 쓰며 신규 취소 액션으로 닫을 수 없다.
    ctrl._editing_origin = "저장작업"
    with pytest.raises(ValueError, match="저장된 작업 편집"):
        ctrl.dispatch("discard_session", {})


# ------------------------------------- U4-E2 #939 편집기 템플릿 슬롯 구조 요약
def _structured_template(tmp_path: Path, name: str = "구간템플릿") -> Path:
    """항목 1 · 선택 2 · 누름틀 3 을 가진 실 HWPX — 표기 → 두 축 컴파일(계약 순서).

    필드 토큰을 먼저(`compile_document`), 구간 표기를 다음(`compile_structure`)에 굽는다.
    순서를 뒤집으면 구조 안의 `{{필드}}` 가 depth>0 이 되어 필드 컴파일에서 조용히 빠진다
    (`docs/UI_CONTRACT.md` 의 변환 순서 계약과 같은 이유). 표기 원본은 S8-02 헬퍼 재사용.
    """
    from hwpxfiller.external.template_inspection import compile_document, compile_structure
    from test_structure_compile import NOTATION, _pkg, _text  # noqa: E402 rootdir 임포트

    pkg = _pkg(*(_text(line) for line in NOTATION))
    compile_document(pkg)
    report = compile_structure(pkg)
    assert (report.modified, report.refusal) == (True, None)
    path = tmp_path / f"{name}.hwpx"
    path.write_bytes(pkg.to_bytes())
    return path


def test_template_slots_stand_for_a_structured_template(tmp_path):
    """구조를 가진 템플릿을 열면 슬롯 축 요약이 선다 — 요약·행은 링1 투영 그대로.

    컨트롤러가 개수를 다시 세지 않는다는 것이 이 단언의 요점이다: 요약 문자열은
    `SlotView.summary()` 소유이고 스냅샷은 그 값을 그대로 나른다.
    """
    ctrl, pushes = _controller(tmp_path)
    tpl = _structured_template(tmp_path)
    ctrl.load_template_path(str(tpl))

    slots = pushes[-1][1]["template_slots"]
    assert slots is not None
    assert slots["summary"] == "항목 1개 · 선택 2개"
    assert slots["path"] == str(tpl) and slots["name"] == tpl.name
    assert slots["diagnostics"] == []
    assert [(r["id"], r["label"], r["option_count"]) for r in slots["rows"]] == [
        ("특약", "특약 사항", 2)
    ]
    assert slots["rows"][0]["options"] == ["지체상금 조항", "하자보수 조항"]
    # tpl 검토가 내는 것과 **같은 모양**이어야 한다(프런트가 스키마를 둘 배우지 않는다).
    assert set(slots) == {"path", "name", "summary", "rows", "diagnostics"}


def test_template_without_structure_stands_no_slot_zone(tmp_path):
    """구조가 없으면 ``None`` — 확인할 것이 없으면 숨긴다(U3 #876)."""
    ctrl, pushes = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    assert pushes[-1][1]["template_slots"] is None
    # 템플릿 이전(빈 세션)에도 서지 않는다.
    ctrl2, _ = _controller(tmp_path)
    assert ctrl2.snapshot()["template_slots"] is None


def test_template_slots_die_with_the_template_swap_and_the_session_reset(tmp_path):
    """수명은 편집 세션 소유 — 템플릿 교체와 세션 초기화가 각각 걷는다."""
    ctrl, pushes = _controller(tmp_path)
    tpl = _structured_template(tmp_path)
    ctrl.load_template_path(str(tpl))
    assert pushes[-1][1]["template_slots"] is not None

    ctrl.load_template_path(str(TPL_COMPILED))            # 구조 없는 템플릿으로 교체
    assert pushes[-1][1]["template_slots"] is None
    assert ctrl.template_slots is None

    ctrl.load_template_path(str(tpl))                     # 다시 세우고
    assert ctrl.template_slots is not None
    ctrl.dispatch("new_session", {})                      # _reset 이 지운다
    assert pushes[-1][1]["template_slots"] is None
    assert ctrl.template_slots is None


# --------------------------------------------------- #16 1·2단계 구조화 렌더 가드
_EDITOR_JS = SOURCE_JS_DIR.parent / "src" / "screens" / "editor.ts"


def test_editor_renders_structured_field_and_data_tables():
    """1·2단계가 나열식 텍스트가 아니라 구조화 표로 렌더돼야 한다(#16 98DDFE96).

    나열식 `.fields-line` 은 제거되고, 1단계는 `schema-fields` 표·2단계는 `data-preview`
    표로 승격. 빈 셀은 ADR-B 대로 "(빈 값)"으로 시끄럽게 표기한다. 실 렌더 되읽기는
    selftest 게이트가 하고, 여기선 마크업 배선의 존재/부재를 정적으로 가드한다.
    """
    src = _EDITOR_JS.read_text(encoding="utf-8")
    assert "fields-line" not in src, "나열식 .fields-line 이 남아 있습니다 — 구조화 표로 교체(#16)."
    assert 'className: "schema-fields"' in src, "1단계 필드 구조화 표(schema-fields)가 없습니다(#16)."
    assert 'className: "data-preview"' in src, "2단계 데이터 미리보기 표(data-preview)가 없습니다(#16)."
    assert "(빈 값)" in src, "2단계 빈 셀의 시끄러운 표기가 없습니다 — ADR-B 위반(#16)."


def test_save_blocks_when_model_schema_mismatches_template(tmp_path):
    """방어층: 모델이 현재 스키마와 어긋나면(혼합) 저장을 시끄럽게 차단(#25 항목4)."""
    ctrl, _ = _controller(tmp_path)
    _build_complete_session(ctrl, "작업A")               # 모델 = COMPILED 스키마
    # new_job_session 을 우회해 low-level 로 스키마만 교체(구버그 경로 재현) → 모델은 A 그대로.
    ctrl.load_template_path(str(TPL_PARTIAL))
    res = ctrl.dispatch("save", {})
    assert res["ok"] is False and "일치하지 않습니다" in res["block_reason"]


# ============================================================ #26 패리티 회수
# 편집 모드(#1)·선언 데이터 자동등록(#3)의 헤드리스 계약.
from hwpxfiller.domain.dataset_reference import DatasetReference
from hwpxfiller.external.dataset_store import DatasetPoolRegistry
from hwpxfiller.domain.job import Job
from hwpxfiller.domain.mapping import FieldMapping, MappingProfile


def _controller26(tmp_path: Path):
    """작업 레지스트리를 tmp 로 격리 주입한 컨트롤러(#347: 풀 주입은 자동등록과 함께 사망)."""
    pushes: list = []
    ctrl = EditorController(
        JobRegistry(tmp_path / "jobs"),
        lambda s, snap: pushes.append((s, snap)),
        clock=_clock,
        template_library=TemplateManagerViewModel(
            paths=[],
            inspect_template=inspect_hwpx_template,
            file_ops=HWPX_TEMPLATE_OPS,
        ),
        text_registry=TextTemplateRegistry(tmp_path / "text_templates"),
    )
    return ctrl, pushes


# ------------------------------------------------------------ 편집 모드(#1)
def test_load_job_restores_edit_session(tmp_path):
    """저장 작업 → load_job: 이름·패턴·확정 매핑·단계가 복원되고 원점이 기록된다."""
    ctrl, _ = _controller26(tmp_path)
    assert _save_named(ctrl, "원본작업")["ok"] is True   # 저장 후 세션 리셋

    ctrl.load_job("원본작업")
    snap = ctrl.snapshot()
    assert snap["section"] == "binding"                             # 매핑 확정 단계로 착지(3단계 접기)
    assert snap["name"] == "원본작업"
    assert snap["editing_origin"] == "원본작업"
    assert snap["is_complete"] is True                   # 1 const + 9 blank 전부 확정 복원
    assert snap["rows"][0]["type"] == "const" and snap["rows"][0]["const"] == "v"
    assert snap["notice"] and "편집합니다" in snap["notice"]["text"]


def test_new_session_action_clears_edit_mode(tmp_path):
    """편집 모드 중 「＋ 새 작업」(new_session) — 편집 원점·복원 notice 까지 소거(F10).

    남으면 새 세션 저장이 '원본작업' 자기-갱신으로 오판되거나 편집 배너가 거짓으로 남는다.
    """
    ctrl, pushes = _controller26(tmp_path)
    _save_named(ctrl, "원본작업")
    ctrl.load_job("원본작업")
    assert ctrl.snapshot()["editing_origin"] == "원본작업"
    ctrl.dispatch("new_session", {})
    snap = pushes[-1][1]
    assert snap["editing_origin"] == "" and snap["name"] == ""
    assert snap["notice"] is None                        # 편집 모드 배너 소거
    assert ctrl.has_unsaved_work() is False


def test_load_job_model_survives_step_navigation(tmp_path):
    """_model_key 함정 봉쇄 — 복원 직후 단계를 오가도 확정 매핑이 초안으로 대체되지 않는다."""
    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "이동작업")
    ctrl.load_job("이동작업")
    ctrl.dispatch("goto_section", {"section": "filename"})              # 저장 단계로
    ctrl.dispatch("goto_section", {"section": "binding"})              # 다시 매핑 진입(_ensure_model 경유)
    snap = ctrl.snapshot()
    assert snap["is_complete"] is True                   # 확정 유지(재생성 아님)
    assert snap["rows"][0]["const"] == "v"


def test_load_job_missing_template_is_loud(tmp_path):
    """템플릿 파일이 사라진 작업의 편집 열기는 조용히 반쯤 열리지 않고 시끄럽게 거절."""
    ctrl, _ = _controller26(tmp_path)
    reg = JobRegistry(tmp_path / "jobs")
    reg.save(Job(name="유실", template_path=str(tmp_path / "없는파일.hwpx")))
    with pytest.raises(ValueError, match="찾을 수 없습니다"):
        ctrl.load_job("유실")


def test_load_job_template_drift_is_restated(tmp_path):
    """저장 매핑에 있으나 현 스키마에 없는 필드는 조용히 누락되지 않고 notice 로 재진술."""
    ctrl, _ = _controller26(tmp_path)
    reg = JobRegistry(tmp_path / "jobs")
    reg.save(Job(
        name="드리프트",
        template_path=str(TPL_COMPILED),
        mapping=MappingProfile(mappings=[
            FieldMapping(template_field="유령필드", source="", type="const", const="x"),
        ]),
    ))
    ctrl.load_job("드리프트")
    snap = ctrl.snapshot()
    assert snap["notice"]["level"] == "warn"
    assert "유령필드" in snap["notice"]["text"]          # 제외 필드 이름 재진술
    assert snap["is_complete"] is False                  # 새 스키마 필드는 미확정(사람 확정 필요)


def test_session_notice_is_dismissible_and_the_trigger_can_stand_it_again(tmp_path):
    """세운 뒤 **지워지는가**(U4 계열1-20 · U3 §1 이 지목한 결함류).

    이 채널에는 세우는 전이만 있고 지우는 전이가 없어서 한 번 선 통지가 사유 해소 뒤에도
    남았다. 닫기는 사용자 몫이고, 트리거는 그대로라 같은 사유가 다시 서면 통지도 다시 선다.
    """
    ctrl, _ = _controller26(tmp_path)
    reg = JobRegistry(tmp_path / "jobs")
    reg.save(Job(
        name="드리프트",
        template_path=str(TPL_COMPILED),
        mapping=MappingProfile(mappings=[
            FieldMapping(template_field="유령필드", source="", type="const", const="x"),
        ]),
    ))
    ctrl.load_job("드리프트")
    assert ctrl.snapshot()["notice"] is not None

    ctrl.dispatch("dismiss_notice", {})
    assert ctrl.snapshot()["notice"] is None

    ctrl.load_job("드리프트")                              # 같은 트리거 → 같은 통지가 다시 선다
    assert ctrl.snapshot()["notice"] is not None


def test_edit_save_self_update_skips_overwrite_and_preserves_meta(tmp_path):
    """편집 원점 그대로 재저장 = 자기-갱신(확인 불요) + 태그·마지막 실행 메타 보존."""
    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "메타작업")
    reg = JobRegistry(tmp_path / "jobs")
    job = reg.load("메타작업")
    job.tags = {"물품": "의약품"}
    job.last_run_at = "2026-07-01T09:00:00"
    job.authority_id = "work-authority-1"
    reg.save(job, allow_overwrite=True)

    ctrl.load_job("메타작업")
    res = ctrl.dispatch("save", {})                      # 같은 이름 재저장
    assert res["ok"] is True                             # needs_overwrite 없이 통과(자기-갱신)
    saved = reg.load("메타작업")
    assert saved.tags == {"물품": "의약품"}              # 태그 조용한 소실 없음
    assert saved.last_run_at == "2026-07-01T09:00:00"
    assert saved.authority_id == "work-authority-1"


def test_edit_save_holds_the_registry_write_lock(tmp_path):
    """저장의 재읽기~쓰기 구간이 **레지스트리 공유 잠금 안**에 있다(#129 리뷰 2R P1).

    보존 값(태그·last_run_at)을 읽은 뒤 저장까지 사이에 생성 스레드의 스탬프가 끼면, 여기서
    만든 Job 이 방금 찍힌 시각을 낡은 값으로 되돌린다. 저장 한 번만 원자적인 것으로는 못 막아
    구간 전체가 잠겨야 하므로, 저장 시점에 잠금이 **다른 스레드에서 잡히지 않는지**로 되읽는다.
    """
    import threading

    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "잠금작업")
    ctrl.load_job("잠금작업")
    seen: "list[bool]" = []
    real_save = ctrl.registry.save

    def spy(job, **kw):
        got = [None]

        def probe():  # 다른 스레드에서 비차단 획득 시도 — 잠겨 있으면 실패해야 한다
            lock = ctrl.registry.write_lock()
            got[0] = lock.acquire(blocking=False)
            if got[0]:
                lock.release()

        t = threading.Thread(target=probe)
        t.start()
        t.join(3)
        seen.append(bool(got[0]))
        return real_save(job, **kw)

    ctrl.registry.save = spy  # type: ignore[method-assign]
    assert ctrl.dispatch("save", {})["ok"] is True
    assert seen and not any(seen), "저장 구간이 쓰기 잠금 밖입니다 — lost update 회귀."


def test_edit_save_renamed_still_confirms_overwrite(tmp_path):
    """편집 중 이름을 다른 기존 작업으로 바꾸면 평소처럼 덮어쓰기 확인을 요구."""
    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "작업일")
    _save_named(ctrl, "작업이")
    ctrl.load_job("작업일")
    ctrl.dispatch("set_name", {"name": "작업이"})        # 다른 작업을 겨냥
    res = ctrl.dispatch("save", {})
    assert res["ok"] is False and res.get("needs_overwrite") is True


def test_ensure_model_carries_values_but_requires_reconfirm_on_data_change(tmp_path):
    """데이터 교체 시 값(소스·유형·서식)은 제안으로 이월하되 확정은 전원 해제(r3 C1).

    이전 확정을 확정 상태 그대로 되살리면 같은 이름 컬럼('금액' 등)이 의미가 다른 새
    데이터에서 사람 검토 없이 ``is_complete`` 를 통과해 저장·실행까지 흐른다 — 구
    불변식 '키 변경 시 전원 미확정 초안'을 복원하고 notice 로 재확정 필요를 재진술한다.

    3단계 접기(블록 2): 데이터 교체는 매핑 단계 관문에서 일어나 **그 자리에서** 모델을
    다시 세운다(load_data_path 가 모델 존재 시 _ensure_model 호출) — 단계 왕복 없이 in-place.
    """
    ctrl, _ = _controller26(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_other_data(ctrl)                               # 1단계 게이트(U6-B) — 갈아탈 첫 결속
    ctrl.dispatch("goto_section", {"section": "binding"})  # 매핑 진입(모델 생성)
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "보존값"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    assert ctrl.snapshot()["is_complete"] is True

    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")  # 관문에서 데이터 겨눔 → in-place 재생성
    snap = ctrl.snapshot()
    assert snap["rows"][0]["const"] == "보존값"          # 값 이월(조용한 소실 없음)
    assert snap["rows"][0]["type"] == "const"
    assert all(row["confirmed"] is False for row in snap["rows"])  # 확정 전원 해제
    assert snap["is_complete"] is False                  # 재확정 없이는 저장 게이트 미통과
    assert snap["notice"] and "다시 확정" in snap["notice"]["text"]
    assert "다시 확정" in snap["notice"]["text"]         # 재확정 필요를 loud 재진술


# --------------------------------------- 선언 데이터 자동등록의 사망(#347, U2 §5.3 D)
def _complete_with_data(ctrl, name: str) -> None:
    """데이터(다중시트 확정) 연결 세션을 저장 직전까지 구성."""
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})   # 매핑 진입(데이터 겨눔 상태 — 3단계 접기)
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("set_name", {"name": name})
    ctrl.dispatch("set_pattern", {"pattern": "p-{{수요기관}}"})


def test_save_with_data_registers_nothing_anywhere(tmp_path, monkeypatch):
    """저장은 데이터를 풀에 등록하지 않는다 — 자동등록(#18·#26)은 U2 §5.3 판정 D 로 폐기.

    편집 세션의 데이터는 검토용 문맥일 뿐이고, 풀 등록은 데이터 선택 면의 「이 데이터
    고정」 명시 행동 하나다. 홈 풀 디렉터리까지 확인해 어떤 경로로도 등록이 새지 않음을
    본다(조용한 durable 쓰기 금지).
    """
    monkeypatch.setenv("HWPXFILLER_HOME", str(tmp_path / "home"))
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "데이터작업")
    snap = ctrl.snapshot()
    assert "dataset_name" not in snap                 # 자동등록 표면째 소멸
    assert "default_dataset" not in snap
    res = ctrl.dispatch("save", {})
    assert res["ok"] is True
    assert "dataset_registered" not in res and "dataset_register_error" not in res
    assert DatasetPoolRegistry(tmp_path / "pool").list_items() == []
    assert DatasetPoolRegistry(tmp_path / "home" / "datasets").list_items() == []


# ------------------------------------------------------- 매핑 프로파일 제거(F22)
def test_removed_dataset_and_profile_actions_are_gone_loudly(tmp_path):
    """폐기한 자동등록·매핑 프로파일 액션은 미지 액션으로 loud 거절된다.

    작업이 매핑을 자족 저장·복원하므로 별도 프로파일 저장 개념은 제거 — 재사용은
    「작업 복제」(홈 clone_job)로 수렴한다. 조용한 no-op 잔존이 아니라 표면째 소멸.
    """
    ctrl, _ = _controller26(tmp_path)
    for action in (
        "frobnicate", "set_dataset_name", "profile_list", "profile_apply", "profile_save",
        "profile_delete",
    ):
        with pytest.raises(ValueError, match="알 수 없는 editor 액션"):
            ctrl.dispatch(action, {"name": "x"})


def test_old_job_json_with_base_mapping_name_still_loads(tmp_path):
    """구 JSON 의 base_mapping_name(제거된 J3 계보 메타)은 미지 키로 무시된다 — 하위호환."""
    ctrl, _ = _controller26(tmp_path)
    assert _save_named(ctrl, "구식작업")["ok"] is True
    reg = JobRegistry(tmp_path / "jobs")
    path = reg.path_for("구식작업")
    import json as _json
    payload = _json.loads(path.read_text(encoding="utf-8"))
    payload["base_mapping_name"] = "지워진베이스"          # 구버전이 남긴 키 시뮬레이션
    path.write_text(_json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    job = reg.load("구식작업")                              # loud raise 없이 로드
    assert job.name == "구식작업"
    assert "base_mapping_name" not in encode_job(job)         # 재저장 시 키 소멸


def test_edit_save_preserves_concurrent_home_tag_edit(tmp_path):
    """편집 세션이 열린 사이 홈에서 단 태그를, 에디터 저장이 stale 스냅샷으로 되돌리지 않는다(#26 #2·#5).

    load_job 시점 태그 스냅샷이 아니라 저장 직전 디스크 상태를 재읽어 보존해야 한다.
    """
    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "태그작업")
    ctrl.load_job("태그작업")                       # 에디터가 빈 태그 스냅샷을 뜬다
    # 편집 세션이 열린 사이 홈 태그 편집(같은 레지스트리 디스크 갱신)을 시뮬레이션.
    reg = JobRegistry(tmp_path / "jobs")
    job = reg.load("태그작업")
    job.tags = {"물품": "의약품"}
    reg.save(job, allow_overwrite=True)

    assert ctrl.dispatch("save", {})["ok"] is True   # 아직 열린 편집 세션 저장
    assert reg.load("태그작업").tags == {"물품": "의약품"}   # 조용한 소실 없음


def test_save_does_not_touch_existing_pool_entries(tmp_path):
    """저장은 기존 풀 항목(보관·메모 포함)을 어떤 방식으로도 건드리지 않는다(#347).

    구 자동등록은 동명 항목의 참조를 갱신했다 — 그 게이트가 §2.8 의 danger 경보
    인플레이션이었고, 폐기 뒤에는 저장이 풀을 읽지도 쓰지도 않는다.
    """
    pool = DatasetPoolRegistry(tmp_path / "pool")
    prior = DatasetReference(
        name="multi_sheet", kind="excel", opts={"path": "old.xlsx"}, note="계약 종료분")
    prior.archive()
    key = pool.add(prior)

    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "재사용작업")
    res = ctrl.dispatch("save", {})
    assert res["ok"] is True and "needs_dataset_confirm" not in res  # 확인 게이트 소멸(§2.8)

    item = pool.load(key)
    assert item.status == "archived" and item.note == "계약 종료분"
    assert item.opts == {"path": "old.xlsx"}          # 참조 불변 — 저장은 풀에 무접촉


# ------------------------------------------------- 작성 출처 provenance(#53-C)
def test_save_stamps_provenance_on_mapping(tmp_path):
    """저장 시 매핑에 작성 출처 지문(템플릿·데이터·스키마·시각)이 새겨진다(#53-C)."""
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "출처작업")
    ctrl.dispatch("save", {})
    prov = JobRegistry(tmp_path / "jobs").load("출처작업").mapping.provenance
    assert prov["template"].endswith(".hwpx")
    assert prov["dataset"] == "multi_sheet"
    assert prov["template_fields"]                    # 템플릿 스키마 지문
    assert prov["authored_at"] == prov["updated_at"] == _NOW.isoformat(timespec="seconds")
    # 순수 메타 — 실행 계약(source_keys)과 별개 축.
    assert isinstance(prov, dict)


def test_edit_save_preserves_authored_at_updates_updated_at(tmp_path):
    """편집 재저장은 최초 작성시각(authored_at)을 보존하고 updated_at 만 갱신한다(#53-C)."""
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "출처편집")
    ctrl.dispatch("save", {})
    first = JobRegistry(tmp_path / "jobs").load("출처편집").mapping.provenance
    authored = first["authored_at"]

    ctrl.load_job("출처편집")
    assert ctrl.snapshot()["provenance"]["template"].endswith(".hwpx")  # 편집 모드 표시
    ctrl.dispatch("set_pattern", {"pattern": "새-{{수요기관}}"})
    ctrl.dispatch("save", {"confirm_overwrite": True})
    second = JobRegistry(tmp_path / "jobs").load("출처편집").mapping.provenance
    assert second["authored_at"] == authored          # 최초 작성시각 보존


def test_new_session_has_no_provenance(tmp_path):
    """저장 전(신규 세션)엔 표시할 작성 출처가 없다."""
    ctrl, _ = _controller26(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    assert ctrl.snapshot()["provenance"] is None


# ------------------------------- 작업↔데이터 결속의 사망(#53-A → #347, U2 §5.3 D)
def test_saved_job_carries_no_dataset_binding(tmp_path):
    """저장된 작업 JSON 에 데이터 결속 키가 없다 — `default_dataset_ref` 는 개념째 폐기.

    작업이 기억하는 것은 스키마(source_keys)뿐이고, 데이터↔작업 결속은 어느 방향으로도
    다시 들이지 않는다(§5.3 뒷문 금지).
    """
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "연결작업")
    assert ctrl.dispatch("save", {})["ok"] is True
    job = JobRegistry(tmp_path / "jobs").load("연결작업")
    assert not hasattr(job, "default_dataset_ref")
    assert "default_dataset_ref" not in encode_job(job)


def test_legacy_default_dataset_ref_key_is_discarded_not_migrated(tmp_path):
    """구 JSON 의 default_dataset_ref 는 미지 키로 무시된다 — 마이그레이션이 아니라 폐기.

    「이 작업이 지난번 쓰던 데이터」 정보는 사용자 확인 하에 사라졌다(U2 §5.3 명시
    확인분). 로드는 loud raise 없이 통과하고 재저장 시 키가 소멸한다.
    """
    ctrl, _ = _controller26(tmp_path)
    assert _save_named(ctrl, "구식결속")["ok"] is True
    reg = JobRegistry(tmp_path / "jobs")
    path = reg.path_for("구식결속")
    import json as _json
    payload = _json.loads(path.read_text(encoding="utf-8"))
    payload["default_dataset_ref"] = "지난달데이터"     # 구버전이 남긴 결속 키
    path.write_text(_json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    job = reg.load("구식결속")                          # loud raise 없이 로드
    assert job.name == "구식결속"
    assert "default_dataset_ref" not in encode_job(job)   # 재저장 시 키 소멸(폐기)


# ------------------------------------------------- 사용할 헤더 선택(#49)
def test_header_selection_lifecycle_defaults_narrows_and_resets(tmp_path):
    """헤더는 전원 활성으로 시작해 개별·일괄 복원되고 새 데이터에서 초기화된다."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    snap = ctrl.snapshot()
    assert snap["source_fields"] == ["업체명", "낙찰금액", "계약일"]       # 전체 헤더 불변
    assert snap["active_source_fields"] == ["업체명", "낙찰금액", "계약일"]  # 기본 전원 활성
    assert snap["active_count"] == 3 and snap["ignored_count"] == 0

    ctrl.dispatch("toggle_source_active", {"field": "낙찰금액"})          # 칩 즉시 토글 off
    ctrl.dispatch("toggle_source_active", {"field": "계약일"})
    snap = ctrl.snapshot()
    assert snap["active_source_fields"] == ["업체명"]                    # 활성만 후보(원 순서)
    assert snap["ignored_source_fields"] == ["낙찰금액", "계약일"]
    assert snap["active_count"] == 1 and snap["ignored_count"] == 2
    assert snap["notice"] and "사용 데이터 열 1개 · 미사용 2개" in snap["notice"]["text"]
    ctrl.dispatch("toggle_source_active", {"field": "낙찰금액"})
    assert "낙찰금액" in ctrl.snapshot()["active_source_fields"]
    ctrl.dispatch("use_all_headers", {})
    assert ctrl.snapshot()["ignored_count"] == 0
    ctrl.dispatch("toggle_source_active", {"field": "낙찰금액"})
    ctrl.dispatch("toggle_source_active", {"field": "계약일"})
    ctrl.load_data_path(str(MULTI_SHEET))
    snap = ctrl.snapshot()
    assert snap["source_fields"] == ["공고명", "추정가격"]
    assert snap["ignored_count"] == 0 and snap["active_source_fields"] == snap["source_fields"]


def test_ignoring_mapped_header_r4_demotes_human_owned_and_restates(tmp_path):
    """사람 소유(확정) 행의 소스 헤더를 끄면 R4 시끄러운 강등 — 확정 해제·이름 재진술(결정 12).
    활성 소스를 쓰는 다른 사람 소유 행은 그대로. 원본 데이터는 불변."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})              # 매핑 진입 → 모델 생성(3단계 접기)
    ctrl.dispatch("set_source", {"index": 0, "source": "낙찰금액"})
    ctrl.dispatch("set_source", {"index": 1, "source": "업체명"})
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": True})
    ctrl.dispatch("set_confirmed", {"index": 1, "confirmed": True})

    ctrl.dispatch("toggle_source_active", {"field": "낙찰금액"})          # 낙찰금액 칩 off
    snap = ctrl.snapshot()
    # 행 0(낙찰금액 사용, 확정)은 R4 강등 — 확정 해제·시스템 소유로(touched=False).
    assert snap["rows"][0]["source"] == "" and snap["rows"][0]["confirmed"] is False
    assert snap["rows"][0]["touched"] is False
    # 행 1(업체명, 활성)은 사람 소유 그대로.
    assert snap["rows"][1]["source"] == "업체명" and snap["rows"][1]["confirmed"] is True
    assert "낙찰금액" not in snap["active_source_fields"]
    assert snap["notice"]["level"] == "warn" and "재확정" in snap["notice"]["text"]


def test_use_none_blocks_on_confirmed_but_allows_when_clean(tmp_path):
    """전체 미사용(결정 13 개정) — 확정 있으면 차단(파괴 방지), 없으면 허용 + 미사용 구역 펼침.

    구 '전부 미사용 무조건 거부'(#62)를 결정 13 이 개정: 되돌릴 수 없는 **확정** 파괴만
    사전 차단하고, 확정이 없으면 '고른다→매핑한다'의 출발점으로 허용한다. 마지막 헤더를
    토글로 끄는 개별 경로는 여전히 '하나 이상'."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_source", {"index": 0, "source": "낙찰금액"})
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": True})

    # 확정 존재 → 전체 미사용 차단(파괴 방지).
    with pytest.raises(ValueError, match="확정한 매핑이 있어"):
        ctrl.dispatch("use_none", {})
    snap = ctrl.snapshot()
    assert snap["rows"][0]["source"] == "낙찰금액" and snap["rows"][0]["confirmed"] is True
    assert snap["ignored_count"] == 0                                # 파괴 없음

    # 마지막 남은 헤더를 토글로 끄는 개별 경로는 '하나 이상'으로 차단.
    ctrl.dispatch("toggle_source_active", {"field": "업체명"})
    ctrl.dispatch("toggle_source_active", {"field": "계약일"})       # 활성=[낙찰금액]
    with pytest.raises(ValueError, match="하나 이상"):
        ctrl.dispatch("toggle_source_active", {"field": "낙찰금액"})

    # 확정 해제 후엔 전체 미사용 허용 + 미사용 구역 펼침(고르는 흐름 시작점).
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": False})
    ctrl.dispatch("use_none", {})
    snap = ctrl.snapshot()
    assert snap["active_count"] == 0 and snap["ignored_count"] == 3
    assert snap["ignored_expanded"] is True


def test_load_job_reedit_starts_all_active(tmp_path):
    """재편집 진입 = 활성 헤더가 저장 매핑에서 파생(#49 핵심 주장) — 미사용 0.

    실제 소스 매핑을 저작해 저장한 뒤 재로드하면 source_fields 가 저장 매핑의 소스 키로
    복원되고(profile_source_vocabulary) 전원 활성이다 — durable ignored 없이도 '매핑이
    곧 기억'이 성립함을 못박는다."""
    ctrl, _ = _controller26(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_source", {"index": 0, "source": "낙찰금액"})   # 실 소스 매핑
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": True})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("set_name", {"name": "재편집대상"})
    ctrl.dispatch("set_pattern", {"pattern": "p-{{수요기관}}"})
    assert ctrl.dispatch("save", {})["ok"] is True

    ctrl.load_job("재편집대상")
    snap = ctrl.snapshot()
    assert "낙찰금액" in snap["source_fields"]            # 저장 매핑 소스로 어휘 복원
    assert snap["ignored_count"] == 0                    # 전원 활성(미사용 0)
    assert snap["active_source_fields"] == snap["source_fields"]


# --------------------- (기본 데이터 연결 상태 재진술(#67)은 #347 에서 참조와 함께 사망 —
#  linked/dead/corrupt/missing 4태 스냅샷은 재진술할 default_dataset_ref 가 없어졌다.
#  스냅샷 키 소멸은 test_save_with_data_registers_nothing_anywhere 가 단언한다.)


# ------------------------------------------------ 에디터 흡수(블록 2 개정, 결정 39~41)
def test_data_gateway_opt_out_is_gone_loudly(tmp_path):
    """「데이터 없이 진행」(구 `skip_data`)은 표면째 사라졌다(#932 U4-C S2-4).

    조용한 no-op 잔존이 아니라 **미지 액션**이다: 데이터 결속이 저장 게이트가 된 이상
    옵트아웃은 저장할 수 없는 세션으로 데려가는 링크였다. 남겨 두면 프런트가 걷힌 뒤에도
    브리지로는 그 세션에 도달할 수 있다.
    """
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    with pytest.raises(ValueError, match="알 수 없는 editor 액션"):
        ctrl.dispatch("skip_data", {})


def test_save_is_blocked_until_data_is_connected(tmp_path):
    """작업은 데이터 결속 없이 저장되지 않는다(#932 U4-C S2-3).

    구 계약(「데이터 없이 진행」)의 자리에 서는 대체 계약이다: 상수만 쓰는 완결 매핑도
    결속이 없으면 차단되고, 차단은 **고칠 자리**(데이터 관문)를 함께 말한다. 결속이
    서면 같은 세션이 그대로 저장된다 — 막힌 것은 데이터 하나뿐임을 대조로 못박는다.
    """
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    # 1단계 게이트를 지나려면 데이터가 있어야 하므로(U6-B) 결속을 세운 뒤 **떼어** 본다 —
    # 저장 게이트의 술어가 그대로 서 있는지가 이 시험의 요점이고, 그 자리는 저장본 편집·
    # 인계 복원 실패처럼 결속이 사라진 세션에서 실제로 도달한다(심층 방어).
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.data_path = ""
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("set_name", {"name": "결속없는작업"})
    ctrl.dispatch("set_pattern", {"pattern": "p-{{수요기관}}"})

    res = ctrl.dispatch("save", {})
    assert res["ok"] is False
    assert "데이터를 연결" in res["block_reason"]
    assert res["blocked_field"] == "data"          # 고칠 자리를 함께 말한다(U2 §2.4)
    assert not JobRegistry(tmp_path / "jobs").exists("결속없는작업")

    _mount_data(ctrl)                              # 결속만 세운다(매핑은 그대로 확정 상태)
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    assert ctrl.dispatch("save", {})["ok"] is True


def test_saved_job_carries_the_data_binding(tmp_path):
    """저장된 작업이 결속 세 성분을 그대로 든다(#932 U4-C) — 경로만 남기지 않는다."""
    ctrl, _ = _controller(tmp_path)
    assert _save_named(ctrl, "결속작업")["ok"] is True
    job = JobRegistry(tmp_path / "jobs").load("결속작업")
    assert job.data_path == str(MULTI_SHEET)
    assert job.data_sheet == "낙찰현황"
    assert job.data_header_row == 0


def test_saved_job_carries_the_session_data_kind(tmp_path):
    """결속의 **종류**도 저장이 다시 짓는 성분이다 — 세션 값이 그대로 durable 로 간다.

    엑셀/CSV 마운트의 종류는 ``""`` 라 기본 동선에서는 눈에 띄지 않는다. 그래서 세션 값을
    직접 세워 **싣는 자리가 있는지**를 못박는다: 저장이 종류를 흘리면 다른 종류의 결속이
    디스크에서 엑셀로 되읽혀 어느 어댑터로 읽을지가 조용히 갈린다.
    """
    ctrl, _ = _controller(tmp_path)
    assert _save_named(ctrl, "종류없는작업")["ok"] is True
    registry = JobRegistry(tmp_path / "jobs")
    assert registry.load("종류없는작업").data_kind == ""    # 파일 마운트 = 엑셀/CSV

    ctrl2, _ = _controller(tmp_path)
    ctrl2.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl2)
    ctrl2.data_kind = "pclm"                               # 마운트 뒤 종류만 갈아 끼운다
    ctrl2.dispatch("goto_section", {"section": "binding"})
    ctrl2.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl2.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl2.dispatch("confirm_all", {})
    ctrl2.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl2.dispatch("set_name", {"name": "종류있는작업"})
    assert ctrl2.dispatch("save", {})["ok"] is True
    assert registry.load("종류있는작업").data_kind == "pclm"


def test_whole_session_discard_compares_all_four_binding_components(tmp_path):
    """버리기의 결속 비교는 **네 성분**이다 — 종류만 갈려도 「되돌렸다」가 참이어야 한다.

    경로·시트·헤더 행만 보면 종류가 갈린 세션은 「아무것도 안 바뀌었다」로 읽혀 문안이
    침묵하고, 화면에는 저장본과 다른 종류의 데이터가 서 있던 사실이 아무 데도 안 남는다.
    """
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "종류버리기")
    ctrl.dispatch("save", {})
    ctrl.load_job("종류버리기")
    assert ctrl.data_kind == ""
    ctrl.data_kind = "pclm"                       # 경로·시트·헤더는 그대로, 종류만 다르다
    ctrl.dispatch("discard_patch", {})            # section 없음 = 세션 전체
    assert ctrl.data_kind == ""                   # 저장본의 종류로 되돌아왔고
    assert "연결된 것으로 되돌렸습니다" in ctrl.notice_text   # 그 사실을 재진술한다


def test_save_landing_keeps_the_session_data(tmp_path):
    """저장 착지가 데이터를 내려놓지 않는다(#932 U4-C S2-1).

    구판은 착지 재로드에 결속을 안 넘겨 저장 한 번마다 `_reset()` 이 데이터·레코드를
    지우고, 소스 어휘가 저장 매핑이 참조하는 키 집합으로 강등돼 표·미리보기가 빈칸이 됐다.
    """
    ctrl, _ = _controller(tmp_path)
    assert _save_named(ctrl, "착지작업")["ok"] is True
    snap = ctrl.snapshot()
    assert snap["data_path"] == str(MULTI_SHEET)
    assert snap["data_sheet"] == "낙찰현황"
    assert snap["record_count"] == 3
    assert snap["source_fields"] == ["업체명", "낙찰금액", "계약일"]
    assert snap["sample_rows"]
    assert snap["notice"]["level"] == "ok"        # 착지 사유는 저장 성공 그대로


def test_editing_a_saved_job_remounts_its_binding(tmp_path):
    """저장본 편집 진입이 결속 데이터를 다시 세운다(#932 U4-C S2-2).

    진입 사유와 무관하다 — 「문서 작업」에서 그냥 연 세션도 데이터를 든 채 선다. 데이터가
    없으면 매핑 표가 어휘 없이 서고, 열을 바꾸러 온 사람이 고를 후보가 화면에 없다.
    """
    ctrl, _ = _controller(tmp_path)
    assert _save_named(ctrl, "재진입작업")["ok"] is True
    ctrl.dispatch("new_session", {})
    assert ctrl.snapshot()["data_path"] == ""     # 세션은 비었다(대조군)

    ctrl.load_job("재진입작업")
    snap = ctrl.snapshot()
    assert snap["data_path"] == str(MULTI_SHEET) and snap["record_count"] == 3
    assert snap["source_fields"] == ["업체명", "낙찰금액", "계약일"]
    assert ctrl.has_unsaved_work() is False       # 결속 재적재는 사람의 변경이 아니다


def test_entry_handoff_beats_the_saved_binding(tmp_path):
    """진입이 들고 온 참조가 저장본 결속을 이긴다(#932 U4-C S2-2) — 더 최신 의사다."""
    ctrl, _ = _controller(tmp_path)
    assert _save_named(ctrl, "인계작업")["ok"] is True
    ctrl.load_job(
        "인계작업",
        entry_reason="document_browser_repair",
        evidence={"데이터": "multi_sheet.xlsx"},
        return_context={"surface": "data"},
        source_ref={"path": str(MULTI_SHEET), "sheet": "공고목록", "header_row": 0},
    )
    snap = ctrl.snapshot()
    assert snap["data_sheet"] == "공고목록"       # 저장본의 낙찰현황이 아니다
    assert snap["source_fields"] == ["공고명", "추정가격"]


def test_a_job_without_a_binding_opens_quietly(tmp_path):
    """구판 작업(결속 없음)은 조용히 열린다(#932 U4-C S2-2) — 여기서 경보를 세우지 않는다.

    「데이터 연결 필요」는 저장 게이트의 문장이라 진입이 그것을 미리 말하면 같은 상태를
    두 곳이 판정한다. 진입은 기존 통지 채널(복원 재진술)을 그대로 쓴다.
    """
    ctrl, _ = _controller(tmp_path)
    assert _save_named(ctrl, "구판작업")["ok"] is True
    registry = JobRegistry(tmp_path / "jobs")
    legacy = registry.load("구판작업")
    legacy.data_path = ""
    legacy.data_sheet = ""
    legacy.data_header_row = 0
    registry.save(legacy, allow_overwrite=True)

    ctrl.dispatch("new_session", {})
    ctrl.load_job("구판작업")
    snap = ctrl.snapshot()
    assert snap["data_path"] == "" and snap["record_count"] == 0
    assert snap["notice"]["level"] == "ok"        # 드리프트 없는 복원은 그대로 ok
    assert "다시 읽지 못했습니다" not in snap["notice"]["text"]


def test_save_landing_restates_a_binding_that_cannot_be_reread(tmp_path):
    """착지 재로드 실패는 성공 문안에 덮이지 않는다(#932 U4-C S2-1).

    저장은 이미 커밋됐으므로 실패했다고 말하지 않되, 화면이 빈 데이터 관문인 채 "저장
    했습니다"만 말하게 두지도 않는다 — 한 통지가 두 사실을 함께 재진술한다.
    """
    moving = tmp_path / "옮길데이터.xlsx"
    moving.write_bytes(MULTI_SHEET.read_bytes())
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(moving), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("set_name", {"name": "사라진데이터작업"})
    ctrl.dispatch("set_pattern", {"pattern": "p-{{수요기관}}"})

    moving.unlink()                                # 저장 직전에 원본이 사라진다
    res = ctrl.dispatch("save", {})
    assert res["ok"] is True                       # 저장 자체는 성립했다
    notice = ctrl.snapshot()["notice"]
    assert notice["level"] == "warn"
    assert "저장했습니다" in notice["text"]
    assert "다시 읽지 못했습니다" in notice["text"]


def test_session_dirty_is_one_python_owned_value(tmp_path):
    """"이 세션이 잃을 것이 있는가"의 **단일 출처**(3R 근본 조치).

    표면이 탭 표지(`dirty_sections`)로 이 판정을 재조립하면 소비자마다 답이 갈린다 —
    실제로 2R 은 이탈 가드만 고쳤고 머리·footer 는 같은 상태를 「저장됨」이라 말했다.
    이름처럼 어느 section 에도 없는 편집(판정 L)이 여기서 갈린다.
    """
    ctrl, _ = _controller26(tmp_path)
    assert _save_named(ctrl, "단일출처")["ok"] is True
    ctrl.load_job("단일출처")
    snap = ctrl.snapshot()
    assert snap["dirty"] is False and snap["dirty_sections"] == []   # 복원 직후 = 저장됨

    ctrl.dispatch("set_name", {"name": "새 이름"})                    # section 밖 편집
    snap = ctrl.snapshot()
    assert snap["dirty_sections"] == []       # 탭 표지엔 안 뜨고
    assert snap["dirty"] is True              # 세션 수준으로는 잃을 것이 있다

    ctrl.dispatch("set_pattern", {"pattern": "다른-{{공고명}}"})
    snap = ctrl.snapshot()
    assert snap["dirty_sections"] == ["filename"] and snap["dirty"] is True


def test_discarding_one_section_keeps_edits_that_live_outside_sections(tmp_path):
    """탭 이동의 자동 버리기는 **그 자리만** 되돌린다(2R P2).

    되돌렸다고 알리는 것은 「그 탭에서 바꾼 것」인데 세션 전체를 되돌리면, 머리에서 고친
    이름처럼 **어느 section 에도 속하지 않는 편집**(§10.13 판정 L 계열)까지 함께 사라진다 —
    되돌리는 범위가 알린 문안보다 넓으면 그건 알린 적 없는 파기다.
    """
    ctrl, _ = _controller26(tmp_path)
    assert _save_named(ctrl, "부분되돌리기")["ok"] is True
    ctrl.load_job("부분되돌리기")
    ctrl.dispatch("set_name", {"name": "새 이름"})                 # section 밖(정체)
    ctrl.dispatch("goto_section", {"section": "filename"})
    ctrl.dispatch("set_pattern", {"pattern": "다른-{{공고명}}"})    # 파일 이름 patch
    assert ctrl.dirty_sections() == ("filename",)

    ctrl.dispatch("discard_patch", {"section": "filename"})
    assert ctrl.dirty_sections() == ()                             # 그 자리는 되돌아갔고
    assert ctrl.job_name == "새 이름"                              # 이름은 살아 있다
    assert ctrl.has_unsaved_work() is True                         # 그래서 아직 버릴 것이 남았다

    # 인자 없는 되돌리기(footer 「변경 버리기」·이탈의 자동 버리기)는 세션 전체가 대상이다.
    ctrl.dispatch("discard_patch", {})
    assert ctrl.job_name == "부분되돌리기" and ctrl.has_unsaved_work() is False


def test_partial_discard_keeps_dirty_while_session_data_is_unsaved(tmp_path):
    """되돌린 뒤 클린 표지는 **정말 남은 것이 없을 때만** 선다(5R P2).

    데이터 선택은 이제 저장되는 **결속**이라 더더욱 미저장 세션 상태다(#932 U4-C) —
    section 밖에 산다는 이유로 안 세면, 남아 있는 편집이 「저장됨」으로 위장하고 이탈이
    아무것도 묻지 않는다. 2R 이 이 줄을 세울 때 이름만 본 것이 연 창이다.

    **다른 시트를 고르는 이유**: 결속이 durable 이 된 뒤로 저장본과 같은 데이터를 다시
    고르는 것은 아무것도 바꾸지 않는다(진입이 이미 그 결속을 세워 둔다). 미저장을 세려면
    실제로 갈리는 선택이어야 한다 — 같은 통합문서의 다른 시트는 다른 데이터다(#33).
    """
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "표지정직")
    ctrl.dispatch("save", {})
    ctrl.load_job("표지정직")
    assert ctrl.snapshot()["dirty"] is False           # 복원 직후 = 저장됨
    assert ctrl.data_sheet == "낙찰현황"                # 진입이 저장된 결속을 세웠다
    ctrl.load_data_path(str(MULTI_SHEET), sheet="공고목록")  # 세션이 다른 데이터를 골랐다
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": False})
    ctrl.dispatch("discard_patch", {"section": "binding"})
    assert ctrl.dirty_sections() == ()                 # 그 자리는 되돌아갔지만
    assert ctrl.snapshot()["dirty"] is True            # 데이터 선택은 아직 미저장이다


def test_partial_discard_refuses_a_tab_this_media_does_not_have(tmp_path):
    """없는 자리를 되돌리라는 요청은 조용히 무시하지 않는다 — 어느 자리를 되돌렸는지가
    문안의 약속이라, 모르는 자리를 받아 넘기면 그 약속을 지켰는지 말할 수 없다."""
    ctrl, _ = _controller26(tmp_path)
    assert _save_named(ctrl, "없는탭")["ok"] is True
    ctrl.load_job("없는탭")
    with pytest.raises(ValueError, match="탭이 없습니다"):
        ctrl.dispatch("discard_patch", {"section": "test"})   # 시험 탭은 기각(§10.17.1) — 영구 미지 section


def test_discarding_a_binding_patch_without_data_restores_the_saved_vocabulary(tmp_path):
    """데이터 없는 편집 세션의 되돌리기는 **저장 매핑의 어휘**로 되세운다.

    어휘를 안 되세우면 되돌린 행이 전부 "(데이터에 없음)"으로 오표시된다 — 되돌렸는데
    화면이 더 나빠 보이면 그 되돌리기는 사용자가 요청한 것이 아니다(load_job 과 동형).
    """
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "어휘복원")
    ctrl.dispatch("save", {})
    ctrl.load_job("어휘복원")                       # 데이터 없이 복원(저장 매핑 어휘로 선다)
    vocabulary = list(ctrl.source_fields)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": False})
    ctrl.dispatch("discard_patch", {"section": "binding"})
    assert ctrl.source_fields == vocabulary and ctrl.dirty_sections() == ()
    assert all(r.confirmed for r in ctrl.model.rows)          # 저장본 그대로 확정 복원


def test_discarding_a_template_patch_keeps_the_name_and_data(tmp_path):
    """템플릿 축은 스키마·매핑이 함께 서야 해 규칙 전체를 다시 세운다 — 그래도 **이름과
    데이터는 남는다**(section 밖에 사는 것은 이 처분의 대상이 아니다)."""
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "템플릿되돌리기")
    ctrl.dispatch("save", {})
    ctrl.load_job("템플릿되돌리기")
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("set_name", {"name": "새 이름"})
    base_template = ctrl.template_path
    ctrl.dispatch("goto_section", {"section": "template"})
    ctrl.load_template_path(str(TPL_PARTIAL))                 # 다른 템플릿으로 갈아 끼움
    assert "template" in ctrl.dirty_sections()
    ctrl.dispatch("discard_patch", {"section": "template"})
    assert ctrl.template_path == base_template                # 템플릿은 되돌아갔고
    assert ctrl.job_name == "새 이름" and ctrl.data_path      # 이름·데이터는 그대로


def test_discarding_a_binding_patch_keeps_the_loaded_data(tmp_path):
    """연결 patch 를 되돌려도 사람이 고른 데이터는 내려놓지 않는다(판정 L).

    데이터 선택은 patch 가 아니라 세션 문맥이다 — 규칙을 되돌린다고 엑셀까지 걷으면
    사용자는 되돌리기 한 번에 관문부터 다시 밟아야 한다.
    """
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "데이터유지")
    ctrl.dispatch("save", {})
    ctrl.load_job("데이터유지")
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    before_rows = len(ctrl.records)
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": False})
    assert ctrl.dirty_sections() == ("binding",)
    ctrl.dispatch("discard_patch", {"section": "binding"})
    assert ctrl.dirty_sections() == ()
    assert ctrl.data_path and len(ctrl.records) == before_rows     # 데이터는 그대로


def test_a_whole_session_discard_returns_the_data_to_the_saved_binding(tmp_path):
    """세션 전체 버리기는 데이터도 **저장된 상태로 되돌린다**(#932 U4-C — 8R P2 승계).

    탭 단위 버리기와 대칭을 이루는 반대쪽이다: 그쪽은 「이 탭에서 바꾼 것만」이라 말하므로
    세션 데이터를 남기고, 이쪽은 「저장된 상태로 되돌린다」고 말하므로 데이터도 그 말을
    따른다. 남기면 버리기를 마친 세션이 여전히 미저장이라 다음 작업을 열 때 방금 버린 것을
    또 묻는다.

    **되돌아가는 자리가 「빈 값」에서 「저장된 결속」으로 갈렸다.** 결속이 durable 이 되기
    전에는 저장본에 데이터가 없어 되돌림이 곧 비움이었다. 지금 비우면 그것은 되돌림이
    아니라 저장본을 넘어선 파기다 — 문안도 함께 갈린다(「내려놨습니다」는 화면에 데이터가
    서 있는 동안 거짓이다).
    """
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "전체버리기")
    ctrl.dispatch("save", {})
    ctrl.load_job("전체버리기")
    ctrl.load_data_path(str(MULTI_SHEET), sheet="공고목록")   # 저장본과 다른 데이터
    ctrl.dispatch("set_name", {"name": "새 이름"})
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": False})
    assert ctrl.has_unsaved_work() is True
    ctrl.dispatch("discard_patch", {})                       # section 없음 = 세션 전체
    assert ctrl.dirty_sections() == () and ctrl.dirty_extras() == ()
    assert ctrl.job_name == "전체버리기"                      # 이름도 저장본으로
    # 데이터도 저장본으로 — 비우는 것이 아니라 연결된 자리로 되돌아간다.
    assert ctrl.data_path == str(MULTI_SHEET) and ctrl.data_sheet == "낙찰현황"
    assert ctrl.records                                      # 되돌린 데이터는 실제로 읽혔다
    # 버린 뒤에는 잃을 것이 없다 — 다음 전환·새 작업이 같은 파기를 두 번 묻지 않는다.
    assert ctrl.has_unsaved_work() is False
    assert "연결된 것으로 되돌렸습니다" in ctrl.notice_text    # 무엇이 되돌아갔는지 재진술


def test_every_session_extra_counts_as_unsaved_work(tmp_path):
    """section 밖 세션 상태의 **열거를 순회해** 판정을 센다(8R 근본 조치).

    F7 리뷰는 같은 결함을 라운드마다 한 값씩 재발견했다(2R 이름 → 3R 자동등록 이름 →
    5R 데이터 선택). 판정이 틀렸던 것이 아니라 열거가 판정마다 손으로 다시 쓰여 있었다 —
    이 테스트는 상수 하나를 순회하므로 목록에 값이 늘면 커버리지가 함께 늘고, 판정이 그
    목록에서 파생되지 않으면 즉시 붉어진다.
    """
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "열거순회")
    ctrl.dispatch("save", {})
    assert ctrl.SESSION_EXTRAS, "section 밖 세션 상태의 열거가 비었습니다."
    # 초안은 비교 대상(base)이 없어 extras 판정이 성립하지 않는다 — 초안 전체의 미저장은
    # 세션 폐기 확인이 따로 지킨다(판정 P). 빈 튜플이어야 「저장본과 다르다」를 참칭하지 않는다.
    draft, _ = _controller26(tmp_path / "draft")
    draft.load_template_path(str(TPL_COMPILED))
    draft.dispatch("set_name", {"name": "초안 이름"})
    assert draft.dirty_extras() == () and draft.has_unsaved_work() is True
    for extra in ctrl.SESSION_EXTRAS:
        ctrl.load_job("열거순회")                             # 매번 깨끗한 세션에서 시작
        assert ctrl.has_unsaved_work() is False, f"{extra}: 복원 직후가 미저장으로 보입니다."
        setattr(ctrl, extra, "손댄 값")
        assert ctrl.dirty_extras() == (extra,), f"{extra}: 열거가 이 값을 세지 않습니다."
        assert ctrl.has_unsaved_work() is True, (
            f"{extra}: section 밖 편집이 「저장됨」으로 위장합니다 — 이탈이 조용히 버립니다."
        )


def test_switching_only_the_sheet_is_unsaved_work(tmp_path):
    """같은 엑셀의 **다른 시트**로 갈아타는 것도 미저장이다 — 열거를 세우자 드러난 자리.

    경로는 그대로이므로 「데이터를 골랐는가」만 보는 판정에는 안 걸린다. 그런데 같은
    워크북의 다른 시트는 다른 데이터라(#33 — §5.3 정체성 축에도 시트가 든다), 그
    갈아타기를 놓치면 사람이 시트를 바꾸고 나갈 때 아무것도 묻지 않고 버린다 — 2R~5R 이
    이름·데이터로 겪은 것과 **같은 결함의 다른 인스턴스**다.
    """
    ctrl, _ = _controller26(tmp_path)
    _complete_with_data(ctrl, "시트갈아타기")
    ctrl.dispatch("save", {})
    ctrl.load_job("시트갈아타기")
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    path_then = ctrl.data_path
    ctrl.dispatch("discard_patch", {})                        # 데이터까지 내려놓고 다시 시작
    ctrl.load_data_path(str(MULTI_SHEET), sheet="공고목록")
    assert ctrl.data_path == path_then                        # 경로는 그대로인데
    assert "data_sheet" in ctrl.dirty_extras()                # 시트는 갈렸다
    assert ctrl.has_unsaved_work() is True


def test_unsaved_work_is_derived_not_flagged(tmp_path):
    """저장된 작업의 미저장 판정은 **파생**이다 — 손으로 켠 클린 표지가 이를 덮지 못한다.

    표지 방식은 변이 자리·되돌리기 자리가 늘 때마다 한 곳이 빠졌고, 빠짐은 「저장됨」이라는
    거짓말이나 되돌린 뒤의 헛확인 둘 중 하나로 나타났다. 파생은 빠질 자리가 없다: 여기서
    표지를 거짓으로 세워 두고도 판정이 patch 를 보는지 확인한다(양방향).
    """
    ctrl, _ = _controller26(tmp_path)
    assert _save_named(ctrl, "파생판정")["ok"] is True
    ctrl.load_job("파생판정")
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": False})
    ctrl._session_clean = True                                # 표지를 거짓으로 세운다
    assert ctrl.has_unsaved_work() is True, "표지가 실재하는 patch 를 덮었습니다."
    ctrl.dispatch("discard_patch", {})
    ctrl._session_clean = False                               # 반대 방향도 표지 무관
    assert ctrl.has_unsaved_work() is False, "되돌린 뒤에도 헛확인을 묻습니다(과경고)."


def test_editing_tabs_move_freely_and_autodiscard_the_blocking_patch(tmp_path):
    """편집 탭은 자유 이동하고(결정 41), **막는 patch 는 묻지 않고 버린다**.

    계약 §5.2·§13-16: 한 편집 진입은 한 section patch 만 가진다 — 다른 탭의 규칙을 손대려면
    지금 것을 처분해야 한다. 종전엔 그 처분을 3택 모달로 물었지만 편집기 한 탭의 작업량은
    확인을 요구할 만큼 크지 않아, 지금은 컨트롤러가 그 자리만 되돌리고 지나간다.

    **조용히 지나가지는 않는다**: 되돌린 사실은 notice 로 재진술되고(확인 대신 알림),
    범위는 종전 「버리고 이동」과 같아 어느 section 에도 속하지 않는 편집(이름)은 살아남는다.
    신규 초안(편집 원점 없음)은 대조군: 거래 밖이지만 전진 게이트가 산다.
    """
    ctrl, _ = _controller26(tmp_path)
    assert _save_named(ctrl, "자유이동")["ok"] is True
    ctrl.load_job("자유이동")
    ctrl.dispatch("goto_section", {"section": "filename"})   # 깨끗한 세션 = 자유 이동
    assert ctrl.section == "filename"
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_name", {"name": "이동해도 사는 이름"})           # section 밖 편집
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": False})   # 연결 patch 발생
    assert ctrl.dirty_sections() == ("binding",)

    assert ctrl.dispatch("goto_section", {"section": "filename"}) is None
    assert ctrl.section == "filename"                         # 막히지 않는다
    assert ctrl.dirty_sections() == ()                        # 막던 자리는 되돌아갔고
    assert ctrl.job_name == "이동해도 사는 이름"                # section 밖 편집은 살아남는다
    notice = ctrl.snapshot()["notice"]
    assert notice and "「연결 확인」" in notice["text"], (
        f"자동으로 버려 놓고 아무 말도 하지 않았습니다: {notice!r}"
    )

    ctrl2, _ = _controller(tmp_path / "new")             # 대조군: 신규 마법사
    ctrl2.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl2)                                                  # 1단계 게이트(U6-B)
    ctrl2.dispatch("goto_section", {"section": "binding"})             # 고르기→연결은 짝이 서야 통과
    with pytest.raises(ValueError, match="조건을 아직 채우지 못해"):
        ctrl2.dispatch("goto_section", {"section": "filename"})         # 매핑 미확정 → 전진 차단


def test_discarding_a_clean_session_is_a_silent_no_op(tmp_path):
    """손대지 않은 세션의 세션 전체 되돌리기는 **아무 일도 하지 않는다**.

    이탈이 확인 없이 `discard_patch {}` 를 무조건 부르게 되면서 생긴 자리다: 판정을 웹에
    두면 같은 상태를 두 곳이 답하므로 게이트를 여기 하나로 뒀다. 게이트가 없으면 클린
    이탈마다 디스크를 다시 읽고 「되돌렸습니다」라는 거짓 통지가 선다(과진술도 부정직이다).
    """
    ctrl, _ = _controller26(tmp_path)
    assert _save_named(ctrl, "클린이탈")["ok"] is True
    ctrl.load_job("클린이탈")
    ctrl.dispatch("dismiss_notice", {})
    assert ctrl.has_unsaved_work() is False

    ctrl.dispatch("discard_patch", {})
    assert ctrl.snapshot()["notice"] is None, "버릴 것이 없는데 버렸다고 말합니다."
    assert ctrl.has_unsaved_work() is False

    ctrl.dispatch("set_name", {"name": "손댐"})               # 양성 대조 — 손댄 뒤엔 말한다
    ctrl.dispatch("discard_patch", {})
    notice = ctrl.snapshot()["notice"]
    assert notice and "되돌렸습니다" in notice["text"]
    assert ctrl.job_name == "클린이탈"


# ---------------------------------------- PR-2 고효율 리뷰 반영(파괴 경로·클린 세션·판정 위치)
def test_save_lands_in_edit_session_of_saved_job(tmp_path):
    """저장 착지 = 방금 저장한 작업의 편집 세션(리뷰 F2 — 빈 마법사 방치·성공 표지 증발 봉합).

    결정 40(저장 제자리)·41(전환점=저장: 초안은 저장으로 작업이 되고 이후 편집은 탭)의 이행.
    성공 재진술은 push 경합에 안 걸리는 notice(ok) 채널로 온다."""
    ctrl, _ = _controller26(tmp_path)
    res = _save_named(ctrl, "착지작업")
    assert res["ok"] is True
    snap = ctrl.snapshot()
    assert snap["editing_origin"] == "착지작업"          # 빈 마법사가 아니라 저장본 위
    assert snap["section"] == "binding" and snap["is_complete"] is True
    assert snap["notice"] and "저장했습니다" in snap["notice"]["text"]
    assert snap["notice"]["level"] == "ok"
    assert ctrl.has_unsaved_work() is False              # 클린 착지 — 직후 전환 헛확인 금지


def test_new_hwpx_save_from_filename_tab_lands_in_place(tmp_path):
    """U2 §2.14 — 실 UI 순서(filename 까지 전진 후 저장)의 신규 hwpx 저장은 **제자리** 착지.

    구판은 신규 세션을 binding 으로 내려 3→2 로 뒤로 갔다(`_save_named` 는 binding 에서
    저장해 이 자리를 안 밟았다 — 조치 전후로 초록인 헬퍼 경로는 위 테스트가 계속 진다).
    착지는 여전히 저장본 편집 세션이다(원점·클린·notice(ok))."""
    ctrl, _ = _controller26(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("goto_section", {"section": "filename"})   # 실 UI: 3단계까지 전진
    ctrl.dispatch("set_name", {"name": "전진저장작업"})
    ctrl.dispatch("set_pattern", {"pattern": "p-{{수요기관}}"})
    assert ctrl.dispatch("save", {})["ok"] is True
    snap = ctrl.snapshot()
    assert snap["section"] == "filename"                     # 제자리 — 뒤로 가지 않는다
    assert snap["editing_origin"] == "전진저장작업"          # 저장본 편집 세션 착지는 그대로
    assert snap["notice"] and snap["notice"]["level"] == "ok"
    assert ctrl.has_unsaved_work() is False


def test_edit_save_preserves_current_tab(tmp_path):
    """편집 저장은 현재 탭을 유지하고 최종 상태만 한 번 렌더한다."""
    ctrl, pushes = _controller26(tmp_path)
    assert _save_named(ctrl, "탭유지작업")["ok"] is True
    ctrl.dispatch("goto_section", {"section": "filename"})             # 작업 저장 탭에서 저장
    before = len(pushes)

    assert ctrl.dispatch("save", {})["ok"] is True

    snap = ctrl.snapshot()
    assert snap["section"] == "filename"
    assert pushes[-1][1]["section"] == "filename"                   # 웹에 전달된 최종 활성 탭도 동일
    assert len(pushes) == before + 1                     # 중간 재로드 렌더 없이 최종 push 1회


def test_load_job_marks_session_clean_until_edited(tmp_path):
    """편집 복원 직후는 클린(디스크 저장본과 동일) — 손대기 전 전환·새 작업이 "저장하지 않은
    세션" 헛확인을 띄우지 않는다(리뷰). 변이 액션 하나로 다시 미저장이 된다."""
    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "클린작업")
    ctrl.load_job("클린작업")
    assert ctrl.has_unsaved_work() is False              # 복원 직후 = 버릴 것 없음
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": False})
    assert ctrl.has_unsaved_work() is True               # 변이 → 미저장


def test_partial_template_saves_once_acked_and_bound(tmp_path):
    """PARTIAL 템플릿도 게이트 확인 + 데이터 결속이면 저장된다(구 리뷰 F6 의 승계 계약).

    구판은 이 자리를 「데이터 없이 진행」으로 밟았다. 옵트아웃이 사라진 뒤로는 관문의
    데이터 선택이 그 자리를 잇고, 게이트 확인이 세션 국소라 저장 착지에서 미확인으로
    돌아온다는 사실(재로드 = 저장본 기준)은 그대로다.
    """
    ctrl, _ = _controller26(tmp_path)
    ctrl.load_template_path(str(TPL_PARTIAL))
    ctrl.dispatch("ack_gate", {})
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("set_name", {"name": "부분템플릿작업"})
    assert ctrl.dispatch("save", {})["ok"] is True       # 저장 착지 = 편집 세션(게이트 미확인 복원)
    assert ctrl.snapshot()["gate"]["acked"] is False


def test_mapping_reset_stakes_judged_by_python_now(tmp_path):
    """관문 파괴 확인의 근거 수치는 Python 이 지금 판정(리뷰 F7 — stale LAST 우회 차단).

    수치 = 이월 대상(확정 + 내용 있는 touched) — _ensure_model carry 와 같은 집합이라
    확인 문안("값은 이월")과 실제 이월이 어긋나지 않는다(리뷰 F1)."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl)                                                  # 1단계 게이트(U6-B)
    assert ctrl.dispatch("mapping_reset_stakes", {})["human"] == 0     # 모델 전
    ctrl.dispatch("goto_section", {"section": "binding"})
    assert ctrl.dispatch("mapping_reset_stakes", {})["human"] == 0     # 미접촉 제안뿐
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    stakes = ctrl.dispatch("mapping_reset_stakes", {})
    assert stakes["human"] == 1                                        # 내용 있는 수동
    # 소스 없는 수동 const 행은 use_none 강등 대상이 아니다 — 문안=파괴 집합(리뷰 F4).
    assert stakes["use_none_manual"] == 0
    # 같은 행이 일괄 재제안에서는 **잃을 것이 있다**(리뷰 R1 P1) — reset_to_system 이 상수를
    # 지운다. 두 관문의 수치가 갈리는 자리라 이름도 소비자별로 갈라 둔다.
    assert stakes["resuggest_manual"] == 1
    assert stakes["confirmed"] == 0                                    # use_none 선차단 근거(F5)
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    stakes = ctrl.dispatch("mapping_reset_stakes", {})
    assert stakes["human"] == ctrl.snapshot()["field_count"]           # 전 행 확정(비움 포함)
    assert stakes["use_none_manual"] == 0                              # 확정 = 미확정 수동 아님
    assert stakes["resuggest_manual"] == 0                             # 확정 행은 재제안 비대상
    assert stakes["confirmed"] == ctrl.snapshot()["field_count"]       # 선차단 수치(F5)


def test_resuggest_stakes_count_every_row_the_loop_resets(tmp_path):
    """일괄 재제안의 확인 수치 = **그 루프가 실제로 리셋하는 행**(리뷰 R1 P1).

    종전엔 use_none 의 수치(`r.source` 를 요구)를 빌려 썼다. 소스 없이 상수만 직접 입력한
    미확정 행은 그 술어에 안 걸려 수치가 0이 되는데, 루프는 그 행도 `revert_to_auto` 로
    리셋한다(const·type·fmt 소거) — 확인 대화 없이 직접 입력이 사라졌다. 수치와 루프가
    같은 술어(`_resuggest_targets`)에서 나오는지 확인한다: 확인 수치 ≥ 실제 잃는 행.
    """
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    # 낙찰현황의 열 셋은 이 템플릿 필드와 하나도 겹치지 않아 자동 제안이 서지 않는다 —
    # 결속만 세우고(1단계 게이트) 「소스를 겨눌 수 없다」는 전제는 그대로 산다.
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "직접 입력한 값"})
    stakes = ctrl.dispatch("mapping_reset_stakes", {})
    assert stakes["use_none_manual"] == 0                 # use_none 은 이 행을 안 건드린다
    assert stakes["resuggest_manual"] == 1                # 재제안은 건드린다 → 확인 근거가 선다

    ctrl.dispatch("resuggest_all", {})
    row = ctrl.snapshot()["rows"][0]
    assert row["const"] == ""                             # 실제로 지운다(그래서 물어야 한다)
    assert row["touched"] is False


def test_ensure_model_carries_touched_unconfirmed_rows(tmp_path):
    """관문 재겨눔이 미확정 수동 편집을 이월한다(리뷰 F1 — carry_profile 실배선).

    확정-전용 이월(to_profile)은 "값은 이월된다"는 확인 문안과 달리 직접 고른 상수를
    조용히 버렸다 — 확정 0·수동 1 세션에서 데이터를 겨눠도 값이 남아야 한다."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_other_data(ctrl)                                            # 1단계 게이트(U6-B)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "수동값"})        # touched·미확정
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")            # 관문 겨눔 = 재초안
    row0 = ctrl.snapshot()["rows"][0]
    assert row0["type"] == "const" and row0["const"] == "수동값"       # 값 이월(소실 금지)
    assert row0["confirmed"] is False                                  # 재검토 강제는 유지


def test_gateway_repick_preserves_touched_unconfirmed_edits(tmp_path):
    """칩-라이브 리뷰 F2 정본(컨트롤러 end-to-end) — 미확정 **수동** 편집(touched)은 관문
    데이터 재겨눔에도 조용히 소실되지 않는다.

    carry_profile 이 확정뿐 아니라 touched 미확정 행도 이월(confirm=False)한다 — 값은 살고
    전 행 미확정으로 재검토를 강제(결정 12 '수동=사람 소유'). 구 to_profile(확정-only)이면
    이 수동 편집은 재초안에서 조용히 사라졌다(F2). 미접촉 제안은 반대로 새 데이터 재제안."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_source", {"index": 0, "source": "낙찰금액"})   # 수동(touched)·미확정
    snap = ctrl.snapshot()
    assert snap["rows"][0]["touched"] is True and snap["rows"][0]["confirmed"] is False

    ctrl.load_data_path(str(MULTI_SHEET))                            # 관문에서 첫 시트로 재겨눔
    snap = ctrl.snapshot()
    assert snap["rows"][0]["source"] == "낙찰금액"                   # 수동 편집 이월(F2 — 소실 아님)
    assert snap["rows"][0]["touched"] is True                       # 사람 소유 유지
    assert snap["rows"][0]["confirmed"] is False                    # 재검토 강제(전 행 미확정)


def test_revert_source_resets_single_row_and_resuggests(tmp_path):
    """↩(자동 제안 복귀, 결정 12) — 그 행만 완전 리셋 후 단일 행 재제안(리뷰 R4).

    무관한 stale 사람 소유 행(비활성 소스 겨눔)은 건드리지 않는다 — 전집합 재계산이면
    조용히 강등됐다. 센티넬 소스값이 아니라 전용 액션이라 동명 실열과도 안 충돌한다(R5)."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_source", {"index": 0, "source": "계약일"})     # 수동 오지정(touched)
    ctrl.dispatch("set_source", {"index": 1, "source": "없는열"})     # 무관 stale 사람 소유
    ctrl.dispatch("revert_source", {"index": 0})
    snap = ctrl.snapshot()
    assert snap["rows"][0]["touched"] is False                       # 시스템 소유 복귀
    assert snap["rows"][1]["source"] == "없는열"                     # 무관 행 불건드림(R4)
    assert snap["rows"][1]["touched"] is True


def test_chip_toggle_leaves_carried_stale_rows_untouched(tmp_path):
    """무관한 칩 조작이 이월 stale 행(현재 데이터에 없는 소스)을 강등하지 않는다(PR-3 리뷰 F1).

    관문 재겨눔이 carry 로 살린 「데이터에 없음」 행은 칩과 무관 — 전집합 강등이면 칩 토글
    한 번에 이월 값이 소실되고 통지는 끈 적 없는 헤더를 지목했다(오귀속)."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_source", {"index": 0, "source": "낙찰금액"})   # 수동
    ctrl.load_data_path(str(MULTI_SHEET))                            # 첫 시트 재겨눔 — carry
    assert ctrl.snapshot()["rows"][0]["source"] == "낙찰금액"         # stale 이월(「데이터에 없음」)
    ctrl.dispatch("toggle_source_active", {"field": "추정가격"})      # 무관 칩 끔
    snap = ctrl.snapshot()
    assert snap["rows"][0]["source"] == "낙찰금액"                    # 이월 값 생존(F1)
    assert snap["rows"][0]["touched"] is True
    assert "낙찰금액" not in (snap["notice"]["text"] if snap["notice"] else "")  # 오귀속 통지 없음


def test_revert_source_refuses_confirmed_rows(tmp_path):
    """↩ 는 확정 행을 거부한다(PR-3 리뷰 F2) — 확정도 touched 라 무가드면 오클릭 한 번에
    확정이 조용히 풀리고 다른 열로 치환된다. 확정 해제(체크박스)가 의식적 1단계."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_source", {"index": 0, "source": "낙찰금액"})
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": True})
    with pytest.raises(ValueError, match="확정을 먼저 해제"):
        ctrl.dispatch("revert_source", {"index": 0})
    assert ctrl.snapshot()["rows"][0]["confirmed"] is True            # 무파괴


def test_resuggest_all_reverts_every_unconfirmed_row(tmp_path):
    """일괄 재제안(U2 §2.4) — 행 단위 ↩ 의 일괄판. 착지가 행 단위와 **같아야** 한다.

    「일괄로 한 것」과 「하나씩 N번 한 것」이 다르면 사용자는 둘 중 어느 쪽이 진짜인지
    알 수 없다 — 그래서 전집합 `apply_active_sources` 가 아니라 행마다
    `revert_to_auto` → `resuggest_row` 로 간다.
    """
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_source", {"index": 0, "source": "계약일"})     # 수동 오지정
    ctrl.dispatch("set_source", {"index": 1, "source": "없는열"})     # 또 다른 수동
    res = ctrl.dispatch("resuggest_all", {})
    snap = ctrl.snapshot()
    assert res["kept_confirmed"] == 0
    assert res["resuggested"] == len(snap["rows"])
    # 행 단위 ↩ 와 같은 착지 — 전부 시스템 소유로 돌아간다.
    assert all(r["touched"] is False for r in snap["rows"])
    assert snap["rows"][1]["source"] != "없는열"


def test_resuggest_all_keeps_confirmed_rows_and_says_so(tmp_path):
    """확정 행은 **거절이 아니라 제외**다(U2 §2.4) — 그리고 제외했다고 수치로 말한다.

    행 단위 ↩ 는 확정 행을 시끄럽게 거절하는데(오클릭 한 번에 확정이 풀리면 안 된다),
    일괄에서 같은 규칙을 쓰면 확정 하나 때문에 나머지 전부를 못 돌려 「확정을 풀었다 다시
    건다」는 우회를 시킨다. 대신 건드린 수와 둔 수를 함께 돌려준다 — 부분 동작을 조용히
    하지 않는다.
    """
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_source", {"index": 0, "source": "낙찰금액"})
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": True})
    total = len(ctrl.snapshot()["rows"])
    res = ctrl.dispatch("resuggest_all", {})
    snap = ctrl.snapshot()
    assert res == {"resuggested": total - 1, "kept_confirmed": 1}
    assert snap["rows"][0]["confirmed"] is True                       # 무파괴
    assert snap["rows"][0]["source"] == "낙찰금액"


def test_resuggest_all_reports_zero_when_everything_is_confirmed(tmp_path):
    """대상이 0개면 0을 돌려준다 — 표면이 「무동작」을 말할 근거다(조용한 소실 금지)."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    # `confirm_all` 액션은 **내용 있는 행만** 확정한다(confirm_content_rows) — 전 행 확정
    # 상태를 만들려면 행마다 명시해야 한다.
    total = len(ctrl.snapshot()["rows"])
    for index in range(total):
        ctrl.dispatch("set_confirmed", {"index": index, "confirmed": True})
    assert ctrl.dispatch("resuggest_all", {}) == {
        "resuggested": 0, "kept_confirmed": total,
    }


def test_same_file_repick_after_use_none_revives_suggestions(tmp_path):
    """use_none 뒤 같은 파일 재겨눔(키 불변) — 관문 재동기화로 제안이 되살아난다(PR-3 리뷰 F3).

    load_data_path 가 칩 상태만 전원 활성으로 리셋하고 모델 키가 그대로면 재초안이 없어,
    「후보 없음」 죽은 제안이 조용히 남았다 — 키 불변이면 apply_active_sources 재동기화."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET))                             # 공고목록: 공고명·추정가격 매치
    ctrl.dispatch("goto_section", {"section": "binding"})
    assert any(r["source"] for r in ctrl.snapshot()["rows"])          # 자동 제안 존재(전제)
    ctrl.dispatch("use_none", {})                                     # 확정 0 — 허용
    assert all(not r["source"] for r in ctrl.snapshot()["rows"])      # 전원 후보 없음
    ctrl.load_data_path(str(MULTI_SHEET))                             # 같은 파일·시트 재겨눔(키 불변)
    snap = ctrl.snapshot()
    assert snap["active_count"] == 2
    assert any(r["source"] for r in snap["rows"])                     # 제안 부활(죽은 표면 아님)


def test_toggle_clears_ignored_expanded_hint(tmp_path):
    """개별 토글은 '전체 미사용' 펼침 힌트를 걷는다(PR-3 리뷰 F7) — 몇 步 전 행동의 stale
    상태가 이후 접힘 렌더를 계속 강제하지 않는다(수동 펼침 보존은 뷰 foldOpen 소관)."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("use_none", {})
    assert ctrl.snapshot()["ignored_expanded"] is True
    ctrl.dispatch("toggle_source_active", {"field": "업체명"})        # 다시 사용(개별)
    assert ctrl.snapshot()["ignored_expanded"] is False


# ---------------------------------- 신규 1단계 = 템플릿 라이브러리(R-info 2부 접합, PR-4)
def _controller_lib(tmp_path, paths=None, lib_dir=None):
    pushes: list = []
    vm = (
        TemplateManagerViewModel(
            lib_dir,
            inspect_template=inspect_hwpx_template,
            file_ops=HWPX_TEMPLATE_OPS,
        )
        if lib_dir is not None
        else TemplateManagerViewModel(
            paths=paths or [],
            inspect_template=inspect_hwpx_template,
            file_ops=HWPX_TEMPLATE_OPS,
        )
    )
    ctrl = EditorController(
        JobRegistry(tmp_path / "jobs"),
        lambda s, snap: pushes.append((s, snap)),
        clock=_clock,
        template_library=vm,
        text_registry=TextTemplateRegistry(tmp_path / "text_templates"),
    )
    return ctrl, pushes


def _lib_rows(ctrl, media="hwpx"):
    """좌 열이 실제로 읽는 목록 — **`tpl` 채널 행 성형**과 같은 술어(U6-B #976).

    편집기 스냅샷의 구 `library` 존은 퇴역했다: 목록 정본이 하나가 됐으므로 여기서는 그
    정본이 세우는 것과 같은 VM 행을 읽는다(스냅샷 계약은 `test_webapp_template` 소관).
    """
    return ctrl.template_library.rows() if media == "hwpx" else list(
        ctrl.text_registry.list_templates()
    )


def test_editor_snapshot_drops_the_library_zone_and_carries_pairing(tmp_path):
    """구 `library` 존은 퇴역하고 스냅샷은 **선택 경로 + 연결 카드**만 낸다(U6-B #976).

    목록을 두 컨트롤러가 그리던 자리를 지운다: 좌 열은 `tpl` 채널, 우 열은 `pool` 채널이
    정본이고 편집기는 「무엇이 골라졌나」와 그 짝의 수치만 답한다. 존이 되살아나면 같은
    목록이 두 스냅샷에 실린다 — 그래서 부재를 음성으로 못박는다.
    """
    ctrl, _ = _controller_lib(tmp_path, paths=[TPL_COMPILED, TPL_PARTIAL])
    snap = ctrl.snapshot()
    assert "library" not in snap
    assert snap["pairing"] == {
        "ready": False, "template_name": "", "data_name": "",
        "field_count": 0, "column_count": 0, "auto_count": 0, "confirm_count": 0,
        "basis": "", "advance_block_reason": "왼쪽에서 템플릿을 고르세요.",
    }
    ctrl.dispatch("use_library_template", {"path": str(TPL_COMPILED)})
    snap = ctrl.snapshot()
    assert snap["template_path"] == str(TPL_COMPILED)   # 선택 경로 하나가 좌 열의 `aria-pressed`
    assert snap["pairing"]["template_name"] == TPL_COMPILED.name
    assert snap["pairing"]["advance_block_reason"] == "오른쪽에서 데이터를 고르세요."


def test_pairing_counts_are_a_readonly_preview_until_the_model_exists(tmp_path):
    """연결 카드 수치의 **출처**를 명시로 든다 — `preview`(순수 함수) ↔ `model`(실제 행).

    1단계는 매핑 모델을 만들지 않는다: 만들면 고르기를 바꿔 보는 것만으로 「전원 미확정
    재생성」 전이가 돌아 확정이 조용히 무너진다. 그래서 모델이 없거나 키가 다르면
    :func:`~hwpxfiller.gui.mapping_state.pairing_preview` 를 읽기 전용으로 돌리고, 모델이
    서 있으면 그 모델의 실제 수치를 낸다. 라벨이 갈리는 근거가 이 한 축이다.
    """
    ctrl, _ = _controller_lib(tmp_path, paths=[TPL_COMPILED])
    ctrl.dispatch("use_library_template", {"path": str(TPL_COMPILED)})
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")
    pairing = ctrl.snapshot()["pairing"]
    assert pairing["ready"] is True and pairing["basis"] == "preview"
    assert pairing["field_count"] == 10 and pairing["column_count"] == 3
    # 낙찰현황 열 셋은 이 템플릿 필드와 하나도 겹치지 않는다 — 제안 0, 나머지가 확인 필요.
    assert (pairing["auto_count"], pairing["confirm_count"]) == (0, 10)
    assert ctrl.model is None, "1단계가 매핑 모델을 만들었습니다(재생성 전이가 열립니다)."

    # 2단계를 다녀오면 모델이 선다 — 그때부터 카드는 **실제 확정 수**를 말한다.
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": True})
    ctrl.dispatch("goto_section", {"section": "template"})
    pairing = ctrl.snapshot()["pairing"]
    assert pairing["basis"] == "model"
    assert (pairing["auto_count"], pairing["confirm_count"]) == (1, 9)


def test_editor_picker_reflects_shared_vm_refresh_without_stale_cache(tmp_path):
    """#137·#138 리뷰 F8 — 공유 VM refresh 가 좌 열에 즉시 반영된다(별도 행 캐시 0).

    U6-B 이후 그 반영은 `tpl` 채널 push 하나가 진다: 편집기는 목록을 성형하지 않고 같은
    VM 을 겨눌 뿐이라, 여기서 재는 것은 **주입이 단일 실체인가**다.
    """
    import shutil

    lib = tmp_path / "lib"
    lib.mkdir()
    vm = TemplateManagerViewModel(
        library_dir=lib,
        inspect_template=inspect_hwpx_template,
        file_ops=HWPX_TEMPLATE_OPS,
    )  # 빈 라이브러리로 시작
    ctrl = EditorController(
        JobRegistry(tmp_path / "jobs"), lambda s, snap: None, template_library=vm,
        clock=_clock,
        text_registry=TextTemplateRegistry(tmp_path / "text_templates"),
    )
    assert _lib_rows(ctrl) == []
    shutil.copy2(TPL_COMPILED, lib / "새서식.hwpx")  # 관리 화면 가져오기 시뮬레이션
    vm.refresh()
    assert "새서식" in {row.name for row in _lib_rows(ctrl)}


def test_use_library_template_rejects_paths_outside_library(tmp_path):
    """라이브러리 밖 경로는 loud 거부(백엔드 화이트리스트) — 웹이 임의 경로를 실어도 생
    파일 직접 로드 경로가 부활하지 않는다(2부: 가져오기=복사가 유일한 바깥 입구)."""
    ctrl, _ = _controller_lib(tmp_path, paths=[TPL_PARTIAL])
    with pytest.raises(ValueError, match="라이브러리에 없는"):
        ctrl.dispatch("use_library_template", {"path": str(TPL_COMPILED)})


def test_import_unification_copies_via_tpl_authority_and_adopts(tmp_path):
    """가져오기 통일(F8 — §10.17.2 판정 C): 복사 권위는 tpl 컨트롤러 하나(잠금·충돌 접미),
    편집기는 사본 채택만 판정한다. 유효 hwpx = 즉시 새 세션(F7 거동 보존), 충돌 접미도
    tpl 권위가 정한 정확한 목적지로 세션이 선다(반환이 이름이 아니라 경로인 이유)."""
    from hwpxfiller.external.text_registry import TextTemplateRegistry as TxtReg
    from hwpxfiller.webapp.screen_template import TemplateController

    from hwpxfiller.external.template_root import TemplateRoot

    lib = tmp_path / "lib"
    lib.mkdir()
    root = TemplateRoot(default_root=lib)      # U6-A: hwpx·txt 가 같은 서식 폴더
    txt_reg = TxtReg(root.path)
    tpl = TemplateController(
        txt_reg, lambda s, snap: None,
        file_store=TemplateFileStore(root.path, txt_reg),
        template_root=root,
        pool_registry=DatasetPoolRegistry(tmp_path / "datasets"),
    )
    ctrl = EditorController(
        JobRegistry(tmp_path / "jobs"), lambda s, snap: None,
        clock=_clock,
        template_library=tpl.vm,          # 앱 조립과 같은 단일 실체 공유
        text_registry=txt_reg,
    )
    dest = tpl.import_into_library(str(TPL_COMPILED))
    assert dest == str(lib / TPL_COMPILED.name)                        # 전체 경로 반환
    assert ctrl.adopt_imported_template(dest) == TPL_COMPILED.name
    assert ctrl.template_path == dest                                  # 세션 = 사본(원본 아님)
    dest2 = tpl.import_into_library(str(TPL_COMPILED))                 # 같은 이름 재가져오기
    assert dest2 != dest and Path(dest2).exists()                      # 접미 회피(조용한 덮기 금지)
    ctrl.adopt_imported_template(dest2)
    assert ctrl.template_path == dest2                                 # 접미 목적지 그대로 채택


def test_adopt_defers_raw_and_broken_copies_with_repair_notice(tmp_path):
    """§10.17.2 판정 C — 시작 불가 사본(RAW·손상·비 UTF-8 TXT)은 **세션 없이 목록 합류** +
    notice 가 수선 경로(행 ⋮ 변환·삭제)를 지목한다. 종전 선거부(무잔재)의 근거는 행 ⋮
    삭제가 서면서 소멸 — 사본은 남고, 지울 수 있고, 세션은 서지 않는다."""
    from test_webapp_template import _write_raw

    lib = tmp_path / "lib"
    lib.mkdir()
    raw = _write_raw(lib / "원본서식.hwpx")
    junk = lib / "깨진.hwpx"
    junk.write_bytes(b"this is not a hwpx zip")
    bad_txt = lib / "잘못.txt"
    bad_txt.write_bytes(b"\xff\xfe\x00 invalid utf8 \x80")
    ctrl, _ = _controller_lib(tmp_path, lib_dir=lib)
    for dest, needle in ((raw, "누름틀로 변환"), (junk, "읽을 수 없습니다"),
                         (bad_txt, "읽을 수 없습니다")):
        assert ctrl.adopt_imported_template(str(dest)) == dest.name
        assert ctrl.template_path == ""                                # 세션 없음
        snap = ctrl.snapshot()
        assert snap["notice"]["level"] == "warn"
        assert needle in snap["notice"]["text"]                        # 수선 경로 지목
        assert dest.exists()                                           # 목록 합류(잔재 아님 — 행)


def test_adopt_starts_txt_session_with_media_branch(tmp_path):
    """§10.17.2 판정 C — TXT 사본도 같은 채택 seam: 판독 가능하면 새 세션이 서고 매체
    분기(F6)가 그대로 탄다(파일 이름 탭 없음)."""
    lib = tmp_path / "lib"
    lib.mkdir()
    txt_dir = tmp_path / "text_templates"
    txt_dir.mkdir()
    doc = txt_dir / "협조전.txt"
    doc.write_text("수신: {{수신}}", encoding="utf-8")
    ctrl, _ = _controller_lib(tmp_path, lib_dir=lib)
    assert ctrl.adopt_imported_template(str(doc)) == "협조전.txt"
    assert ctrl.template_path == str(doc)
    assert ctrl.snapshot()["sections"] == ["template", "binding"]      # TXT 매체 탭(§3.2)


def test_pattern_preview_uses_real_renderer_on_save_stage(tmp_path):
    """F26 — 저장 분류의 파일명 라이브 예시는 실제 생성기(make_output_filename)와 같은
    함수로 만든 표본 1행(seq=1) 렌더다(예시 ≠ 산출물의 조용한 어긋남 금지)."""
    ctrl, _ = _controller(tmp_path)
    ctrl.load_template_path(str(TPL_COMPILED))
    _mount_data(ctrl)                                    # 1단계 게이트(U6-B)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "수기값"})
    field = ctrl.snapshot()["rows"][0]["template_field"]
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("goto_section", {"section": "filename"})
    ctrl.dispatch("set_pattern", {"pattern": f"x-{{{{{field}}}}}-{{{{seq:001}}}}"})
    assert ctrl.snapshot()["pattern_preview"] == "x-수기값-001.hwpx"
    ctrl.dispatch("goto_section", {"section": "binding"})
    assert ctrl.snapshot()["pattern_preview"] == ""                    # 저장 분류 밖은 미계산


# (구 test_import_template_rejects_broken_file_without_residue 삭제 — F8 판정 C: 선거부의
#  근거(인앱 삭제 어포던스 부재)가 행 ⋮ 삭제로 소멸. 손상 사본은 이제 삭제 가능한 오류
#  행으로 목록에 합류하고 세션만 서지 않는다 — 위 test_adopt_defers_raw_and_broken_copies.)


def test_use_library_rejection_refreshes_stale_list(tmp_path):
    """화이트리스트 거절은 갱신된 목록을 먼저 push 한다(PR-4 리뷰 F7) — 외부 삭제된 파일의
    stale 행이 화면에 남아 같은 클릭을 반복하게 만드는 무행동 안내 금지."""
    lib = tmp_path / "lib"
    lib.mkdir()
    ghost = lib / "유령.hwpx"
    ghost.write_bytes(TPL_COMPILED.read_bytes())
    ctrl, pushes = _controller_lib(tmp_path, lib_dir=lib)
    assert [row.name for row in _lib_rows(ctrl)] == ["유령"]
    ghost.unlink()                                                     # 외부 삭제
    before = len(pushes)
    with pytest.raises(ValueError, match="라이브러리에 없는"):
        ctrl.dispatch("use_library_template", {"path": str(ghost)})
    # 거절 **전에** 재스캔이 돌아 목록이 스스로 걷힌다(그 목록의 표면은 `tpl` 채널이다).
    assert _lib_rows(ctrl) == []
    assert len(pushes) > before


# ------------------------------------------------- 덮어쓰기 확인의 잠금·문안 대조(#149)
def test_overwrite_confirm_requires_the_text_the_user_actually_read(tmp_path):
    """확인 플래그만으로는 덮지 않는다 — **본 문안**을 함께 되돌려야 통과한다(#149).

    무엇을 보고 확정했는지 모르면 그 확인이 지금 상태에 대한 것인지 알 수 없다. 덮어쓰기는
    되돌릴 수 없으므로, 검증할 수 없는 확인은 통과가 아니라 재확인(fail-closed)이다.
    """
    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "작업일")
    _save_named(ctrl, "작업이")
    ctrl.load_job("작업일")
    ctrl.dispatch("set_name", {"name": "작업이"})
    res = ctrl.dispatch("save", {})
    assert res["needs_overwrite"] is True
    again = ctrl.dispatch("save", {"confirm_overwrite": True})       # 문안 없이 확정
    assert again["ok"] is False and again["needs_overwrite"] is True
    ok = ctrl.dispatch(
        "save",
        {"confirm_overwrite": True, "confirmed_overwrite_text": res["overwrite_text"]},
    )
    assert ok["ok"] is True


def test_overwrite_confirm_reasks_when_the_situation_changed_under_the_modal(tmp_path):
    """모달을 읽는 사이 상태가 바뀌면 새 문안으로 **다시 묻는다**(#149).

    사용자는 '작업이를 덮는다'를 확정했는데 그 사이 원본이 바뀌면, 확정은 더 이상 지금
    일어날 일에 대한 확인이 아니다 — 확인한 내용과 실제 집합이 갈라지는 자리.
    """
    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "원본작업")
    ctrl.load_job("원본작업")
    ctrl.dispatch("set_pattern", {"pattern": "새-{{수요기관}}"})
    assert ctrl.dispatch("save", {})["ok"] is True                   # 무드리프트 자기-갱신

    ctrl.load_job("원본작업")
    reg = JobRegistry(tmp_path / "jobs")
    job = reg.load("원본작업")
    job.filename_pattern = "외부-{{ID}}"                              # 편집 사이 외부 변경
    reg.save(job, allow_overwrite=True)
    res = ctrl.dispatch("save", {})
    assert res["needs_overwrite"] is True and "외부" in res["overwrite_text"]

    # 확정 왕복 중 원본이 통째로 사라진다 → 덮을 것이 없으니 그냥 저장(묻지 않는다).
    reg.path_for("원본작업").unlink()
    ok = ctrl.dispatch(
        "save",
        {"confirm_overwrite": True, "confirmed_overwrite_text": res["overwrite_text"]},
    )
    assert ok["ok"] is True


def test_overwrite_gate_is_judged_inside_the_write_lock(tmp_path):
    """게이트 판정이 **쓰기 잠금 안**이다(#149) — 판정과 실행 사이 창을 없앤다.

    잠금 밖 선판정은 판정 후 저장까지 사이에 디스크가 바뀔 수 있어, 확인 없이 외부 변경을
    덮거나 읽은 문안과 다른 자리를 덮는다. 판정 시점에 잠금이 다른 스레드에서 잡히지
    않는지로 되읽는다(저장 구간 잠금 테스트와 동형).
    """
    import threading

    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "게이트작업")
    ctrl.load_job("게이트작업")
    held: "list[bool]" = []
    real_gate = ctrl._overwrite_gate

    def spy() -> str:
        got = [None]

        def probe() -> None:
            lock = ctrl.registry.write_lock()
            got[0] = lock.acquire(blocking=False)
            if got[0]:
                lock.release()

        t = threading.Thread(target=probe)
        t.start()
        t.join(3)
        held.append(not got[0])
        return real_gate()

    ctrl._overwrite_gate = spy  # type: ignore[method-assign]
    assert ctrl.dispatch("save", {})["ok"] is True
    assert held and all(held), "덮어쓰기 게이트가 쓰기 잠금 밖입니다 — 판정·실행 창 회귀."


def test_edit_save_preserves_group_and_favorite(tmp_path):
    """편집 저장이 **이 화면이 편집하지 않는** durable 메타를 조용히 떨어뜨리지 않는다.

    태그·마지막 실행은 이미 보존됐지만 그룹과 즐겨찾기(슬라이스 2 신설)는 저장이 새로
    조립한 Job 에 실리지 않아 소실됐다 — 좌 목록 구획과 메인 후보 순위가 편집 한 번에
    조용히 초기화된다(confirm-or-alarm: 편집이 파괴한 것을 아무도 말하지 않는다).
    """
    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "메타보존작업")
    reg = JobRegistry(tmp_path / "jobs")
    reg.set_group("메타보존작업", "조달")
    reg.set_favorite("메타보존작업", True, "2026-07-26T09:00:00")

    ctrl.load_job("메타보존작업")
    assert ctrl.dispatch("save", {})["ok"] is True

    saved = reg.load("메타보존작업")
    assert saved.group == "조달"
    assert saved.favorited_at == "2026-07-26T09:00:00"


def test_overwriting_another_job_keeps_the_victims_ranking_meta(tmp_path):
    """이름을 바꿔 남의 자리를 덮어도 **그 작업의** 분류·이력·즐겨찾기는 그대로다(리뷰 3R P2).

    확인 문안이 약속한 것은 "그 파일을 덮어쓴다"뿐이다 — 원점 메타를 실으면 남의 즐겨찾기가
    조용히 꺼지거나 남의 카드에 내 실행 이력이 붙는다(문안 ≠ 실제 집합).
    """
    ctrl, _ = _controller26(tmp_path)
    reg = JobRegistry(tmp_path / "jobs")
    _save_named(ctrl, "덮힐작업")
    reg.set_group("덮힐작업", "대상그룹")
    reg.set_favorite("덮힐작업", True, "2026-07-01T09:00:00")
    reg.stamp_last_run("덮힐작업", "2026-07-02T09:00:00")

    _save_named(ctrl, "원점작업")                              # 두 번째 작업(편집 원점)
    reg.set_group("원점작업", "원점그룹")
    reg.set_favorite("원점작업", True, "2026-07-20T09:00:00")
    reg.stamp_last_run("원점작업", "2026-07-21T09:00:00")

    ctrl.load_job("원점작업")
    ctrl.dispatch("set_name", {"name": "덮힐작업"})            # 남의 자리로 이름 변경
    blocked = ctrl.dispatch("save", {})
    assert blocked["needs_overwrite"] is True
    res = ctrl.dispatch("save", {
        "confirm_overwrite": True,
        "confirmed_overwrite_text": blocked["overwrite_text"],
    })
    assert res["ok"] is True

    saved = reg.load("덮힐작업")
    assert saved.group == "대상그룹"                           # 남의 분류 불변
    assert saved.favorited_at == "2026-07-01T09:00:00"        # 남의 즐겨찾기 불변
    assert saved.last_run_at == "2026-07-02T09:00:00"         # 남의 이력 불변(이식 없음)


def test_saving_under_a_new_name_inherits_class_but_not_history(tmp_path):
    """빈 자리에 새 이름 = 새 identity — 그룹·태그는 따라가고 이력·즐겨찾기는 계승하지 않는다."""
    ctrl, _ = _controller26(tmp_path)
    reg = JobRegistry(tmp_path / "jobs")
    _save_named(ctrl, "원본")
    reg.set_group("원본", "조달")
    reg.set_favorite("원본", True, "2026-07-20T09:00:00")
    reg.stamp_last_run("원본", "2026-07-21T09:00:00")

    ctrl.load_job("원본")
    ctrl.dispatch("set_name", {"name": "새이름"})
    assert ctrl.dispatch("save", {})["ok"] is True

    fresh = reg.load("새이름")
    assert fresh.group == "조달"                               # 분류는 편집을 따라간다
    assert fresh.favorited_at == "" and fresh.last_run_at == ""  # 위조 금지
    assert reg.load("원본").favorited_at == "2026-07-20T09:00:00"  # 원본 불변


def _txt_draft_named(ctrl, tmp_path, name: str) -> dict:
    """TXT 초안을 이름까지 세워 저장 직전으로 만드는 최소 흐름 — 저장 결과를 돌려준다.

    데이터 결속은 **매체를 가리지 않는다**(#932 U4-C S2-3): TXT 작업도 레코드를
    읽어 기안을 세우므로(``workbench_entry_gate`` 서열 = 데이터→행→템플릿) 저장
    게이트가 같은 것을 요구한다. 매체로 갈리는 것은 파일명 패턴 축 하나다.
    """
    path = _txt_template(tmp_path)
    ctrl.dispatch("use_library_template", {"path": str(path)})
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "물품 구매"})
    blanks = ctrl.dispatch("confirm_all", {})["blanks"]
    ctrl.dispatch("confirm_blanks", {"fields": blanks})
    ctrl.dispatch("set_name", {"name": name})
    return ctrl.dispatch("save", {})


def test_saving_a_txt_draft_under_an_existing_hwpx_name_is_rejected(tmp_path):
    """TXT 초안을 기존 HWPX 작업 이름으로 저장 = 거절(§10.16 판정 D).

    F6 PR-B 가 편집기에 TXT 를 들이며 열린 경로다: 일반 덮어쓰기 확인만 거치면
    `_preserved_for_target` 이 victim 의 이력·즐겨찾기·검토 기준선을 보존해 이력 위조가
    된다(`last_run_at` 의 뜻은 매체가 정한다 — §19.4). 확인 승격이 아니라 **거절**이다 —
    `confirm_overwrite` 로도 뚫리지 않고 victim 은 디스크에서 불변이다.
    """
    ctrl, _ = _controller(tmp_path)
    assert _save_named(ctrl, "공고작업")["ok"] is True
    reg = JobRegistry(tmp_path / "jobs")
    reg.set_favorite("공고작업", True, "2026-07-01T09:00:00")
    reg.stamp_last_run("공고작업", "2026-07-02T09:00:00")

    res = _txt_draft_named(ctrl, tmp_path, "공고작업")
    assert res["ok"] is False
    assert "형식이 다른" in res["block_reason"] and "HWPX 문서 생성" in res["block_reason"]
    assert "needs_overwrite" not in res          # 거절될 저장에 확인 문안을 겹치지 않는다
    res = ctrl.dispatch("save", {"confirm_overwrite": True,
                                 "confirmed_overwrite_text": "아무 문안"})
    assert res["ok"] is False and "형식이 다른" in res["block_reason"]
    victim = reg.load("공고작업")
    assert victim.media == "hwpx"                # 자리 불변(위조 없음)
    assert victim.favorited_at == "2026-07-01T09:00:00"
    assert victim.last_run_at == "2026-07-02T09:00:00"


def test_saving_an_hwpx_draft_over_a_txt_job_name_is_rejected(tmp_path):
    """역방향 — HWPX 초안이 TXT 작업의 자리를 덮는 것도 같은 거절(§10.16 판정 D).

    이쪽은 PR-B 이전부터 이론상 열려 있던 방향이다(편집기는 hwpx 만 만들었다) — 표면에
    새 매체가 들어오면 기존 동사의 매체 가정을 함께 세어야 한다는 규칙의 대칭 가드.

    저장 뒤의 세션은 그 TXT 작업의 **편집 세션**이므로 새 초안은 실 UI 그대로 세션 시작
    (`new_session`)으로 연다 — 편집 세션 위에 다른 템플릿을 얹는 것은 어느 표면도 하지
    않는 조작이고(템플릿 선택 seam 은 언제나 세션을 먼저 끊는다), 그 상태로 재면 이
    테스트가 겨눈 저장 게이트가 아니라 탭 이동의 자동 버리기를 재게 된다.
    """
    ctrl, _ = _controller(tmp_path)
    assert _txt_draft_named(ctrl, tmp_path, "기안작업")["ok"] is True

    ctrl.dispatch("new_session", {})              # 실 UI: 새 작업은 세션을 먼저 끊는다
    res = _save_named(ctrl, "기안작업")          # hwpx 초안이 같은 이름을 겨눈다
    assert res["ok"] is False
    assert "형식이 다른" in res["block_reason"]
    assert "온나라 기안 검토·복사" in res["block_reason"]
    assert JobRegistry(tmp_path / "jobs").load("기안작업").media == "txt"  # 자리 불변


def test_saving_over_a_valid_unknown_media_job_is_rejected(tmp_path):
    """미상 매체(.docx) **정상** 작업의 자리도 덮어쓰기 거절(리뷰 4R P2).

    손상 victim(로드 실패)은 보존 메타가 빈 값으로 서서 위조가 없지만, 미상 매체 victim 은
    로드가 성공해 `_preserved_for_target` 이 이력·즐겨찾기를 그대로 보존한다 — 어느 매체의
    술어로도 읽을 수 없는 이력이 새 방식에 이식되는 같은 위조다. 그 작업의 정도는
    덮어쓰기가 아니라 relink 복구(미상→기지, §10.16 판정 C)다.
    """
    ctrl, _ = _controller(tmp_path)
    reg = JobRegistry(tmp_path / "jobs")
    docx = tmp_path / "구양식.docx"
    docx.write_text("x", encoding="utf-8")
    reg.save(Job(name="구양식작업", template_path=str(docx)))
    reg.stamp_last_run("구양식작업", "2026-07-02T09:00:00")

    res = _save_named(ctrl, "구양식작업")
    assert res["ok"] is False and "형식이 다른" in res["block_reason"]
    assert "지원 작업 방식 확인 필요" in res["block_reason"]  # 미상 라벨 fail-closed
    saved = reg.load("구양식작업")
    assert saved.media == "" and saved.last_run_at == "2026-07-02T09:00:00"  # 자리 불변


def test_edit_save_preserves_the_review_baseline(tmp_path):
    """3R P2 — 규칙을 하나도 안 바꾸고 저장만 해도 기준선이 비면 §13-2 의 조용한 반복이
    깨지고 다음 실행이 가장 무거운 검토를 다시 요구한다.

    에디터가 소유하는 것은 **규칙**(템플릿·매핑·파일명)이고 "마지막 완주가 그중 무엇을
    썼는가"는 실행 이력의 일이다 — 그룹·태그·이력과 같은 부류의 비-편집 메타다.
    """
    from hwpxfiller.domain.job import rules_fingerprints
    from hwpxfiller.gui.review_state import review_requirement

    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "기준선작업")
    reg = JobRegistry(tmp_path / "jobs")
    job = reg.load("기준선작업")
    job.last_run_at = "2026-07-01T09:00:00"
    job.reviewed_rules = rules_fingerprints(job)
    reg.save(job, allow_overwrite=True)
    assert not review_requirement(reg.load("기준선작업")).required

    ctrl.load_job("기준선작업")
    assert ctrl.dispatch("save", {})["ok"] is True
    after = reg.load("기준선작업")
    assert after.reviewed_rules == job.reviewed_rules, "저장이 검토 기준선을 지웠습니다."
    assert not review_requirement(after).required


def test_changing_the_rules_keeps_the_old_baseline_so_review_stands(tmp_path):
    """보존은 **무효화를 막지 않는다**: 기준선은 「마지막 완주가 쓴 것」 그대로 남고,
    바뀐 규칙과 어긋나 검토 요구가 선다(보존이 곧 승인은 아니다)."""
    from hwpxfiller.domain.job import rules_fingerprints
    from hwpxfiller.gui.review_state import review_requirement

    ctrl, _ = _controller26(tmp_path)
    _save_named(ctrl, "규칙변경작업")
    reg = JobRegistry(tmp_path / "jobs")
    job = reg.load("규칙변경작업")
    job.last_run_at = "2026-07-01T09:00:00"
    job.reviewed_rules = rules_fingerprints(job)
    reg.save(job, allow_overwrite=True)

    ctrl.load_job("규칙변경작업")
    ctrl.dispatch("set_pattern", {"pattern": "바뀐-{{seq:001}}"})
    assert ctrl.dispatch("save", {})["ok"] is True
    after = reg.load("규칙변경작업")
    assert after.reviewed_rules == job.reviewed_rules      # 기준선은 그대로
    assert review_requirement(after).required              # 규칙이 갈려 요구가 선다


# ------------------------------------------ TXT 매체 분기(F6 PR-B — 「기안」 생성 경로 승계)
def test_txt_template_loads_with_token_schema_and_two_tabs(tmp_path):
    """TXT 선택 → {{토큰}} 동형 스키마 + 탭 2개(템플릿·연결 — 파일 이름 탭 없음, §3.2).

    스키마가 :class:`TemplateSchema` 동형이라 `_ensure_model`·`validate_save` 의 스키마
    대조 술어가 hwpx 와 같은 길을 돈다(판정 단일 출처). PARTIAL 게이트는 서지 않는다.
    """
    ctrl, _ = _controller(tmp_path)
    path = _txt_template(tmp_path)  # 건명(2회)·금액
    ctrl.dispatch("use_library_template", {"path": str(path)})
    snap = ctrl.snapshot()
    assert snap["sections"] == ["template", "binding"]
    assert snap["template_media"] == "txt"
    assert snap["field_count"] == 2 and snap["gate"] is None and not snap["gate_error"]
    by_name = {f["name"]: f for f in snap["fields"]}
    assert by_name["건명"]["occurrences"] == 2          # 등장 횟수는 세그먼트 단일 출처로 센다
    assert by_name["금액"]["inferred_type"] == "amount"  # 이름 휴리스틱(infer_type) 공유
    assert ctrl.can_advance("template") is False        # 데이터가 남았다(U6-B)
    _mount_data(ctrl)
    assert ctrl.can_advance("template") is True
    with pytest.raises(ValueError, match="'filename' 탭이 없습니다"):
        ctrl.dispatch("goto_section", {"section": "filename"})


def test_txt_band_lists_templates_and_reads_errors_loud(tmp_path):
    """스냅샷 TXT 밴드 — 목록·필드 수·current 표지, 손상 파일은 오류 행으로 loud."""
    ctrl, _ = _controller(tmp_path)
    path = _txt_template(tmp_path)
    (tmp_path / "text_templates" / "손상.txt").write_bytes("한글".encode("cp949"))
    # 목록 성형은 U6-B 에서 `tpl` 채널 하나로 모였다(`test_webapp_template` 소관) —
    # 여기서는 편집기가 **같은 레지스트리를 겨눈다**는 것과 선택 표지를 잰다.
    by_name = {t.name: t for t in _lib_rows(ctrl, "txt")}
    assert len(by_name["기안"].fields()) == 2
    with pytest.raises(UnicodeDecodeError):
        by_name["손상"].fields()                        # 판독 실패는 숨기지 않는다
    ctrl.dispatch("use_library_template", {"path": str(path)})
    assert ctrl.snapshot()["template_path"] == str(path)


def test_txt_path_outside_registry_is_rejected(tmp_path):
    """레지스트리 밖 .txt 는 loud 거부 + 최신 목록 선 push(hwpx 화이트리스트와 같은 규율)."""
    ctrl, pushes = _controller(tmp_path)
    outside = tmp_path / "바깥.txt"
    outside.write_text("{{건명}}", encoding="utf-8")
    before = len(pushes)
    with pytest.raises(ValueError, match="라이브러리에 없는 템플릿"):
        ctrl.dispatch("use_library_template", {"path": str(outside)})
    assert len(pushes) > before  # 거절 전에 갱신 목록을 먼저 보여준다


def test_txt_template_without_tokens_blocks(tmp_path):
    """토큰 0 TXT = hwpx RAW 동형 차단 — 매체에 맞는 문안(누름틀·fieldize 언급 금지)."""
    ctrl, _ = _controller(tmp_path)
    path = _txt_template(tmp_path, name="맹탕", body="토큰이 하나도 없는 본문")
    ctrl.dispatch("use_library_template", {"path": str(path)})
    snap = ctrl.snapshot()
    assert "{{토큰}}" in snap["raw_block"] and "누름틀" not in snap["raw_block"]
    assert snap["fields"] == [] and ctrl.can_advance("template") is False


def test_txt_template_non_utf8_read_is_loud(tmp_path):
    """비 UTF-8 TXT 는 조용한 빈 스키마가 아니라 loud raise(confirm-or-alarm)."""
    ctrl, _ = _controller(tmp_path)
    bad = tmp_path / "text_templates"
    bad.mkdir(parents=True, exist_ok=True)
    p = bad / "구형.txt"
    p.write_bytes("건명: {{건명}}".encode("cp949"))
    with pytest.raises(ValueError, match="TXT 템플릿을 읽을 수 없습니다"):
        ctrl.dispatch("use_library_template", {"path": str(p)})


def test_txt_draft_saves_without_pattern_gate_and_reopens_with_two_tabs(tmp_path):
    """TXT 초안 저장 — 파일명 패턴 게이트가 서지 않고(매체 인지, §10.15.15 판정), 저장
    Job 형상은 구 「기안」 저장과 같다(pattern 미편집 = 기본값 그대로). 재편집 왕복도
    같은 2탭 구성으로 돌아온다(사망 점검표 5행)."""
    from hwpxfiller.domain.job import DEFAULT_FILENAME_PATTERN, template_media

    ctrl, _ = _controller(tmp_path)
    path = _txt_template(tmp_path)
    ctrl.dispatch("use_library_template", {"path": str(path)})
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "물품 구매"})
    blanks = ctrl.dispatch("confirm_all", {})["blanks"]
    ctrl.dispatch("confirm_blanks", {"fields": blanks})
    ctrl.dispatch("set_name", {"name": "TXT기안작업"})
    assert ctrl.dispatch("save", {})["ok"] is True
    # 착지 = 제자리(U2 §2.14) — txt 는 binding 이 마지막 탭이라 저장 자리가 곧 착지다.
    assert ctrl.snapshot()["section"] == "binding"

    saved = JobRegistry(tmp_path / "jobs").load("TXT기안작업")
    assert template_media(saved.template_path) == "txt"
    assert saved.filename_pattern == DEFAULT_FILENAME_PATTERN  # 미편집 기본값(기안 저장 동형)

    ctrl.load_job("TXT기안작업")
    snap = ctrl.snapshot()
    assert snap["sections"] == ["template", "binding"]
    assert snap["section"] == "binding" and snap["editing_origin"] == "TXT기안작업"


def test_load_job_with_target_lands_on_the_target_section_and_roundtrips(tmp_path):
    """target 이 서면 착지 탭도 target 이 정한다(값의 앞 절 = section) + 스냅샷 왕복.

    겨눈다고 말하고 다른 탭에 내리는 반쪽 착지 금지. 행 단위 조준(스크롤·포커스)은 뷰
    소관이고 스키마 드리프트는 뷰가 fail-open 한다 — 여기는 탭·문맥 축만 고정한다.
    """
    ctrl, _ = _controller26(tmp_path)
    assert _save_named(ctrl, "겨눔작업")["ok"] is True

    # 진입 사유·복귀 표면 표본은 `preview_result`/`preview` 였다 — #957 슬라이스 ③ 에서
    # 그 어휘가 걷혔으므로 살아 있는 결과 표면으로 겨눈다(축은 그대로: 착지 탭과 문맥 왕복).
    ctrl.load_job(
        "겨눔작업", entry_reason="run_failure",
        return_context={"surface": "result", "reopen_drawer": True, "preview_index": 2},
        target="filename/filenamePattern",
    )
    snap = ctrl.snapshot()
    assert snap["section"] == "filename"                     # landing_section 인자 없이도
    assert snap["context"]["target"] == "filename/filenamePattern"
    assert snap["context"]["return_context"]["preview_index"] == 2

    ctrl.load_job("겨눔작업", entry_reason="run_failure",
                  return_context={"surface": "result"}, target="binding/공고명")
    assert ctrl.snapshot()["section"] == "binding"

    import pytest
    with pytest.raises(ValueError, match="deep-link target"):
        ctrl.load_job("겨눔작업", target="template/x")       # fail-closed 관통(make_context)


def test_binding_commit_failure_reports_partial_success_without_rollback(tmp_path) -> None:
    calls: list[str] = []

    def fail_after_save(work_ref: str) -> None:
        calls.append(work_ref)
        raise RuntimeError("binding commit failed")

    ctrl, _ = _controller(tmp_path, after_mapping_saved=fail_after_save)

    assert _save_named(ctrl, "\ubd80\ubd84\uc131\uacf5")["ok"] is True
    ctrl.load_job(
        "\ubd80\ubd84\uc131\uacf5",
        entry_reason="run_failure",   # #957: `preview_result` 진입 사유는 어휘에서 걷혔다
        return_context={"surface": "data"},
        target="binding/\uacf5\uace0\uba85",
    )
    result = ctrl.dispatch("save", {})

    assert result["ok"] is False
    assert result["legacy_saved"] is True
    assert result["binding_commit_ok"] is False
    assert calls == ["\ubd80\ubd84\uc131\uacf5"]
    saved = JobRegistry(tmp_path / "jobs").load("\ubd80\ubd84\uc131\uacf5")
    assert saved.mapping.mappings
    assert ctrl.snapshot()["notice"]["level"] == "danger"
    assert "Field Binding" in result["block_reason"]


# ── 연결 확정 대기(#911) — 무장 사유를 **더한다**(dirty 술어는 무회귀) ──────────────────────
def test_binding_confirm_pending_is_absent_for_an_ordinary_job(tmp_path) -> None:
    """비관리 작업에선 확정 대기가 거짓이다 — 확정할 것이 없는 자리에 동사를 세우지 않는다.

    미주입 조립(테스트 단독·비관리 앱)도 같은 갈래다: 없는 표면을 있다고 말하지 않는다.
    """
    ctrl, _ = _controller(tmp_path)
    assert _save_named(ctrl, "평범작업")["ok"] is True
    ctrl.load_job("평범작업")

    snap = ctrl.snapshot()
    assert snap["binding_confirm"]["pending"] is False
    # 라벨·설명은 대기 여부와 무관하게 늘 실린다(표면이 문안을 발명하지 않는다).
    assert snap["binding_confirm"]["label"] and snap["binding_confirm"]["hint"]


def test_binding_confirm_pending_survives_a_clean_reentry_and_clears_on_confirm(
    tmp_path,
) -> None:
    """확정 대기 사실은 **손댄 것이 없는** 재진입에서도 참이고, 무변경 확정으로 걷힌다.

    이 두 값이 #911 의 결함과 수리다: 종전에는 dirty 하나가 두 동사를 잠가, 매핑이 이미
    옳은 작업은 확정을 요구받으면서 그 확정을 수행할 활성 동사가 없었다.
    """
    name = "확정대기"
    probe_calls: list[str] = []
    committed: list[str] = []
    pending = {"value": True}

    def probe(work_ref: str) -> bool:
        probe_calls.append(work_ref)
        return pending["value"]

    def commit(work_ref: str) -> None:
        committed.append(work_ref)
        pending["value"] = False       # 확정이 성립하면 대기는 사라진다(실 서비스의 전이).

    ctrl, _ = _controller(
        tmp_path, after_mapping_saved=commit, binding_confirm_pending=probe
    )
    assert _save_named(ctrl, name)["ok"] is True
    registry = JobRegistry(tmp_path / "jobs")
    job = registry.load(name)
    job.authority_id = "managed-work-911"
    registry.save(job, allow_overwrite=True)

    ctrl.load_job(name)
    snap = ctrl.snapshot()
    assert snap["dirty"] is False, "손대지 않은 재진입이라 변경 기반 무장은 닫혀 있다"
    assert snap["binding_confirm"]["pending"] is True

    result = ctrl.dispatch("save", {})    # 무변경 확정 — payload 는 평소 저장과 같다

    assert result["ok"] is True
    assert committed == [name], "무변경 저장도 결속 확정을 부른다"
    assert ctrl.snapshot()["binding_confirm"]["pending"] is False, (
        "확정이 성립했으면 확정 동사는 스스로 걷힌다"
    )
    assert probe_calls, "판정은 백엔드에 물어본다(표면 추론 금지)"


def test_binding_confirm_pending_stays_true_when_the_commit_fails(tmp_path) -> None:
    """확정이 실패하면 대기는 그대로다 — 실패를 성공처럼 접어 동사를 걷지 않는다."""
    name = "확정실패"

    def fail(work_ref: str) -> None:
        raise RuntimeError("binding commit failed")

    ctrl, _ = _controller(
        tmp_path, after_mapping_saved=fail, binding_confirm_pending=lambda _ref: True
    )
    assert _save_named(ctrl, name)["ok"] is True
    registry = JobRegistry(tmp_path / "jobs")
    job = registry.load(name)
    job.authority_id = "managed-work-911"
    registry.save(job, allow_overwrite=True)
    ctrl.load_job(name)

    result = ctrl.dispatch("save", {})

    assert result["ok"] is False and result["legacy_saved"] is True
    assert ctrl.snapshot()["binding_confirm"]["pending"] is True


def test_binding_confirm_probe_failure_never_arms_the_verb(tmp_path) -> None:
    """판정이 터지면 확정 대기는 거짓이다 — 눌러도 아무 일 없는 동사를 세우지 않는다."""
    def boom(_work_ref: str) -> bool:
        raise RuntimeError("authority store unreadable")

    ctrl, _ = _controller(tmp_path, binding_confirm_pending=boom)
    assert _save_named(ctrl, "판정불가")["ok"] is True
    ctrl.load_job("판정불가")

    assert ctrl.snapshot()["binding_confirm"]["pending"] is False


def test_ordinary_managed_mapping_save_runs_binding_sync(tmp_path) -> None:
    ctrl, _ = _controller(tmp_path)
    name = "\uad00\ub9ac\uc791\uc5c5"
    assert _save_named(ctrl, name)["ok"] is True
    registry = JobRegistry(tmp_path / "jobs")
    job = registry.load(name)
    job.authority_id = "managed-work-1"
    registry.save(job, allow_overwrite=True)
    calls: list[str] = []
    ctrl._after_mapping_saved = calls.append

    ctrl.load_job(name)
    result = ctrl.dispatch("save", {})

    assert result["ok"] is True
    assert calls == [name]


# ------------------------------- 계약 목록(pclm) 결속(#937)
#
# 편집기가 계약 목록을 **고르고 · 저장하고 · 다시 여는** 한 바퀴. 세 자리가 같은 성분 한
# 벌(db=경로 · 뷰=시트 · 0 · pclm)을 돌려야 저장본이 어느 어댑터로 읽힐지가 추측이 아니다.
_PCLM_VIEW = "v_통합_v1"


def _pclm_db(tmp_path: Path, *, name: str = "pclm.db", rows=None) -> str:
    """pclm 이 내는 모양을 흉내 낸 SQLite — 표 하나 위에 계약면 뷰를 얹는다."""
    import sqlite3

    db = tmp_path / name
    connection = sqlite3.connect(db)
    connection.execute('CREATE TABLE 계약 ("계약건명" TEXT, "계약금액" TEXT);')
    connection.executemany(
        "INSERT INTO 계약 VALUES (?, ?);",
        rows if rows is not None else [("잔류항생제분석기", "170,309,180")],
    )
    connection.execute(f'CREATE VIEW "{_PCLM_VIEW}" AS SELECT * FROM 계약;')
    connection.commit()
    connection.close()
    return str(db)


def _pool_editor(tmp_path: Path):
    """풀 seam 을 주입한 편집기 — 「고정한 데이터에서 고르기」가 서는 조립."""
    pool = DatasetPoolRegistry(tmp_path / "datasets")
    pushes: list = []
    ctrl = EditorController(
        JobRegistry(tmp_path / "jobs"),
        lambda s, snap: pushes.append((s, snap)),
        clock=_clock,
        pool_registry=pool,
        template_library=TemplateManagerViewModel(
            paths=[], inspect_template=inspect_hwpx_template, file_ops=HWPX_TEMPLATE_OPS,
        ),
        text_registry=TextTemplateRegistry(tmp_path / "text_templates"),
    )
    return ctrl, pool


def test_pool_select_block_reason_admits_pclm_and_speaks_its_own_broken_line(tmp_path):
    """「이 데이터를 쓸 수 있는가」의 판정 자리는 **`pool` 스냅샷 하나**다(U6-B #976).

    종전에는 셋이 각자 답했다: 데이터 선택 다이얼로그의 웹 함수(`usableReason`), 편집기
    축약 목록의 `screen_editor.pool_option_block`, 그리고 마운트 관문. 앞의 둘이 한
    컴포넌트가 되면서 그 어긋남이 화면 안에서 드러나므로 판정을 스냅샷 행으로 올렸다.

    **끊김 처방은 종류가 가른다**(#937): 엑셀에는 「다시 연결」이 있고 계약 목록에는 없다.
    """
    from hwpxfiller.application.dataset_pool import DatasetPoolRow
    from hwpxfiller.webapp.screen_pool import select_block_reason

    db = tmp_path / "pclm.db"
    db.write_bytes(b"x")
    live = DatasetPoolRow.from_item(
        "k1", DatasetReference(name="계약목록", kind="pclm", opts={"db": str(db), "view": "v"})
    )
    assert select_block_reason(live) == ""

    gone = DatasetPoolRow.from_item(
        "k2",
        DatasetReference(
            name="사라진목록", kind="pclm",
            opts={"db": str(tmp_path / "none.db"), "view": "v"},
        ),
    )
    reason = select_block_reason(gone)
    assert "참조가 끊겼습니다" in reason
    assert "계약 목록 DB 파일" in reason      # 엑셀 전용 동사(「다시 연결」)를 지시하지 않는다
    assert "다시 연결" not in reason

    frozen = DatasetPoolRow.from_item(
        "k3", DatasetReference(name="나라", kind="nara", opts={})
    )
    assert "작업 데이터로 연결할 수 없습니다" in select_block_reason(frozen)


def test_use_pool_data_mounts_a_pclm_view_and_the_save_carries_the_binding(tmp_path):
    """고르기 → 마운트 → 저장 한 바퀴 — 저장본이 db·뷰·0·pclm 을 그대로 든다."""
    ctrl, pool = _pool_editor(tmp_path)
    db = _pclm_db(tmp_path)
    key = pool.add(
        DatasetReference(name="계약목록", kind="pclm", opts={"db": db, "view": _PCLM_VIEW}),
    )
    ctrl.load_template_path(str(TPL_COMPILED))

    assert ctrl.dispatch("use_pool_data", {"key": key}) == {"ok": True, "label": "계약목록"}

    assert (ctrl.data_path, ctrl.data_sheet, ctrl.data_header_row, ctrl.data_kind) == (
        db, _PCLM_VIEW, 0, "pclm",
    )
    assert ctrl.source_fields == ["계약건명", "계약금액"]  # 컬럼이 곧 어휘(엑셀 헤더 동형)
    # 겨눈 슬롯을 스냅샷이 든다(U6-B #976) — 우 열의 `aria-current` 는 이 값이고, 표면이
    # 경로를 대조해 되추측하면 kind-스코프 정체성 규칙(#347)이 두 곳에 산다.
    snap = ctrl.snapshot()
    assert snap["data_pool_key"] == key
    assert snap["pairing"]["ready"] is True
    assert snap["pairing"]["data_name"] == Path(db).name
    assert snap["pairing"]["column_count"] == 2

    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("set_name", {"name": "계약작업"})
    assert ctrl.dispatch("save", {})["ok"] is True

    job = JobRegistry(tmp_path / "jobs").load("계약작업")
    assert (job.data_path, job.data_sheet, job.data_header_row, job.data_kind) == (
        db, _PCLM_VIEW, 0, "pclm",
    )


def test_reopening_a_pclm_bound_job_restores_the_view(tmp_path):
    """저장본을 다시 열면 결속이 **데이터를 먼저** 세운다 — 종류가 해석기를 가른다."""
    ctrl, pool = _pool_editor(tmp_path)
    db = _pclm_db(tmp_path)
    key = pool.add(
        DatasetReference(name="계약목록", kind="pclm", opts={"db": db, "view": _PCLM_VIEW}),
    )
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.dispatch("use_pool_data", {"key": key})
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("set_name", {"name": "계약작업"})
    ctrl.dispatch("save", {})

    fresh, _ = _pool_editor(tmp_path)
    fresh.load_job("계약작업")

    assert fresh.data_kind == "pclm" and fresh.data_sheet == _PCLM_VIEW
    assert fresh.source_fields == ["계약건명", "계약금액"]
    assert fresh.records[0]["계약건명"] == "잔류항생제분석기"


def test_load_source_ref_still_refuses_a_kind_it_cannot_read(tmp_path):
    """이름 없는 종류는 그대로 시끄럽게 거절한다 — 파일 갈래로 흘려보내지 않는다."""
    ctrl, _ = _controller(tmp_path)

    with pytest.raises(ValueError, match="복원할 수 없습니다"):
        ctrl._load_source_ref({"path": "C:/d/x.bin", "sheet": "", "kind": "미래소스"})


def test_whole_session_discard_returns_to_the_saved_pclm_binding(tmp_path):
    """세션 전체 버리기는 저장본의 **계약 목록 결속**으로 되돌아간다(빈 값이 아니다)."""
    ctrl, pool = _pool_editor(tmp_path)
    db = _pclm_db(tmp_path)
    key = pool.add(
        DatasetReference(name="계약목록", kind="pclm", opts={"db": db, "view": _PCLM_VIEW}),
    )
    ctrl.load_template_path(str(TPL_COMPILED))
    ctrl.dispatch("use_pool_data", {"key": key})
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_type", {"index": 0, "type": "const"})
    ctrl.dispatch("set_const", {"index": 0, "const": "v"})
    r = ctrl.dispatch("confirm_all", {})
    ctrl.dispatch("confirm_blanks", {"fields": r["blanks"]})
    ctrl.dispatch("set_name", {"name": "계약작업"})
    ctrl.dispatch("save", {})
    ctrl.load_job("계약작업")
    ctrl.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")  # 엑셀로 갈아탄 뒤

    ctrl.dispatch("discard_patch", {})  # section 없음 = 세션 전체

    assert (ctrl.data_path, ctrl.data_sheet, ctrl.data_kind) == (db, _PCLM_VIEW, "pclm")
    assert "연결된 것으로 되돌렸습니다" in ctrl.notice_text


def test_new_work_handoff_carries_the_pclm_kind_across_the_two_screens(tmp_path):
    """「이 데이터로 새 작업」의 참조는 종류를 관통한다 — 받는 쪽이 경로로 되추측하지 않는다."""
    db = _pclm_db(tmp_path)

    fresh, _ = _pool_editor(tmp_path)
    fresh.new_draft_with_data(
        {"path": db, "sheet": _PCLM_VIEW, "header_row": 0, "kind": "pclm"},
        entry_reason="document_browser_new_work",
    )

    assert fresh.data_kind == "pclm" and fresh.data_path == db
    assert fresh.source_fields == ["계약건명", "계약금액"]


# ------------------------------- U6-B 리뷰 반영(#976 리뷰 1·3·7)
def test_reselecting_the_same_template_is_a_no_op(tmp_path):
    """이미 고른 템플릿을 다시 고르면 **세션을 끊지 않는다**(리뷰 1).

    종전 표면에서는 현재 항목이 클릭 핸들러 없는 span 이라 이 호출이 구조적으로
    불가능했다. 고르기 화면은 현재 항목도 누를 수 있고 끌어 놓을 수도 있어, 통과시키면
    :meth:`new_job_session` 이 이름·매핑·단계를 통째로 끊는다 — 누른 사람은 「이미 고른
    것을 다시 골랐을」 뿐이다. 판정은 **표면과 여기 둘 다**에 선다: 표면만 막으면
    프로브·다른 호출자가 그대로 뚫는다.
    """
    ctrl, _ = _controller_lib(tmp_path, paths=[TPL_COMPILED])
    ctrl.dispatch("use_library_template", {"path": str(TPL_COMPILED)})
    _mount_data(ctrl)
    ctrl.dispatch("goto_section", {"section": "binding"})
    ctrl.dispatch("set_confirmed", {"index": 0, "confirmed": True})
    ctrl.dispatch("set_name", {"name": "지켜야 할 이름"})

    ctrl.dispatch("use_library_template", {"path": str(TPL_COMPILED)})   # 같은 템플릿 재선택

    snap = ctrl.snapshot()
    assert snap["section"] == "binding"                    # 단계가 1단계로 되감기지 않는다
    assert snap["name"] == "지켜야 할 이름"
    assert snap["data_path"] == str(MULTI_SHEET)
    assert snap["rows"][0]["confirmed"] is True            # 확정이 살아 있다
    # 대조군 — 다른 템플릿이면 종전대로 새 세션이다.
    ctrl2, _ = _controller_lib(tmp_path, paths=[TPL_COMPILED, TPL_PARTIAL])
    ctrl2.dispatch("use_library_template", {"path": str(TPL_COMPILED)})
    ctrl2.dispatch("set_name", {"name": "끊길 이름"})
    ctrl2.dispatch("use_library_template", {"path": str(TPL_PARTIAL)})
    assert ctrl2.snapshot()["name"] == ""


def test_pairing_is_not_ready_when_the_template_has_no_fields(tmp_path):
    """경로가 둘 다 있어도 **채울 필드가 0 이면 짝이 선 것이 아니다**(리뷰 3).

    종전에는 경로 유무만 봐서 「필드 0개 · 자동 연결 0」 카드가 비활성 CTA 위에 섰다 —
    화면이 「짝이 섰다」고 말하면서 다음으로 못 가는 자리다. 사유는 링1 문안 하나가 낸다.
    """
    ctrl, _ = _controller(tmp_path)
    txt = _txt_template(tmp_path, "빈템플릿", "토큰이 하나도 없는 본문")
    ctrl.dispatch("use_library_template", {"path": str(txt)})
    _mount_data(ctrl)

    pairing = ctrl.snapshot()["pairing"]
    assert pairing["template_name"] and pairing["data_name"]   # 둘 다 골랐다는 사실은 그대로
    assert pairing["ready"] is False, "채울 필드가 0 인데 짝이 섰다고 말합니다"
    assert pairing["field_count"] == 0
    assert (pairing["auto_count"], pairing["confirm_count"], pairing["basis"]) == (0, 0, "")
    from hwpxfiller.webapp.screens import TXT_RAW_BLOCK
    assert pairing["advance_block_reason"] == TXT_RAW_BLOCK   # 링1 문안 그대로
    assert ctrl.can_advance("template") is False


def test_pairing_counts_are_computed_only_on_the_choosing_stage(tmp_path, monkeypatch):
    """수치는 **고르기 단계에서만** 세고 같은 짝에서는 한 번만 센다(리뷰 7).

    ``suggest_mappings`` 는 필드×열 SequenceMatcher 라 매핑 편집의 잦은 push 마다 지불할
    것이 아니다. 세지 않은 자리는 ``basis=""`` 로 **세지 않았음을 명시**한다 — 0 을 사실처럼
    말하지 않는다.
    """
    calls: list = []
    import hwpxfiller.webapp.screen_editor as mod

    real = mod.pairing_preview

    def counted(fields, sources):
        calls.append((tuple(fields), tuple(sources)))
        return real(fields, sources)

    monkeypatch.setattr(mod, "pairing_preview", counted)

    ctrl, _ = _controller_lib(tmp_path, paths=[TPL_COMPILED])
    ctrl.dispatch("use_library_template", {"path": str(TPL_COMPILED)})
    _mount_data(ctrl)
    assert ctrl.snapshot()["pairing"]["basis"] == "preview"
    assert len(calls) == 1
    ctrl.snapshot()
    ctrl.snapshot()
    assert len(calls) == 1, f"같은 짝에서 다시 셌습니다: {calls!r}"

    # 2단계에서는 아예 세지 않는다 — 그리고 그 사실을 `basis` 가 말한다.
    ctrl.dispatch("goto_section", {"section": "binding"})
    binding = ctrl.snapshot()["pairing"]
    assert binding["basis"] == "" and binding["ready"] is True
    assert (binding["auto_count"], binding["confirm_count"]) == (0, 0)
    assert len(calls) == 1

    # 데이터를 갈아타면 정체 키가 움직여 다시 센다(캐시가 옛 수치를 말하지 않는다).
    # 모델이 서기 전 갈래를 재야 하므로 2단계를 다녀오지 않은 세션으로 본다 — 모델이
    # 있으면 수치의 출처가 `model` 로 갈리고(그쪽은 memo 대상이 아니다) 이 축이 안 보인다.
    fresh, _ = _controller_lib(tmp_path / "fresh", paths=[TPL_COMPILED])
    fresh.dispatch("use_library_template", {"path": str(TPL_COMPILED)})
    fresh.load_data_path(str(MULTI_SHEET), sheet="낙찰현황")   # 마운트 push 가 한 번 센다
    settled = len(calls)
    assert fresh.snapshot()["pairing"]["basis"] == "preview"
    assert len(calls) == settled, "같은 짝에서 다시 셌습니다"

    fresh.load_data_path(str(MULTI_SHEET))            # 다른 시트 = 다른 정체 키
    counts = fresh.snapshot()["pairing"]
    assert counts["basis"] == "preview" and counts["column_count"] == 2
    assert len(calls) > settled, "짝이 바뀌었는데 옛 수치를 재사용했습니다"
