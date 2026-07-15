"""CLI 테스트 — 기본 채우기 흐름 + 데이터 소스 선택(엑셀/나라장터) + 프로파일 매핑.

나라장터는 네트워크 없이: ``NaraStdDataSource._fetch`` 를 픽스처 바이트로 몽키패치.
"""

from __future__ import annotations

import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from hwpxfiller.cli import main
from hwpxfiller.core.engine import HwpxEngine
from hwpxfiller.core.mapping import FieldMapping, MappingProfile

CORPUS = Path(__file__).parent / "corpus" / "real"
FIXTURES = Path(__file__).parent / "fixtures"
TEMPLATE = str(CORPUS / "bid_notice_limited_under100m.hwpx")
FIELDS = ["입찰공고번호", "공고명", "계약방법", "추정가격", "개찰일시"]


def _xlsx(path: Path, rows: "list[list[str]]") -> str:
    wb = Workbook()
    ws = wb.active
    ws.append(FIELDS)
    for r in rows:
        ws.append(r)
    wb.save(path)
    return str(path)


def _outputs(d: Path) -> "list[str]":
    return sorted(p.name for p in d.glob("*.hwpx"))


def _covered_profile(*mappings: FieldMapping, name: str = "p") -> MappingProfile:
    """실 코퍼스 전 필드를 값 매핑 또는 명시 blank로 전건 확정한 프로파일."""
    covered = {m.template_field for m in mappings}
    blanks = [
        FieldMapping(field, type="blank")
        for field in HwpxEngine().required_fields(TEMPLATE)
        if field not in covered
    ]
    return MappingProfile(name=name, mappings=[*mappings, *blanks])


# --------------------------------------------------------------------- --fields
def test_fields_lists_required_fields(capsys):
    rc = main(["--template", TEMPLATE, "--fields"])
    assert rc == 0
    out = capsys.readouterr().out
    for f in FIELDS:
        assert f in out


# ---------------------------------------------------------------- 엑셀 채우기
def test_excel_fill_generates_documents(tmp_path):
    data = _xlsx(tmp_path / "d.xlsx",
                 [["R26BK00000001", "관급자재 구매", "일반경쟁", "12000000", "2026-08-01 10:00"]])
    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--data", data, "--out", str(out),
               "--pattern", "공고-{{입찰공고번호}}"])
    assert rc == 0
    assert _outputs(out) == ["공고-R26BK00000001.hwpx"]
    with zipfile.ZipFile(out / "공고-R26BK00000001.hwpx") as zf:
        assert zf.infolist()[0].filename == "mimetype"


def test_excel_missing_data_errors():
    # --data 없이 --source excel(기본) → argparse error → SystemExit.
    with pytest.raises(SystemExit):
        main(["--template", TEMPLATE])


# ------------------------------------------------- 다중 시트 loud 게이트(#2)
def _xlsx_multi(path: Path, sheets: "list[str]") -> str:
    """여러 시트 워크북 — 각 시트에 FIELDS 헤더 + 유효 1행."""
    wb = Workbook()
    wb.remove(wb.active)
    for name in sheets:
        ws = wb.create_sheet(name)
        ws.append(FIELDS)
        ws.append(["R26BK00000001", "관급자재 구매", "일반경쟁", "12000000", "2026-08-01 10:00"])
    wb.save(path)
    return str(path)


def test_cli_fill_multisheet_without_sheet_errors(tmp_path, capsys):
    """다중 시트인데 --sheet 미지정 → loud 실패(시트 나열), 산출물 미생성(#2)."""
    data = _xlsx_multi(tmp_path / "m.xlsx", ["공고목록", "낙찰현황"])
    out = tmp_path / "out"
    with pytest.raises(SystemExit):
        main(["--template", TEMPLATE, "--data", data, "--out", str(out),
              "--pattern", "공고-{{입찰공고번호}}"])
    err = capsys.readouterr().err
    assert "공고목록" in err and "낙찰현황" in err  # 어느 시트가 있는지 재진술
    assert not out.exists() or _outputs(out) == []  # 조용한 생성 없음


def test_cli_fill_multisheet_with_sheet_proceeds(tmp_path):
    """--sheet 명시 → 게이트 통과, 해당 시트로 생성."""
    data = _xlsx_multi(tmp_path / "m.xlsx", ["공고목록", "낙찰현황"])
    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--data", data, "--out", str(out),
               "--sheet", "낙찰현황", "--pattern", "공고-{{입찰공고번호}}"])
    assert rc == 0
    assert _outputs(out) == ["공고-R26BK00000001.hwpx"]


def test_cli_render_multisheet_without_sheet_errors(tmp_path, capsys):
    """render 경로도 동일 게이트 — 다중 시트 --sheet 미지정 시 loud 실패."""
    data = _xlsx_multi(tmp_path / "m.xlsx", ["공고목록", "낙찰현황"])
    tpl = tmp_path / "t.txt"
    tpl.write_text("공고: {{공고명}}", encoding="utf-8")
    with pytest.raises(SystemExit):
        main(["render", str(tpl), "--data", data])
    err = capsys.readouterr().err
    assert "공고목록" in err and "낙찰현황" in err


def test_cli_singlesheet_no_gate(tmp_path):
    """단일 시트는 --sheet 없이도 조용히 진행(회귀 방지) — 모호할 때만 묻는다."""
    data = _xlsx(tmp_path / "single.xlsx",
                 [["R26BK00000001", "관급자재 구매", "일반경쟁", "12000000", "2026-08-01 10:00"]])
    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--data", data, "--out", str(out),
               "--pattern", "공고-{{입찰공고번호}}"])
    assert rc == 0
    assert _outputs(out) == ["공고-R26BK00000001.hwpx"]


# ------------------------------------------------------- 프로파일 매핑(엑셀 소스)
def test_profile_maps_source_keys_to_template_fields(tmp_path):
    # 소스 헤더가 영문이라도 프로파일이 한글 필드로 잇는다.
    wb = Workbook()
    ws = wb.active
    ws.append(["bidNtceNo", "presmptPrce"])
    ws.append(["R26BK00000009", "5000000"])
    data = tmp_path / "eng.xlsx"
    wb.save(data)

    profile = _covered_profile(
        FieldMapping("입찰공고번호", "bidNtceNo"),
        FieldMapping("추정가격", "presmptPrce", type="amount"),
    )
    pf = tmp_path / "map.json"
    profile.save(pf)

    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--data", str(data), "--profile", str(pf),
               "--out", str(out), "--pattern", "p-{{입찰공고번호}}"])
    assert rc == 0
    gen = out / "p-R26BK00000009.hwpx"
    assert gen.exists()
    import zipfile as _z
    with _z.ZipFile(gen) as zf:
        blob = b"".join(zf.read(n) for n in zf.namelist() if n.endswith(".xml")).decode("utf-8")
    assert "5,000,000원" in blob


# ------------------------------------------------------------- 생성 원장(--ledger)
def test_ledger_is_optin_and_writes_evidence_sidecar(tmp_path, capsys):
    data = _xlsx(tmp_path / "d.xlsx",
                 [["R26BK00000001", "관급자재 구매", "일반경쟁", "12000000", "2026-08-01 10:00"]])

    # 기본은 opt-in — 플래그 없으면 사이드카를 만들지 않는다.
    out0 = tmp_path / "out0"
    assert main(["--template", TEMPLATE, "--data", data, "--out", str(out0),
                 "--pattern", "공고-{{입찰공고번호}}"]) == 0
    assert not list(out0.glob("fill-ledger*.json"))

    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--data", data, "--out", str(out),
               "--pattern", "공고-{{입찰공고번호}}", "--ledger"])
    assert rc == 0
    # 실행별 타임스탬프 파일명(RC-02) — 고정 이름이 아니라 fill-ledger-<시각>.json.
    (sidecar,) = list(out.glob("fill-ledger-*.json"))
    payload = json.loads(sidecar.read_text(encoding="utf-8"))
    assert payload["kind"] == "hwpx-fill-ledger"
    assert payload["source"] == f"file:{data}"           # 포인터-온리
    (entry,) = payload["outputs"]
    assert entry["ok"] and entry["verify_error"] == ""
    rows = {r["field"]: r for r in entry["rows"]}
    # dry-run 결과값 + 생성물 되읽기 증거(주장 아닌 관측).
    assert rows["입찰공고번호"]["preview_text"] == "R26BK00000001"
    assert rows["입찰공고번호"]["injected"] is True
    assert rows["공고명"]["injected"] is True
    # 소스 실제형 프로파일 — 잠정 라벨은 추정 표기.
    profs = {p["key"]: p for p in payload["profiles"]}
    assert profs["추정가격"]["samples"] == ["12000000"]
    assert profs["추정가격"]["tentative_type"] == "정수(추정)"
    assert "HWPX 렌더" in capsys.readouterr().err       # 미리보기 ≠ 렌더 고지


def test_ledger_rerun_accumulates_evidence(tmp_path):
    """원장 재실행 — 이전 실행의 증거 사이드카를 덮지 않고 축적한다(RC-02)."""
    data = _xlsx(tmp_path / "d.xlsx",
                 [["R26BK00000001", "관급자재 구매", "일반경쟁", "12000000", "2026-08-01 10:00"]])
    out = tmp_path / "out"
    args = ["--template", TEMPLATE, "--data", data, "--out", str(out),
            "--pattern", "공고-{{입찰공고번호}}", "--ledger"]
    assert main(args) == 0
    assert main(args + ["--overwrite"]) == 0  # 산출물 재생성은 명시 확정, 원장은 새 파일
    assert len(list(out.glob("fill-ledger-*.json"))) == 2


# ------------------------------------------------------- 덮어쓰기 계약(RC-02)
def test_cli_blocks_overwrite_by_default(tmp_path, capsys):
    """같은 폴더 재실행 기본 차단 — exit 1 + 기존 파일(수기 보정본) 무손상 + 안내."""
    data = _xlsx(tmp_path / "d.xlsx",
                 [["R26BK00000001", "관급자재 구매", "일반경쟁", "12000000", "2026-08-01 10:00"]])
    out = tmp_path / "out"
    args = ["--template", TEMPLATE, "--data", data, "--out", str(out),
            "--pattern", "공고-{{입찰공고번호}}"]
    assert main(args) == 0
    sentinel = out / "공고-R26BK00000001.hwpx"
    sentinel.write_bytes(b"user-edited")  # 발송 후 수기 보정본 모사

    rc = main(args)
    assert rc == 1
    assert sentinel.read_bytes() == b"user-edited"  # 무경고 파괴 금지
    err = capsys.readouterr().err
    assert "덮어쓰" in err and "--overwrite" in err  # 시끄럽게 + 다음 행동 안내


def test_cli_overwrite_optin_passes(tmp_path):
    """--overwrite 옵트인 — 명시 확정 시에만 기존 산출물을 교체한다."""
    data = _xlsx(tmp_path / "d.xlsx",
                 [["R26BK00000001", "관급자재 구매", "일반경쟁", "12000000", "2026-08-01 10:00"]])
    out = tmp_path / "out"
    args = ["--template", TEMPLATE, "--data", data, "--out", str(out),
            "--pattern", "공고-{{입찰공고번호}}"]
    assert main(args) == 0
    sentinel = out / "공고-R26BK00000001.hwpx"
    sentinel.write_bytes(b"user-edited")

    assert main(args + ["--overwrite"]) == 0
    assert sentinel.read_bytes()[:2] == b"PK"  # 재생성본으로 교체됨


# ------------------------------------------------------- 빈값 게이트(RC-03, ADR-E)
def test_cli_blocks_empty_values_by_default(tmp_path, capsys):
    """값이 빈 필드는 기본 차단(exit 1) — GUI 와 동일 입력에서 문서 내용이 갈라지지 않는다."""
    data = _xlsx(tmp_path / "d.xlsx",
                 [["", "관급자재 구매", "일반경쟁", "12000000", "2026-08-01 10:00"]])
    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--data", data, "--out", str(out),
               "--pattern", "공고-{{공고명}}"])
    assert rc == 1
    assert not out.exists()  # 생성 전 차단(부분 산출물 0)
    err = capsys.readouterr().err
    assert "입찰공고번호" in err and "--ack-empty" in err  # 시끄럽게 + 다음 행동 안내


def test_cli_ack_empty_injects_marker(tmp_path):
    """--ack-empty 옵트인 — GUI 와 같은 미입력 표식을 넣고 진행(조용한 누름틀 잔존 금지)."""
    from hwpxfiller.core.fields import read_fields

    data = _xlsx(tmp_path / "d.xlsx",
                 [["", "관급자재 구매", "일반경쟁", "12000000", "2026-08-01 10:00"]])
    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--data", data, "--out", str(out),
               "--pattern", "공고-{{공고명}}", "--ack-empty"])
    assert rc == 0
    (doc,) = list(out.glob("*.hwpx"))
    fields = read_fields(str(doc))
    assert fields["입찰공고번호"] == "〘미입력·입찰공고번호〙"  # GUI 와 동일 표식(단일 출처)
    assert fields["공고명"] == "관급자재 구매"


def test_cli_ack_empty_ledger_records_marker_evidence(tmp_path):
    """--ack-empty + --ledger — 원장이 문서 실상(표식 주입)을 증거한다(RC-03 원장 갈림 봉합)."""
    import json as _json

    data = _xlsx(tmp_path / "d.xlsx",
                 [["", "관급자재 구매", "일반경쟁", "12000000", "2026-08-01 10:00"]])
    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--data", data, "--out", str(out),
               "--pattern", "공고-{{공고명}}", "--ack-empty", "--ledger"])
    assert rc == 0
    (sidecar,) = list(out.glob("fill-ledger-*.json"))
    payload = _json.loads(sidecar.read_text(encoding="utf-8"))
    rows = {r["field"]: r for r in payload["outputs"][0]["rows"]}
    # 표식 주입 필드: 미충족(missing)으로 분류 + 실제 들어간 값(표식) 되읽기 증거.
    assert rows["입찰공고번호"]["status"] == "missing"
    assert rows["입찰공고번호"]["preview_text"] == "〘미입력·입찰공고번호〙"
    assert rows["입찰공고번호"]["injected"] is True


# --------------------------------------------------------------- 나라장터 소스
def _patch_nara(monkeypatch):
    """NaraStdDataSource._fetch 를 실 응답 픽스처로 대체(네트워크 차단).

    실 OS 자격증명 저장소 무접촉을 위해 ``DATA_GO_KR_KEY`` 환경변수도 비운다(키 해석이
    저장소로 폴백하지 않도록). 호출부는 빈 :class:`MemorySecretStore` 를 주입한다.
    """
    fixture = (FIXTURES / "nara_std_response.json").read_bytes()
    from hwpxfiller.data.nara import NaraStdDataSource
    monkeypatch.setattr(NaraStdDataSource, "_fetch", lambda self: fixture)
    monkeypatch.delenv("DATA_GO_KR_KEY", raising=False)


def test_nara_missing_service_key_errors(monkeypatch):
    from hwpxfiller.data.secret_store import MemorySecretStore
    _patch_nara(monkeypatch)
    # 인라인 키·파일·환경변수·저장소 모두 비어 있으면 시끄럽게 종료.
    with pytest.raises(SystemExit):
        main(["--template", TEMPLATE, "--source", "nara",
              "--bgn", "202606010000", "--end", "202606302359"],
             secret_store=MemorySecretStore())


def test_nara_source_with_profile_fills_template(tmp_path, monkeypatch, capsys):
    _patch_nara(monkeypatch)
    profile = _covered_profile(
        FieldMapping("입찰공고번호", "bidNtceNo"),
        FieldMapping("공고명", "bidNtceNm"),
        FieldMapping("추정가격", "presmptPrce", type="amount"),
        FieldMapping("개찰일시", "opengDate", type="date"),
        name="나라",
    )
    pf = tmp_path / "nara.json"
    profile.save(pf)
    out = tmp_path / "out"

    rc = main(["--template", TEMPLATE, "--source", "nara",
               "--service-key", "DUMMY", "--bgn", "202606010000", "--end", "202606302359",
               "--profile", str(pf), "--out", str(out), "--pattern", "n-{{입찰공고번호}}"])
    assert rc == 0
    # 픽스처는 2레코드 → 2파일.
    assert len(_outputs(out)) == 2
    gen = out / "n-R26BK01561738.hwpx"
    assert gen.exists()
    with zipfile.ZipFile(gen) as zf:
        blob = b"".join(zf.read(n) for n in zf.namelist() if n.endswith(".xml")).decode("utf-8")
    assert "65,454,545원" in blob
    assert "2026년 6월 15일" in blob
    # 취득 로그가 stderr 로 나온다.
    assert "[나라장터]" in capsys.readouterr().err


def test_nara_auth_failure_is_loud_exit_1(tmp_path, monkeypatch, capsys):
    """인증 실패(resultCode=07)는 '0건 취득 + exit 0' 조용한 성공이 아니라 exit 1(RC-03 C1)."""
    auth_fail = (
        b'{"response":{"header":{"resultCode":"07",'
        b'"resultMsg":"INVALID_REQUEST_PARAMETER_ERROR"},"body":{}}}'
    )
    from hwpxfiller.data.nara import NaraStdDataSource
    monkeypatch.setattr(NaraStdDataSource, "_fetch", lambda self: auth_fail)
    monkeypatch.delenv("DATA_GO_KR_KEY", raising=False)
    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--source", "nara",
               "--service-key", "DUMMY", "--bgn", "202606010000", "--end", "202606302359",
               "--out", str(out), "--pattern", "n-{{bidNtceNo}}", "--ledger"])
    assert rc == 1
    assert not out.exists()  # 문서도 원장도 없다 — 실패를 정상 실행으로 문서화하지 않는다
    err = capsys.readouterr().err
    assert "취득 실패" in err and "[07]" in err


def test_nara_period_over_one_month_blocked(tmp_path, monkeypatch, capsys):
    """기간 6개월은 데이터 경계 검증(RC-03)에 걸려 exit 1 — GUI 취득 경로와 동일 엄격도."""
    _patch_nara(monkeypatch)
    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--source", "nara",
               "--service-key", "DUMMY", "--bgn", "202601010000", "--end", "202607010000",
               "--out", str(out)])
    assert rc == 1
    assert "1개월" in capsys.readouterr().err
    assert not out.exists()


def test_profile_template_drift_is_cli_hard_gate(tmp_path, capsys):
    data = _xlsx(tmp_path / "d.xlsx",
                 [["1", "공고", "일반", "100", "2026-01-01 10:00"]])
    profile = MappingProfile(mappings=[FieldMapping("입찰공고번호", "입찰공고번호")])
    pf = tmp_path / "partial.json"
    profile.save(pf)
    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--data", data, "--profile", str(pf), "--out", str(out)])
    assert rc == 1 and not out.exists()
    assert "구조 드리프트" in capsys.readouterr().err


def test_nara_without_profile_warns(tmp_path, monkeypatch, capsys):
    _patch_nara(monkeypatch)
    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--source", "nara",
               "--service-key", "DUMMY", "--bgn", "202606010000", "--end", "202606302359",
               "--out", str(out), "--pattern", "x-{{bidNtceNo}}"])
    # 프로파일 없이도 실행은 되지만(영문키라 대부분 빈칸) 경고를 낸다.
    assert rc == 0
    assert "--profile 없이는" in capsys.readouterr().err


# ------------------------------------------------ 최상위 오류 번역 경계(RC-16)
def test_missing_template_is_one_line_korean_error_exit_2(tmp_path, capsys):
    """일상 실패(부재 파일)는 원시 traceback 대신 '[오류]' 1줄 + exit 2(게이트 1과 구분)."""
    rc = main(["--template", str(tmp_path / "없는템플릿.hwpx"), "--fields"])
    assert rc == 2
    err = capsys.readouterr().err
    assert "[오류]" in err
    assert "Traceback" not in err


def test_out_path_colliding_with_existing_file_exit_2(tmp_path, capsys):
    """--out 자리에 기존 파일(FX3b) → FileExistsError traceback 대신 번역 진단."""
    data = _xlsx(tmp_path / "d.xlsx",
                 [["1", "공고", "일반", "100", "2026-01-01 10:00"]])
    clash = tmp_path / "outfile"
    clash.write_text("x", encoding="utf-8")
    rc = main(["--template", TEMPLATE, "--data", data, "--out", str(clash),
               "--pattern", "공고-{{입찰공고번호}}"])
    assert rc == 2
    assert "[오류]" in capsys.readouterr().err


def test_lint_crash_exit_2_distinct_from_issue_gate_exit_1(tmp_path, capsys):
    """손상 템플릿 lint 크래시(2)가 '위생 이슈 게이트'(1)와 종료코드로 구분된다."""
    corrupt = tmp_path / "손상.hwpx"
    corrupt.write_text("zip 아님", encoding="utf-8")
    rc = main(["lint", str(corrupt)])
    assert rc == 2
    err = capsys.readouterr().err
    assert "[오류]" in err
    assert "Traceback" not in err


def test_ledger_failure_diagnosed_but_generation_exit_kept(tmp_path, capsys, monkeypatch):
    """원장 실패(FX6)는 '[원장 실패]' 진단만 — 성공 배치를 실패(exit 1)로 위장하지 않는다.

    사이드카는 실행별 타임스탬프 파일명(RC-02)이라 경로 선점으로는 실패를 주입할 수
    없다 — export 함수 자체에 실패를 주입한다.
    """
    import hwpxfiller.cli as cli_mod

    def boom(*args, **kwargs):
        raise OSError("사이드카 쓰기 실패 주입")

    monkeypatch.setattr(cli_mod, "_export_ledger", boom)
    data = _xlsx(tmp_path / "d.xlsx",
                 [["1", "공고", "일반", "100", "2026-01-01 10:00"]])
    out = tmp_path / "out"
    rc = main(["--template", TEMPLATE, "--data", data, "--out", str(out),
               "--pattern", "공고-{{입찰공고번호}}", "--ledger"])
    assert rc == 0                       # 생성 성패 기준 exit 유지
    assert "[원장 실패]" in capsys.readouterr().err
    assert _outputs(out)                 # 생성물은 실재


def test_nara_fetch_error_translated_exit_1(monkeypatch, capsys):
    """NaraFetchError 는 '[오류] 나라장터 취득 실패' 1줄 + exit 1(데이터 경계 게이트).

    취득 실패는 크래시(exit 2)가 아니라 RC-03 데이터 경계 게이트다 — 인증 실패·기간
    위반·연결 실패를 '0건 성공'으로 넘기지 않되, 자동화가 게이트로 분류하게 한다.
    """
    from hwpxfiller.data.nara import NaraFetchError, NaraStdDataSource

    def boom(self):
        raise NaraFetchError("connection timed out")

    monkeypatch.setattr(NaraStdDataSource, "_fetch", boom)
    monkeypatch.delenv("DATA_GO_KR_KEY", raising=False)
    rc = main(["--template", TEMPLATE, "--source", "nara", "--service-key", "DUMMY",
               "--bgn", "202606010000", "--end", "202606302359"])
    assert rc == 1
    err = capsys.readouterr().err
    assert "[오류] 나라장터 취득 실패" in err
    assert "Traceback" not in err


# ------------------------------------------------- 파일명 패턴 계약(RC-20)
def test_pattern_token_missing_from_data_fails_loud(tmp_path, monkeypatch, capsys):
    """기본 패턴의 {{ID}} 가 나라 레코드에 없음 → 'output-{{ID}}.hwpx' 조용 생성 대신 오류."""
    _patch_nara(monkeypatch)
    out = tmp_path / "out"
    with pytest.raises(SystemExit) as ei:
        main(["--template", TEMPLATE, "--source", "nara",
              "--service-key", "DUMMY", "--bgn", "202606010000", "--end", "202606302359",
              "--out", str(out)])
    assert ei.value.code == 2
    assert "파일명 패턴의 토큰이 데이터에 없어" in capsys.readouterr().err
    assert not out.exists()              # 미치환 이름의 파일이 만들어지지 않았다


def test_cli_pattern_default_is_single_source():
    from hwpxfiller.cli import DEFAULT_FILENAME_PATTERN as CLI_DEFAULT
    from hwpxfiller.core.job import DEFAULT_FILENAME_PATTERN
    assert CLI_DEFAULT is DEFAULT_FILENAME_PATTERN


# ------------------------------------------------- 최상위 --help 표면(RC-21)
def test_top_level_help_lists_subcommands(capsys):
    """--help 가 pre-argparse 수동 디스패치 하위명령 6종을 전부 표기한다."""
    with pytest.raises(SystemExit) as ei:
        main(["--help"])
    assert ei.value.code == 0
    out = capsys.readouterr().out
    for name in ("schema", "fieldize", "lint", "drift", "render", "diff"):
        assert name in out


# ------------------------------------------------- lint --vocab BOM(RC-33)
def test_lint_vocab_utf8_bom_does_not_pollute_first_entry(tmp_path, capsys):
    """메모장/PowerShell 이 남긴 BOM(U+FEFF)이 첫 어휘 항목을 오염시키지 않는다."""
    vocab = tmp_path / "vocab.txt"
    vocab.write_bytes(("﻿" + "입찰공고번호\n공고명\n").encode("utf-8"))  # BOM 파일
    rc = main(["lint", TEMPLATE, "--vocab", str(vocab)])
    out = capsys.readouterr().out
    assert "﻿" not in out           # BOM 잔존 없음
    assert "'입찰공고번호'" not in out   # 어휘에 있는 첫 항목이 오탐되지 않는다
    assert rc == 1                       # 다른 필드는 여전히 어휘 밖 — 게이트 계약 유지


# ------------------------------------------------- 키 소스 우선순위(스펙 고정)
def _key_args(**over):
    import argparse
    ns = argparse.Namespace(service_key_file=None, service_key=None)
    for k, v in over.items():
        setattr(ns, k, v)
    return ns


def test_service_key_precedence(tmp_path, monkeypatch):
    """우선순위: --service-key-file > DATA_GO_KR_KEY > --service-key(비권장) > 저장소.

    핵심: 1급 입력(파일·환경변수)이 비권장 인라인 플래그보다 위. 비권장 인라인은 수동
    저장된 키보다는 위(명시적 override).
    """
    import argparse

    from hwpxfiller.cli import _resolve_service_key
    from hwpxfiller.data.secret_store import NARA_SERVICE_KEY_NAME, MemorySecretStore

    ap = argparse.ArgumentParser()
    store = MemorySecretStore({NARA_SERVICE_KEY_NAME: "STORED"})
    kf = tmp_path / "key.txt"
    kf.write_text("  FILEKEY\n", encoding="utf-8")

    # 파일이 최우선(환경변수·인라인·저장소 모두 있어도).
    monkeypatch.setenv("DATA_GO_KR_KEY", "ENVKEY")
    args = _key_args(service_key_file=str(kf), service_key="INLINE")
    assert _resolve_service_key(ap, args, store) == "FILEKEY"

    # 파일 없으면 환경변수가 인라인보다 우선(스펙 핵심 수정).
    args = _key_args(service_key="INLINE")
    assert _resolve_service_key(ap, args, store) == "ENVKEY"

    # 환경변수 없으면 비권장 인라인이 저장소보다 우선.
    monkeypatch.delenv("DATA_GO_KR_KEY", raising=False)
    args = _key_args(service_key="INLINE")
    assert _resolve_service_key(ap, args, store) == "INLINE"

    # 아무 입력 없으면 저장소 폴백.
    args = _key_args()
    assert _resolve_service_key(ap, args, store) == "STORED"

    # 저장소도 비었으면 None.
    args = _key_args()
    assert _resolve_service_key(ap, args, MemorySecretStore()) is None


# ---------------------------------------------------- 전역 용어 정렬(RC-26, U12)
def test_cli_help_uses_canonical_terms(capsys):
    """CLI 도움말이 정준 용어를 노출한다 — --profile 재사용성('매핑 프로파일'),
    fieldize 대응 병기('누름틀 변환'), drift 판본쌍('구판/신판')."""
    with pytest.raises(SystemExit):
        main(["--template", TEMPLATE, "--help"])
    top = capsys.readouterr().out
    assert "매핑 프로파일" in top          # GUI 산출물 재사용 명기(1개념 1이름)

    with pytest.raises(SystemExit):
        main(["fieldize", "--help"])
    fz = capsys.readouterr().out
    assert "누름틀 변환" in fz              # fieldize = 누름틀 변환(사용자 문구 대응)

    with pytest.raises(SystemExit):
        main(["drift", "--help"])
    dr = capsys.readouterr().out
    assert "구판" in dr and "신판" in dr    # 판본 쌍 정준(CLI가 GUI/HTML과 일치)
