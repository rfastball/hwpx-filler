"""Excel/CSV 데이터 소스와 공유 행 성형 계약.

첫 시트(또는 지정 시트)의 ``header_row``가 필드명이고 이후 각 행이 문서 1건이다.
빈 행은 건너뛰고 짧은 행은 빈 값으로 채운다. 빈/중복 헤더와 헤더보다 긴 비어 있지
않은 행은 조용한 데이터 소실을 막기 위해 거절한다. XLSX 수식은 계산하지 않고 저장된
cache만 읽으며 cache가 없으면 시끄럽게 실패한다.
"""

from __future__ import annotations

import csv
from collections.abc import Iterable
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


def _cell_text(value: object) -> str:
    """CSV와 Excel이 공유하는 결정론적 scalar→text 정책."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, float):
        return format(value, ".15g")
    return str(value)


def _normalize_headers(raw: "Iterable[object] | None") -> "list[str]":
    headers = [_cell_text(value).strip() for value in (raw or ())]
    blank = [str(index + 1) for index, header in enumerate(headers) if not header]
    if blank:
        raise ValueError(f"빈 헤더 열은 사용할 수 없습니다: {', '.join(blank)}")
    seen: set[str] = set()
    duplicate: list[str] = []
    for header in headers:
        if header in seen and header not in duplicate:
            duplicate.append(header)
        seen.add(header)
    if duplicate:
        raise ValueError(f"중복 헤더는 사용할 수 없습니다: {', '.join(duplicate)}")
    return headers


def sheet_overview(path: "str | Path") -> "list[tuple[str, int, int]]":
    """통합문서의 시트를 열거해 ``(시트명, 행수, 열수)`` 목록으로 준다.

    xlsx/xlsm 은 openpyxl read_only 메타(max_row/max_column — 저장 시점 dimension
    기반 **근사치**)로 전 시트를 통합문서 순서 그대로 열거한다. CSV 는 시트 개념이
    없으므로 **빈 목록** — 소비자(시트 선택 다이얼로그)의 생략 판정 단일 출처다
    (빈 목록 = 물을 것이 없음, 길이 1 = 유일 시트라 물을 필요 없음).
    """
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return []
    wb = load_workbook(str(path), read_only=True)
    try:
        return [
            (name, wb[name].max_row or 0, wb[name].max_column or 0)
            for name in wb.sheetnames
        ]
    finally:
        wb.close()


def ambiguous_sheets(path: "str | Path") -> "list[tuple[str, int, int]]":
    """시트 확정이 필요한(2+ 시트) 워크북이면 개요를, 아니면 빈 목록을 준다.

    "모호할 때만 묻는다"의 판정 **단일 출처** — :func:`sheet_overview` 위에 얇게 얹는다:
    CSV·단일 시트는 물을 것이 없어 빈 목록, 2+ 시트는 조용한 첫 시트 선택을 막고 사용자에게
    확정을 요구할 근거(시트 목록)를 돌려준다. CLI(``--sheet`` 게이트)와 웹(시트 선택
    다이얼로그, #33)이 같은 이 판정을 공유한다 — 두 표면의 판정 드리프트를 원천 차단.
    """
    overview = sheet_overview(path)
    return overview if len(overview) >= 2 else []


def ambiguous_sheet_error(path: "str | Path", *, prefix: str = "") -> "str | None":
    """#33 멀티시트 게이트의 공유 판정+문구 — 모호(2+ 시트)면 거절 문구, 아니면 ``None``.

    정책 3요소의 단일 출처(수동 등록=pool 화면·겨눔=``load_pool_item_checked`` 공유 —
    두 사이트에 복붙돼 문구가 표류하던 것을 여기로 수렴):

    - 판정은 :func:`ambiguous_sheets` (빈 목록=CSV·단일 시트 → 통과=``None``).
    - **읽기 실패(경로 부재·잠김·손상)는 통과(``None``)** — 참조 등록 의미(파일 미개봉)를
      지키고, 죽은 참조는 이어지는 실제 로드/겨눔 관문이 시끄럽게 재진술한다.
    - 2+ 시트면 시트 목록을 병기한 거절 문구를 돌려준다(첫 시트 자동 선택은 조용한
      오독 위험 — 시끄럽게 확정을 요구).

    ``prefix`` 는 호출 컨텍스트 병기용(예: ``"등록 데이터 'x' 에 시트가 지정되지 않았습니다 — "``).
    """
    try:
        overview = ambiguous_sheets(path)
    except Exception:  # noqa: BLE001 — 읽기 실패는 참조 의미상 통과(후속 관문이 재방어)
        return None
    if not overview:
        return None
    names = ", ".join(n for n, _r, _c in overview)
    return (
        f"{prefix}워크북에 시트가 여러 개입니다({names}). "
        "데이터 관리에서 시트를 지정해 등록하세요."
    )


class ExcelDataSource:
    def __init__(self, path: str, sheet: "str | None" = None, header_row: int = 1):
        if isinstance(header_row, bool) or not isinstance(header_row, int) or header_row < 1:
            raise ValueError("header_row는 1 이상의 정수여야 합니다.")
        self.path = path
        self.sheet = sheet
        self.header_row = header_row
        self._headers: "list[str]" = []
        self._records: "list[dict[str, str]]" = []
        self._loaded = False

    def _load(self) -> None:
        if self._loaded:
            return
        ext = Path(self.path).suffix.lower()
        if ext == ".csv":
            self._load_csv()
        else:
            self._load_xlsx()
        self._loaded = True

    def _load_xlsx(self) -> None:
        values_wb = load_workbook(self.path, data_only=True, read_only=True)
        formulas_wb = load_workbook(self.path, data_only=False, read_only=True)
        try:
            values_ws = (
                values_wb[self.sheet] if self.sheet else values_wb[values_wb.sheetnames[0]]
            )
            formulas_ws = (
                formulas_wb[self.sheet]
                if self.sheet
                else formulas_wb[formulas_wb.sheetnames[0]]
            )
            value_rows = list(values_ws.iter_rows(values_only=True))
            formula_rows = list(formulas_ws.iter_rows(values_only=True))
        finally:
            values_wb.close()
            formulas_wb.close()
        if len(value_rows) < self.header_row:
            return
        header_formulas = formula_rows[self.header_row - 1]
        if any(isinstance(value, str) and value.startswith("=") for value in header_formulas):
            raise ValueError("헤더에는 Excel 수식을 사용할 수 없습니다.")
        self._headers = _normalize_headers(value_rows[self.header_row - 1])
        for offset, raw in enumerate(value_rows[self.header_row:], start=self.header_row + 1):
            formulas = formula_rows[offset - 1] if offset - 1 < len(formula_rows) else ()
            resolved: list[object] = []
            width = max(len(raw), len(formulas))
            for column in range(width):
                value = raw[column] if column < len(raw) else None
                formula = formulas[column] if column < len(formulas) else None
                if isinstance(formula, str) and formula.startswith("=") and value is None:
                    raise ValueError(
                        f"Excel 수식 cache가 없습니다: 행 {offset}, 열 {column + 1}"
                    )
                resolved.append(value)
            self._append_record(resolved, row_number=offset)

    def _load_csv(self) -> None:
        with open(self.path, "r", encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
        if len(rows) < self.header_row:
            return
        self._headers = _normalize_headers(rows[self.header_row - 1])
        for offset, raw in enumerate(rows[self.header_row:], start=self.header_row + 1):
            self._append_record(raw, row_number=offset)

    def _append_record(self, raw, *, row_number: int) -> None:
        if raw is None:
            return
        cells = [_cell_text(value) for value in raw]
        if all(c.strip() == "" for c in cells):
            return  # 빈 행 스킵
        overflow = cells[len(self._headers):]
        if any(value.strip() for value in overflow):
            raise ValueError(
                f"헤더보다 값이 많은 행입니다: 행 {row_number}, "
                f"헤더 {len(self._headers)}열"
            )
        rec = {
            header: cells[index] if index < len(cells) else ""
            for index, header in enumerate(self._headers)
        }
        self._records.append(rec)

    # ---------------------------------------------------------- DataSource
    def records(self) -> "list[dict[str, str]]":
        self._load()
        return self._records

    def fields(self) -> "list[str]":
        self._load()
        return [h for h in self._headers if h]

    def field_labels(self) -> "dict[str, str]":
        """Excel/CSV 헤더는 이미 사람이 읽는 라벨이라 별도 어휘가 없다(빈 dict).

        영문 코드 키를 쓰는 소스(예: 나라장터 API)만 자기 어휘를 선언한다.
        """
        return {}
